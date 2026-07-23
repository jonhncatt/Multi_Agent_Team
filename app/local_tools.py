from __future__ import annotations

import json
import fnmatch
import hashlib
import itertools
import importlib
import os
import re
import secrets
import shlex
import shutil
import ssl
import subprocess
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from html import unescape
from pathlib import Path
from typing import Any, Callable

from PIL import Image, ImageEnhance, ImageOps

from app.action_validator import (
    blocked_supply_chain_command,
    is_dangerous_command,
    missing_supply_chain_allowed_commands,
    parse_compound_shell_command,
    shell_command_uses_compound_syntax,
    validate_compound_shell_command as validate_compound_shell_command_shared,
    validate_command_path_args,
)
from app.browser_runtime import BrowserToolManager
from app.config import AppConfig, get_access_roots, normalize_permission_profile
from app.context_meter import count_tokens, truncate_text_to_token_limit
from app.i18n import normalize_locale
from app.document_text import (
    extract_heading_entries_from_pages,
    extract_pdf_page_texts_from_path,
    extract_pdf_tables_from_path,
    extract_pdf_text_from_bytes,
    extract_pdf_text_from_path,
    normalize_lookup_text,
    truncate_text,
)
from app.sandbox import DockerSandboxManager
from app.storage import ProjectStore
from app.tool_trace_summary import safe_error_message
from app.tool_result_store import ToolResultStore

try:
    from pillow_heif import register_heif_opener

    register_heif_opener()
except Exception:
    pass


def _is_within(path: Path, root: Path) -> bool:
    return path == root or root in path.parents


_BROAD_GLOB_GUIDANCE_THRESHOLD = 300

APPLY_PATCH_TOOL_DESCRIPTION = (
    "Apply one atomic file-oriented patch under allowed writable roots. "
    "Use `*** Add File:` only for a target known not to exist, `*** Update File:` for every "
    "existing or previously read file, and `*** Delete File:` only for an existing file. "
    "Never use Add File to replace an existing file."
)

APPLY_PATCH_ARGUMENT_DESCRIPTION = (
    "Patch DSL enclosed by `*** Begin Patch` and `*** End Patch`. Each operation starts with "
    "`*** Add File: path`, `*** Update File: path`, or `*** Delete File: path`. Add File content "
    "uses `+` lines. Update File requires one or more `@@` hunks whose context/removal/addition "
    "lines begin with space, `-`, or `+`; it may be followed by `*** Move to: new_path`. Paths "
    "may be relative to cwd or absolute, but must remain under an allowed writable root."
)


def _project_relative_or_absolute(path: Path, project_root: Path) -> str:
    try:
        return str(path.resolve().relative_to(project_root.resolve())).replace("\\", "/")
    except Exception:
        return str(path.resolve())


def _display_model_path(path: Path, *, project_root: Path, cwd: Path | None = None) -> str:
    resolved = path.resolve()
    root = project_root.resolve()
    try:
        return str(resolved.relative_to(root)).replace("\\", "/")
    except Exception:
        pass
    if cwd is not None:
        try:
            return str(resolved.relative_to(cwd.resolve())).replace("\\", "/")
        except Exception:
            pass
    return str(resolved)


def _path_payload(path: Path, *, project_root: Path, cwd: Path | None = None) -> dict[str, Any]:
    resolved = path.resolve()
    model_path = _display_model_path(resolved, project_root=project_root, cwd=cwd)
    root_ref = "absolute"
    try:
        resolved.relative_to(project_root.resolve())
        root_ref = "project_root"
    except Exception:
        if cwd is not None:
            try:
                resolved.relative_to(cwd.resolve())
                root_ref = "cwd"
            except Exception:
                root_ref = "absolute"
    return {
        "path": model_path,
        "display_path": model_path,
        "root_ref": root_ref,
        "resolved_path": str(resolved),
    }


def _is_broad_glob_pattern(pattern: str) -> bool:
    normalized = str(pattern or "").strip().replace("\\", "/")
    return normalized in {"*", "*.*", "**", "**/*", "./**/*", "**/*.*"}


_UPLOAD_META_SAFE_PATTERN = re.compile(r"[^a-zA-Z0-9._-]+")


def _safe_upload_meta_name(name: str) -> str:
    return _UPLOAD_META_SAFE_PATTERN.sub("_", str(name or "")).strip("._") or "file"


def _meta_matches_upload_key(meta: dict[str, Any], raw: str, raw_basename: str) -> bool:
    candidate_path = Path(str(meta.get("path") or ""))
    keys = {
        str(meta.get("id") or "").strip(),
        str(meta.get("original_name") or meta.get("name") or "").strip(),
        str(meta.get("safe_name") or "").strip(),
        candidate_path.name if str(candidate_path) else "",
    }
    return raw in keys or (raw_basename and raw_basename in keys)


def _upload_meta_path_for(config: AppConfig, raw: str) -> Path:
    return config.uploads_dir / ".meta" / f"{_safe_upload_meta_name(raw)}.json"


def _build_path_candidates(
    config: AppConfig,
    raw_path: str,
    *,
    workspace_root: Path | None = None,
    access_roots: list[Path] | None = None,
) -> list[Path]:
    raw = (raw_path or ".").strip() or "."
    path = Path(raw).expanduser()
    seen: set[str] = set()
    candidates: list[Path] = []
    base_root = (workspace_root or config.workspace_root).resolve()
    roots = [root.resolve() for root in (access_roots or get_access_roots(config))]

    def add(p: Path) -> None:
        resolved = p.resolve()
        key = str(resolved)
        if key in seen:
            return
        seen.add(key)
        candidates.append(resolved)

    if path.is_absolute():
        add(path)
        return candidates

    normalized = raw.replace("\\", "/").strip("/").lower()
    normalized_slash = raw.replace("\\", "/").strip("/")
    if normalized:
        # High-priority alias mapping, e.g. "workbench/a.txt" -> "<allowed_root_named_workbench>/a.txt"
        # Also support short aliases from allowed root tails, e.g. "master/source" -> "<...>/master/source".
        for root in roots:
            root_norm = str(root).replace("\\", "/").rstrip("/").lower()
            if normalized == root_norm or normalized == root.name.lower():
                add(root)
                continue
            prefix = f"{root.name.lower()}/"
            if normalized.startswith(prefix):
                suffix = normalized_slash[len(prefix) :]
                add(root / suffix)

            parent_name = root.parent.name.lower()
            if parent_name:
                if normalized == parent_name:
                    add(root)
                parent_prefix = f"{parent_name}/"
                if normalized.startswith(parent_prefix):
                    suffix = normalized_slash[len(parent_prefix) :]
                    if suffix == root.name.lower():
                        add(root)
                    elif suffix.startswith(f"{root.name.lower()}/"):
                        add(root / suffix[len(root.name) + 1 :])
                    else:
                        add(root / suffix)

                parent_child = f"{parent_name}/{root.name.lower()}"
                if normalized == parent_child:
                    add(root)
                parent_child_prefix = f"{parent_child}/"
                if normalized.startswith(parent_child_prefix):
                    suffix = normalized_slash[len(parent_child_prefix) :]
                    add(root / suffix)

    # Default mapping keeps backward compatibility.
    add(base_root / path)
    for root in roots:
        if root == base_root:
            continue
        add(root / path)

    return candidates


def _resolve_workspace_path(
    config: AppConfig,
    raw_path: str,
    *,
    workspace_root: Path | None = None,
    access_roots: list[Path] | None = None,
    allow_any_path: bool = False,
) -> Path:
    base_root = (workspace_root or config.workspace_root).resolve()
    roots = [root.resolve() for root in (access_roots or get_access_roots(config))]
    if base_root not in roots:
        roots = [base_root, *roots]
    if allow_any_path:
        path = Path((raw_path or ".").strip() or ".").expanduser()
        if not path.is_absolute():
            path = base_root / path
        path = path.resolve()
        return path

    candidates = _build_path_candidates(config, raw_path, workspace_root=base_root, access_roots=roots)

    # Prefer existing paths in allowed roots for better UX with relative inputs.
    for path in candidates:
        for root in roots:
            if _is_within(path, root) and path.exists():
                return path

    # Fall back to first allowed candidate even if it does not exist,
    # prefer a candidate whose parent directory exists.
    for path in candidates:
        for root in roots:
            if _is_within(path, root) and path.parent.exists():
                return path

    # Last resort: return first allowed candidate even if parent does not exist,
    # so upper layers can return a clear "not found" error.
    for root in roots:
        for path in candidates:
            if _is_within(path, root):
                return path

    allowed = ", ".join(str(p) for p in roots)
    raise ValueError(f"Path out of allowed roots: {raw_path}. Allowed roots: {allowed}")


def _resolve_source_path(
    config: AppConfig,
    raw_path: str,
    *,
    workspace_root: Path | None = None,
    access_roots: list[Path] | None = None,
    allow_any_path: bool = False,
) -> Path:
    """
    Resolve existing source file path with upload-name fallback.
    If raw_path is only an original upload filename (e.g. a.zip),
    try matching uploads_dir entry like <uuid>__a.zip.
    """
    roots = [root.resolve() for root in (access_roots or get_access_roots(config))]
    resolved = _resolve_workspace_path(
        config,
        raw_path,
        workspace_root=workspace_root,
        access_roots=roots,
        allow_any_path=allow_any_path,
    )
    if resolved.exists():
        return resolved

    raw = (raw_path or "").strip()
    if not raw:
        return resolved

    try:
        raw_basename = Path(raw.replace("\\", "/")).name
        direct_meta_path = _upload_meta_path_for(config, raw)
        if direct_meta_path.exists():
            direct_meta = json.loads(direct_meta_path.read_text(encoding="utf-8"))
            if isinstance(direct_meta, dict):
                direct_path = Path(str(direct_meta.get("path") or "")).expanduser().resolve()
                for root in roots:
                    if direct_path.exists() and _is_within(direct_path, root):
                        return direct_path
        uploads_index_path = config.uploads_dir / "index.json"
        if uploads_index_path.exists():
            upload_index = json.loads(uploads_index_path.read_text(encoding="utf-8"))
            if isinstance(upload_index, dict):
                direct_hit = upload_index.get(raw)
                if isinstance(direct_hit, dict):
                    direct_path = Path(str(direct_hit.get("path") or "")).expanduser().resolve()
                    for root in roots:
                        if direct_path.exists() and _is_within(direct_path, root):
                            return direct_path
                for meta in upload_index.values():
                    if not isinstance(meta, dict):
                        continue
                    candidate_path = Path(str(meta.get("path") or "")).expanduser().resolve()
                    if _meta_matches_upload_key(meta, raw, raw_basename):
                        for root in roots:
                            if candidate_path.exists() and _is_within(candidate_path, root):
                                return candidate_path
        meta_dir = config.uploads_dir / ".meta"
        if meta_dir.exists():
            meta_files = sorted(meta_dir.glob("*.json"), key=lambda m: m.stat().st_mtime, reverse=True)[:500]
            for meta_file in meta_files:
                meta = json.loads(meta_file.read_text(encoding="utf-8"))
                if not isinstance(meta, dict) or not _meta_matches_upload_key(meta, raw, raw_basename):
                    continue
                candidate_path = Path(str(meta.get("path") or "")).expanduser().resolve()
                for root in roots:
                    if candidate_path.exists() and _is_within(candidate_path, root):
                        return candidate_path
    except Exception:
        pass

    p = Path(raw.replace("\\", "/"))
    if p.is_absolute():
        return resolved

    basename = p.name
    if not basename:
        return resolved

    try:
        matches = sorted(
            config.uploads_dir.glob(f"*__{basename}"),
            key=lambda m: m.stat().st_mtime,
            reverse=True,
        )
    except Exception:
        return resolved

    for match in matches:
        candidate = match.resolve()
        for root in roots:
            if _is_within(candidate, root):
                return candidate
    return resolved


def _truncate_output(text: str, max_chars: int = 12000) -> str:
    if len(text) <= max_chars:
        return text
    return f"{text[:max_chars]}\n\n[output truncated: {len(text)} chars]"


def _looks_like_html(content_type: str, text: str) -> bool:
    lower_ct = (content_type or "").lower()
    if "text/html" in lower_ct or "application/xhtml+xml" in lower_ct:
        return True
    head = text[:400].lower()
    return "<html" in head or "<!doctype html" in head


def _extract_html_text(raw_html: str, max_chars: int) -> str:
    html = re.sub(r"(?is)<!--.*?-->", " ", raw_html)
    html = re.sub(r"(?is)<(script|style|noscript).*?>.*?</\1>", " ", html)
    html = re.sub(r"(?i)<br\s*/?>", "\n", html)
    html = re.sub(r"(?i)</(p|div|li|tr|h1|h2|h3|h4|h5|h6|section|article)>", "\n", html)
    html = re.sub(r"(?s)<[^>]+>", " ", html)
    html = unescape(html)

    lines: list[str] = []
    for line in html.splitlines():
        normalized = re.sub(r"\s+", " ", line).strip()
        if normalized:
            lines.append(normalized)

    out = "\n".join(lines)
    if len(out) > max_chars:
        out = out[:max_chars]
    return out


def _find_html_meta_content(raw_html: str, attr_name: str, attr_value: str) -> str:
    pattern = re.compile(
        rf'(?is)<meta[^>]*{attr_name}\s*=\s*["\']{re.escape(attr_value)}["\'][^>]*content\s*=\s*["\'](.*?)["\']'
    )
    match = pattern.search(raw_html or "")
    if match:
        return _clean_html_fragment(match.group(1) or "")
    return ""


def _extract_html_metadata(raw_html: str, base_url: str = "") -> dict[str, str]:
    html = raw_html or ""
    title = ""
    title_match = re.search(r"(?is)<title[^>]*>(.*?)</title>", html)
    if title_match:
        title = _clean_html_fragment(title_match.group(1) or "")
    if not title:
        title = _find_html_meta_content(html, "property", "og:title") or _find_html_meta_content(html, "name", "twitter:title")

    published_at = (
        _find_html_meta_content(html, "property", "article:published_time")
        or _find_html_meta_content(html, "name", "article:published_time")
        or _find_html_meta_content(html, "property", "og:updated_time")
        or _find_html_meta_content(html, "name", "pubdate")
        or _find_html_meta_content(html, "name", "publish-date")
    )

    canonical_url = ""
    canonical_match = re.search(r'(?is)<link[^>]*rel\s*=\s*["\']canonical["\'][^>]*href\s*=\s*["\'](.*?)["\']', html)
    if canonical_match:
        canonical_url = unescape(canonical_match.group(1) or "").strip()
        if canonical_url and base_url:
            canonical_url = urllib.parse.urljoin(base_url, canonical_url)

    return {
        "title": title,
        "published_at": published_at,
        "canonical_url": canonical_url,
    }


def _tokenize_query(query: str) -> list[str]:
    text = (query or "").strip()
    if not text:
        return []

    ascii_stopwords = {
        "a",
        "an",
        "and",
        "are",
        "at",
        "for",
        "from",
        "how",
        "in",
        "is",
        "it",
        "latest",
        "mlb",
        "news",
        "npb",
        "of",
        "on",
        "recent",
        "score",
        "scores",
        "the",
        "today",
        "what",
        "when",
        "where",
        "who",
        "why",
    }
    cjk_fillers = (
        "查一下",
        "查下",
        "搜一下",
        "搜索",
        "帮我查",
        "请问",
        "最近",
        "近期",
        "最新",
        "新闻",
        "消息",
        "今天",
        "今日",
        "现在",
        "目前",
        "在不在",
        "在哪",
        "是否",
        "一下",
    )
    cjk_stopwords = {"新闻", "消息", "今天", "今日", "最近", "近期", "一下", "查下", "搜索"}

    seen: set[str] = set()
    tokens: list[str] = []

    def add(token: str) -> None:
        normalized = str(token or "").strip().lower()
        if not normalized or normalized in seen:
            return
        seen.add(normalized)
        tokens.append(normalized)

    for part in re.split(r"[^a-z0-9]+", text.lower()):
        if len(part) < 2 or part in ascii_stopwords:
            continue
        add(part)

    cjk_text = text
    for filler in cjk_fillers:
        cjk_text = cjk_text.replace(filler, " ")

    for segment in re.findall(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]+", cjk_text):
        cleaned = segment.strip()
        if len(cleaned) < 2:
            continue
        max_size = min(4, len(cleaned))
        for size in range(max_size, 1, -1):
            for start in range(0, len(cleaned) - size + 1):
                token = cleaned[start : start + size]
                if token in cjk_stopwords:
                    continue
                add(token)
                if len(tokens) >= 24:
                    return tokens

    return tokens


def _query_relevance_score(query: str, item: dict[str, Any]) -> float:
    title = str(item.get("title") or "").lower()
    snippet = str(item.get("snippet") or "").lower()
    domain = str(item.get("domain") or "").lower()
    tokens = _tokenize_query(query)
    score = 0.0
    for token in tokens:
        weight = 1.0 + min(len(token), 4) * 0.35
        if token in title:
            score += 4.0 * weight
        if token in snippet:
            score += 2.0 * weight
        if token.isascii() and token in domain:
            score += 1.5 * weight
    return score


def _score_web_result(query: str, item: dict[str, Any]) -> float:
    domain = str(item.get("domain") or "").lower()
    score = _query_relevance_score(query, item)
    if domain.endswith(".gov") or domain.endswith(".edu"):
        score += 2.5
    if any(flag in domain for flag in ("official", "docs", "developer", "openai.com", "github.com")):
        score += 1.5
    if item.get("published_at"):
        score += 0.5
    return score


def _looks_like_script_payload(text: str) -> bool:
    sample = (text or "")[:6000].lower()
    if not sample:
        return False

    if "sourcemappingurl=" in sample:
        return True

    markers = [
        "function(",
        "var ",
        "const ",
        "let ",
        "window.",
        "document.",
        "=>",
    ]
    hits = sum(1 for m in markers if m in sample)
    longest_line = max((len(line) for line in sample.splitlines()), default=0)
    punct = sum(ch in "{}[]();=<>/\\*" for ch in sample)
    alpha = sum(ch.isalpha() for ch in sample) or 1
    punct_ratio = punct / alpha

    return (hits >= 3 and longest_line >= 220) or punct_ratio >= 0.45


def _extract_search_query(url: str) -> str | None:
    try:
        parsed = urllib.parse.urlsplit(url)
    except Exception:
        return None

    host = (parsed.hostname or "").lower()
    if not host:
        return None

    q = urllib.parse.parse_qs(parsed.query or "")
    key = None
    if "google." in host or "bing." in host:
        key = "q"
    elif "yahoo." in host:
        key = "p"
    elif "baidu." in host:
        key = "wd"

    if not key:
        return None
    vals = q.get(key) or []
    if not vals:
        return None
    out = (vals[0] or "").strip()
    return out or None


def _clean_html_fragment(raw_html: str) -> str:
    text = re.sub(r"(?s)<[^>]+>", " ", raw_html or "")
    text = unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _decode_ddg_redirect(raw_url: str) -> str:
    if not raw_url:
        return raw_url
    url = unescape(raw_url).strip()
    absolute = urllib.parse.urljoin("https://duckduckgo.com", url)
    try:
        parsed = urllib.parse.urlsplit(absolute)
    except Exception:
        return absolute

    host = (parsed.hostname or "").lower()
    if host.endswith("duckduckgo.com") and parsed.path == "/l/":
        q = urllib.parse.parse_qs(parsed.query or "")
        target = (q.get("uddg") or [""])[0].strip()
        if target:
            return urllib.parse.unquote(target)
    return absolute


def _extract_ddg_results(raw_html: str, max_results: int) -> list[dict[str, str]]:
    html = raw_html or ""
    limit = max(1, min(20, int(max_results)))
    patterns = [
        re.compile(
            r'(?is)<a[^>]*class="[^"]*result__a[^"]*"[^>]*href="([^"]+)"[^>]*>(.*?)</a>'
        ),
        re.compile(
            r"(?is)<a[^>]*class=['\"][^'\"]*result-link[^'\"]*['\"][^>]*href=['\"]([^'\"]+)['\"][^>]*>(.*?)</a>"
        ),
    ]

    seen: set[str] = set()
    out: list[dict[str, str]] = []

    for pattern in patterns:
        for match in pattern.finditer(html):
            href = _decode_ddg_redirect(match.group(1) or "")
            title = _clean_html_fragment(match.group(2) or "")
            if not href or not title:
                continue
            try:
                parsed = urllib.parse.urlsplit(href)
            except Exception:
                continue
            if parsed.scheme not in {"http", "https"}:
                continue
            host = (parsed.hostname or "").lower()
            if host.endswith("duckduckgo.com") and parsed.path == "/y.js":
                continue

            key = f"{href}|{title}".lower()
            if key in seen:
                continue
            seen.add(key)

            snippet = ""
            window = html[match.end() : match.end() + 2400]
            snippet_match = re.search(
                r'(?is)<(?:a|div|span)[^>]*class="[^"]*result__snippet[^"]*"[^>]*>(.*?)</(?:a|div|span)>',
                window,
            )
            if not snippet_match:
                snippet_match = re.search(
                    r"(?is)<td[^>]*class=['\"][^'\"]*result-snippet[^'\"]*['\"][^>]*>(.*?)</td>",
                    window,
                )
            if snippet_match:
                snippet = _clean_html_fragment(snippet_match.group(1) or "")

            out.append({"title": title, "url": href, "snippet": snippet})
            if len(out) >= limit:
                return out

    return out


def _normalize_url_for_request(raw_url: str) -> str:
    """
    Make URL safe for urllib by encoding non-ASCII host/path/query.
    """
    url = (raw_url or "").strip()
    parsed = urllib.parse.urlsplit(url)

    scheme = (parsed.scheme or "").lower()
    if scheme not in {"http", "https"}:
        raise ValueError("Only http/https URLs are supported")
    if not parsed.netloc:
        raise ValueError("Invalid URL")

    host = parsed.hostname or ""
    if not host:
        raise ValueError("Invalid URL")
    host_ascii = host.encode("idna").decode("ascii")

    auth = ""
    if parsed.username is not None:
        auth = urllib.parse.quote(parsed.username, safe="")
        if parsed.password is not None:
            auth += ":" + urllib.parse.quote(parsed.password, safe="")
        auth += "@"

    port = f":{parsed.port}" if parsed.port else ""
    netloc = f"{auth}{host_ascii}{port}"

    path = urllib.parse.quote(urllib.parse.unquote(parsed.path or "/"), safe="/%:@!$&'()*+,;=-._~")
    query = urllib.parse.quote(urllib.parse.unquote(parsed.query or ""), safe="=&%:@!$'()*+,;/-._~")

    return urllib.parse.urlunsplit((scheme, netloc, path, query, ""))


def _is_cert_verify_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "certificate_verify_failed" in text or "certificate verify failed" in text


def _normalize_search_query(query: str) -> str:
    return re.sub(r"\s+", " ", (query or "").strip())


def _expand_search_variants(query: str) -> list[str]:
    normalized = _normalize_search_query(query)
    if not normalized:
        return []

    variants: list[str] = []
    seen: set[str] = set()

    def add(value: str) -> None:
        clean = _normalize_search_query(value)
        key = clean.lower()
        if not clean or key in seen:
            return
        seen.add(key)
        variants.append(clean)

    add(normalized)

    hex_patterns = list(re.finditer(r"(?i)\b(?:0x([0-9a-f]{1,4})|([0-9a-f]{1,4})h)\b", normalized))
    for match in hex_patterns:
        digits = (match.group(1) or match.group(2) or "").upper()
        if not digits:
            continue
        token_variants = [f"{digits}h", f"{digits} h", f"0x{digits}"]
        for token in token_variants:
            add(token)
            replaced = normalized[: match.start()] + token + normalized[match.end() :]
            add(replaced)

    return variants


def _build_search_pattern(query: str) -> re.Pattern[str] | None:
    normalized = _normalize_search_query(query)
    if not normalized:
        return None
    parts = [re.escape(part) for part in normalized.split(" ") if part]
    if not parts:
        return None
    body = r"\s+".join(parts)
    if len(parts) == 1 and re.fullmatch(r"(?i)(?:0x)?[0-9a-f]{1,4}h?", normalized):
        body = rf"(?<![0-9A-F]){body}(?![0-9A-F])"
    return re.compile(body, flags=re.IGNORECASE)


def _page_hint_for_offset(text: str, offset: int) -> int | None:
    page = None
    for match in re.finditer(r"--- Page (\d+) ---", text):
        if match.start() > offset:
            break
        try:
            page = int(match.group(1))
        except Exception:
            page = None
    return page


def _spans_overlap(left: tuple[int, int], right: tuple[int, int]) -> bool:
    return left[0] < right[1] and right[0] < left[1]


def _looks_like_pdf_path(path: Path) -> bool:
    if path.suffix.lower() == ".pdf":
        return True
    try:
        with path.open("rb") as fp:
            return fp.read(5).startswith(b"%PDF-")
    except Exception:
        return False


def _xlsx_cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        try:
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
        except Exception:
            pass
        return str(value)
    if hasattr(value, "isoformat"):
        try:
            return str(value.isoformat())
        except Exception:
            pass
    return str(value).strip()


def _heading_score(query_norm: str, heading_norm: str) -> float:
    if not query_norm or not heading_norm:
        return 0.0
    if query_norm == heading_norm:
        return 10.0
    if query_norm in heading_norm:
        return 7.0 + len(query_norm) / max(1, len(heading_norm))
    if heading_norm in query_norm:
        return 6.0 + len(heading_norm) / max(1, len(query_norm))
    query_tokens = set(query_norm.split())
    heading_tokens = set(heading_norm.split())
    if not query_tokens or not heading_tokens:
        return 0.0
    overlap = len(query_tokens & heading_tokens)
    if not overlap:
        return 0.0
    return overlap / max(1, len(query_tokens | heading_tokens))


def _find_best_heading(
    headings: list[dict[str, object]],
    query: str,
) -> dict[str, object] | None:
    query_norm = normalize_lookup_text(query)
    best: tuple[float, dict[str, object] | None] = (0.0, None)
    for heading in headings:
        heading_norm = str(heading.get("normalized") or "")
        score = _heading_score(query_norm, heading_norm)
        if score > best[0]:
            best = (score, heading)
    if best[0] <= 0.0:
        return None
    return best[1]


def _line_matches_heading(line: str, heading: dict[str, object]) -> bool:
    line_norm = normalize_lookup_text(line)
    heading_norm = str(heading.get("normalized") or "")
    return bool(line_norm and heading_norm and _heading_score(line_norm, heading_norm) >= 6.0)


def _extract_section_from_pdf_pages(
    pages: list[tuple[int, str]],
    headings: list[dict[str, object]],
    heading_query: str,
    max_chars: int,
) -> dict[str, Any]:
    match = _find_best_heading(headings, heading_query)
    if not match:
        return {"ok": False, "error": f"Heading not found: {heading_query}"}

    ordered = sorted(headings, key=lambda item: (int(item.get("page") or 0), int(item.get("line_index") or 0)))
    match_idx = ordered.index(match)
    next_heading = ordered[match_idx + 1] if match_idx + 1 < len(ordered) else None

    collecting = False
    chunks: list[str] = []
    total = 0
    page_start = int(match.get("page") or 0)
    page_end = page_start

    for page_num, body in pages:
        if page_num < page_start:
            continue
        lines = body.splitlines()
        started_here = False
        for line_idx, line in enumerate(lines, start=1):
            if not collecting:
                if page_num == page_start and _line_matches_heading(line, match):
                    collecting = True
                    started_here = True
                else:
                    continue
            if (
                next_heading
                and page_num == int(next_heading.get("page") or 0)
                and _line_matches_heading(line, next_heading)
                and not (started_here and page_num == page_start and line_idx == int(match.get("line_index") or 0))
            ):
                collecting = False
                break
            line_text = line.rstrip()
            if not line_text:
                continue
            chunks.append(line_text)
            total += len(line_text) + 1
            page_end = page_num
            if total >= max_chars:
                collecting = False
                break
        if not collecting and chunks:
            break

    content = truncate_text("\n".join(chunks).strip(), max(512, int(max_chars)))
    return {
        "ok": True,
        "matched_heading": str(match.get("heading") or heading_query),
        "matched_section": str(match.get("section") or ""),
        "page_start": page_start,
        "page_end": page_end,
        "content": content,
    }


def _derive_fact_check_queries(claim: str) -> list[str]:
    text = (claim or "").strip()
    if not text:
        return []
    queries: list[str] = []
    seen: set[str] = set()

    def add(value: str) -> None:
        normalized = _normalize_search_query(value)
        key = normalized.lower()
        if not normalized or key in seen:
            return
        seen.add(key)
        queries.append(normalized)

    for match in re.finditer(r"(?i)\b(?:0x[0-9a-f]{1,4}|[0-9a-f]{1,4}h)\b", text):
        add(match.group(0))
    for match in re.finditer(r'"([^"]+)"|“([^”]+)”|\'([^\']+)\'', text):
        for group in match.groups():
            if group:
                add(group)
    for match in re.finditer(r"\b\d+(?:\.\d+){1,5}\b", text):
        add(match.group(0))

    tokens = [
        token
        for token in re.findall(r"[A-Za-z][A-Za-z0-9/_-]{3,}", text)
        if token.lower() not in {"that", "with", "from", "there", "which", "this", "does", "have"}
    ]
    for token in tokens[:4]:
        add(token)
    if not queries:
        add(text)
    return queries[:8]


def _is_negative_claim(claim: str) -> bool:
    text = (claim or "").strip().lower()
    markers = (
        " not ",
        " no ",
        "none",
        "without",
        "does not",
        "is not",
        "isn't",
        "没有",
        "不存在",
        "未找到",
        "不是",
        "不支持",
    )
    padded = f" {text} "
    return any(marker in padded for marker in markers)


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r"[^\w.\-]+", "_", (name or "").strip(), flags=re.UNICODE).strip("._")
    if not cleaned:
        cleaned = "download.bin"
    return cleaned[:180]


def _guess_filename_from_response(url: str, content_type: str, content_disposition: str) -> str:
    cd = content_disposition or ""
    filename_match = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd, flags=re.IGNORECASE)
    if filename_match:
        name = urllib.parse.unquote(filename_match.group(1) or "").strip()
        if name:
            return _safe_filename(name)

    parsed = urllib.parse.urlsplit(url)
    candidate = Path(urllib.parse.unquote(parsed.path or "")).name
    if candidate:
        return _safe_filename(candidate)

    ct = (content_type or "").lower()
    if "application/pdf" in ct:
        return "download.pdf"
    if "application/json" in ct:
        return "download.json"
    if "text/html" in ct:
        return "download.html"
    if "text/plain" in ct:
        return "download.txt"
    return "download.bin"


def _find_subsequence(lines: list[str], chunk: list[str], start: int = 0) -> int:
    if not chunk:
        return max(0, min(len(lines), start))
    upper = len(lines) - len(chunk) + 1
    for index in range(max(0, start), max(0, upper)):
        if lines[index : index + len(chunk)] == chunk:
            return index
    return -1


def _parse_workspace_patch(patch_text: str) -> list[dict[str, Any]]:
    lines = str(patch_text or "").splitlines()
    if not lines or lines[0] != "*** Begin Patch":
        raise ValueError("patch must start with '*** Begin Patch'")
    if "*** End Patch" not in lines:
        raise ValueError("patch must end with '*** End Patch'")

    operations: list[dict[str, Any]] = []
    index = 1
    while index < len(lines):
        line = lines[index]
        if line == "*** End Patch":
            return operations
        if not line.strip():
            index += 1
            continue
        if line.startswith("*** Add File: "):
            raw_path = line[len("*** Add File: ") :].strip()
            if not raw_path:
                raise ValueError("Add File requires a target path")
            index += 1
            content_lines: list[str] = []
            while index < len(lines) and not lines[index].startswith("*** "):
                current = lines[index]
                if not current.startswith("+"):
                    raise ValueError(f"Add File expects '+' lines only: {current}")
                content_lines.append(current[1:])
                index += 1
            operations.append(
                {
                    "op": "add",
                    "path": raw_path,
                    "content": "\n".join(content_lines) + ("\n" if content_lines else ""),
                }
            )
            continue
        if line.startswith("*** Delete File: "):
            raw_path = line[len("*** Delete File: ") :].strip()
            if not raw_path:
                raise ValueError("Delete File requires a target path")
            operations.append({"op": "delete", "path": raw_path})
            index += 1
            continue
        if line.startswith("*** Update File: "):
            raw_path = line[len("*** Update File: ") :].strip()
            if not raw_path:
                raise ValueError("Update File requires a target path")
            index += 1
            move_to = raw_path
            if index < len(lines) and lines[index].startswith("*** Move to: "):
                move_to = lines[index][len("*** Move to: ") :].strip() or move_to
                index += 1
            hunks: list[list[str]] = []
            while index < len(lines) and not lines[index].startswith("*** "):
                header = lines[index]
                if not header.startswith("@@"):
                    raise ValueError(f"Unsupported patch section: {header}")
                index += 1
                hunk_lines: list[str] = []
                while index < len(lines) and not lines[index].startswith("@@") and not lines[index].startswith("*** "):
                    current = lines[index]
                    if current == "*** End of File":
                        index += 1
                        continue
                    if not current:
                        raise ValueError("Patch hunk lines must start with ' ', '+', or '-'")
                    prefix = current[:1]
                    if prefix not in {" ", "+", "-"}:
                        raise ValueError(f"Unsupported patch line: {current}")
                    hunk_lines.append(current)
                    index += 1
                if not hunk_lines:
                    raise ValueError(f"Empty patch hunk for {raw_path}")
                hunks.append(hunk_lines)
            if not hunks:
                raise ValueError(f"Update File requires at least one hunk: {raw_path}")
            operations.append(
                {
                    "op": "update",
                    "path": raw_path,
                    "move_to": move_to,
                    "hunks": hunks,
                }
            )
            continue
        raise ValueError(f"Unsupported patch operation: {line}")
    raise ValueError("patch ended unexpectedly before '*** End Patch'")


class LocalToolExecutor:
    def __init__(self, config: AppConfig) -> None:
        self.config = config
        self._runtime_ctx = threading.local()
        self._web_cache_lock = threading.Lock()
        self._docker_cache_lock = threading.Lock()
        self._taint_registry_lock = threading.Lock()
        self._command_sessions_lock = threading.Lock()
        self._command_sessions: dict[int, dict[str, Any]] = {}
        self._command_session_ids = itertools.count(1)
        self._docker_sandbox_cache: dict[tuple[str, ...], DockerSandboxManager] = {}
        self._web_cache_dir = (config.workspace_root / "app" / "data" / "web_cache").resolve()
        self._web_cache_dir.mkdir(parents=True, exist_ok=True)
        self._runtime_data_dir = (config.workspace_root / "app" / "data" / "runtime").resolve()
        self._runtime_data_dir.mkdir(parents=True, exist_ok=True)
        self._tool_result_store = ToolResultStore(config.sessions_dir.parent / "tool_results")
        self._taint_registry_path = (self._runtime_data_dir / "taint_registry.json").resolve()
        self._project_store = ProjectStore(config.projects_registry_path, default_root=config.workspace_root)
        self._browser_manager = BrowserToolManager(
            artifacts_dir=(config.workspace_root / "app" / "data" / "browser_artifacts").resolve(),
            mode=config.browser_mode,
            channel=config.browser_channel,
            headless=config.browser_headless,
            user_data_dir=config.browser_user_data_dir,
            executable_path=config.browser_executable_path,
            proxy_server=config.browser_proxy_server,
            ignore_https_errors=config.browser_ignore_https_errors,
            chromium_sandbox=config.browser_chromium_sandbox,
            disable_password_manager=config.browser_disable_password_manager,
        )
        self._image_read_handler: Callable[..., dict[str, Any]] | None = None
        self._docker_sandbox = DockerSandboxManager(
            workspace_root=config.workspace_root,
            allowed_roots=get_access_roots(config),
            docker_bin=config.docker_bin,
            image=config.docker_image,
            network=config.docker_network,
            memory=config.docker_memory,
            cpus=config.docker_cpus,
            pids_limit=config.docker_pids_limit,
            container_prefix=config.docker_container_prefix,
        )

    def set_runtime_context(
        self,
        *,
        execution_mode: str | None = None,
        session_id: str | None = None,
        project_id: str | None = None,
        project_root: str | None = None,
        cwd: str | None = None,
        model: str | None = None,
        locale: str | None = None,
        permission_profile: str | None = None,
        runtime_boundary: dict[str, Any] | None = None,
        run_id: str | None = None,
        cancel_event: Any | None = None,
        skill_writer: Any | None = None,
        task_writer: Any | None = None,
        reserved_skill_roots: list[str] | None = None,
        builtin_skill_roots: list[str] | None = None,
        team_skill_roots: list[str] | None = None,
        subagent_runner: Any | None = None,
        subagent_waiter: Any | None = None,
        subagent_read_only: bool = False,
    ) -> None:
        mode = (execution_mode or "").strip().lower()
        if mode not in {"host", "docker"}:
            mode = self.config.execution_mode
        self._runtime_ctx.execution_mode = mode
        sid = str(session_id or "").strip() or "__anon__"
        self._runtime_ctx.session_id = sid
        self._runtime_ctx.project_id = str(project_id or "").strip()
        self._runtime_ctx.project_root = str(project_root or "").strip()
        self._runtime_ctx.cwd = str(cwd or "").strip()
        self._runtime_ctx.model = str(model or "").strip()
        self._runtime_ctx.run_id = str(run_id or "").strip()
        self._runtime_ctx.locale = normalize_locale(locale, self.config.default_locale)
        self._runtime_ctx.permission_profile = normalize_permission_profile(
            permission_profile or getattr(self.config, "permission_profile", "auto")
        )
        self._runtime_ctx.runtime_boundary = dict(runtime_boundary or {})
        self._runtime_ctx.cancel_event = cancel_event
        self._runtime_ctx.skill_writer = skill_writer
        self._runtime_ctx.task_writer = task_writer
        self._runtime_ctx.reserved_skill_roots = [
            str(item) for item in list(reserved_skill_roots or []) if str(item or "").strip()
        ]
        self._runtime_ctx.builtin_skill_roots = [
            str(item) for item in list(builtin_skill_roots or []) if str(item or "").strip()
        ]
        self._runtime_ctx.team_skill_roots = [
            str(item) for item in list(team_skill_roots or []) if str(item or "").strip()
        ]
        self._runtime_ctx.subagent_runner = subagent_runner
        self._runtime_ctx.subagent_waiter = subagent_waiter
        self._runtime_ctx.subagent_read_only = bool(subagent_read_only)

    def clear_runtime_context(self) -> None:
        for key in ("execution_mode", "session_id", "project_id", "project_root", "cwd", "model", "run_id", "locale", "permission_profile", "runtime_boundary", "cancel_event", "skill_writer", "task_writer", "reserved_skill_roots", "builtin_skill_roots", "team_skill_roots", "subagent_runner", "subagent_waiter", "subagent_read_only"):
            try:
                delattr(self._runtime_ctx, key)
            except Exception:
                pass

    def _current_execution_mode(self) -> str:
        mode = str(getattr(self._runtime_ctx, "execution_mode", "") or "").strip().lower()
        if mode in {"host", "docker"}:
            return mode
        return self.config.execution_mode

    def _current_session_id(self) -> str:
        return str(getattr(self._runtime_ctx, "session_id", "") or "__anon__")

    def _current_run_id(self) -> str:
        return str(getattr(self._runtime_ctx, "run_id", "") or "")

    def _persist_tool_result(
        self,
        *,
        call_id: str,
        tool_name: str,
        content: str,
        token_count: int,
    ) -> str:
        return self._tool_result_store.save(
            thread_id=self._current_session_id(),
            run_id=self._current_run_id(),
            call_id=call_id,
            tool_name=tool_name,
            content=content,
            token_count=token_count,
        )

    def read_tool_result(
        self,
        result_ref: str,
        cursor: int = 0,
        max_tokens: int = 4000,
    ) -> dict[str, Any]:
        payload = self._tool_result_store.load(
            thread_id=self._current_session_id(),
            result_ref=str(result_ref or ""),
        )
        if payload is None:
            return {
                "ok": False,
                "error": {"kind": "tool_result_not_found", "result_ref": str(result_ref or "")},
                "summary": "The tool result reference is unavailable in this Thread.",
            }
        content = str(payload.get("content") or "")
        start = max(0, min(len(content), int(cursor or 0)))
        model_visible_cap = max(512, int(getattr(self.config, "tool_output_token_limit", 10_000) or 10_000))
        token_limit = max(
            128,
            min(8000, int(max_tokens or 4000), max(128, model_visible_cap - 256)),
        )
        chunk = truncate_text_to_token_limit(
            content[start:],
            model=str(getattr(self._runtime_ctx, "model", "") or self.config.default_model),
            max_tokens=token_limit,
        )
        if start < len(content) and not chunk:
            chunk = content[start : start + 1]
        next_cursor = start + len(chunk)
        complete = next_cursor >= len(content)
        return {
            "ok": True,
            "result_ref": str(result_ref or ""),
            "tool_name": str(payload.get("tool_name") or "unknown_tool"),
            "cursor": start,
            "next_cursor": None if complete else next_cursor,
            "cursor_unit": "characters",
            "complete": complete,
            "total_chars": len(content),
            "total_tokens": max(0, int(payload.get("token_count") or count_tokens(content, self.config.default_model))),
            "content": chunk,
            "summary": "Tool result continuation returned.",
        }

    def _current_project_id(self) -> str:
        return str(getattr(self._runtime_ctx, "project_id", "") or "")

    def _current_project_root(self) -> Path:
        raw = str(getattr(self._runtime_ctx, "project_root", "") or "").strip()
        if raw:
            return Path(raw).expanduser().resolve()
        default_project = self._project_store.ensure_default_project()
        return Path(str(default_project.get("root_path") or self.config.workspace_root)).resolve()

    def _current_cwd_hint(self) -> str:
        raw = str(getattr(self._runtime_ctx, "cwd", "") or "").strip()
        if raw:
            return raw
        return str(self._current_project_root())

    def _current_model_hint(self) -> str:
        return str(getattr(self._runtime_ctx, "model", "") or "").strip()

    def _current_run_id(self) -> str:
        return str(getattr(self._runtime_ctx, "run_id", "") or "").strip()

    def _current_cancel_requested(self) -> bool:
        event = getattr(self._runtime_ctx, "cancel_event", None)
        return bool(event and hasattr(event, "is_set") and event.is_set())

    def _current_subagent_read_only(self) -> bool:
        return bool(getattr(self._runtime_ctx, "subagent_read_only", False))

    def _current_locale_hint(self) -> str:
        fallback_locale = str(getattr(self.config, "default_locale", "ja-JP") or "ja-JP")
        return normalize_locale(getattr(self._runtime_ctx, "locale", ""), fallback_locale)

    @staticmethod
    def _extract_outlook_msg_payload_compat(
        extractor: Callable[..., dict[str, Any] | None],
        path: str,
        *,
        max_chars: int,
        locale: str,
    ) -> dict[str, Any]:
        try:
            return dict(extractor(path, max_chars=max_chars, locale=locale) or {})
        except TypeError as exc:
            if "locale" not in str(exc):
                raise
            return dict(extractor(path, max_chars=max_chars) or {})

    def set_image_read_handler(self, handler: Callable[..., dict[str, Any]] | None) -> None:
        self._image_read_handler = handler

    @staticmethod
    def _normalize_ocr_text(text: str, *, max_output_chars: int) -> str:
        raw = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
        lines = [line.rstrip() for line in raw.split("\n")]
        cleaned = "\n".join(line for line in lines if line.strip())
        cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
        if len(cleaned) > max_output_chars:
            cleaned = cleaned[:max_output_chars]
        return cleaned

    @staticmethod
    def _short_preview(value: Any, *, limit: int = 240) -> str:
        return _truncate_output(str(value or ""), max_chars=max(1, int(limit))).strip()

    @staticmethod
    def _extract_rapidocr_text(payload: Any) -> str:
        items = payload[0] if isinstance(payload, tuple) and payload else payload
        if not isinstance(items, list):
            return ""
        parts: list[str] = []
        for item in items:
            text = ""
            if isinstance(item, dict):
                text = str(item.get("text") or item.get("value") or "").strip()
            elif isinstance(item, (list, tuple)):
                if len(item) >= 2:
                    candidate = item[1]
                    if isinstance(candidate, dict):
                        text = str(candidate.get("text") or candidate.get("value") or "").strip()
                    elif isinstance(candidate, (list, tuple)):
                        text = str(candidate[0] or "").strip() if candidate else ""
                    else:
                        text = str(candidate or "").strip()
                elif len(item) == 1 and isinstance(item[0], str):
                    text = str(item[0] or "").strip()
            if text:
                parts.append(text)
        return "\n".join(parts)

    @staticmethod
    def _probe_rapidocr_status() -> tuple[bool, str]:
        rapidocr_spec = importlib.util.find_spec("rapidocr_onnxruntime")
        if rapidocr_spec is None:
            return False, "rapidocr unavailable: No module named 'rapidocr_onnxruntime'"
        onnx_spec = importlib.util.find_spec("onnxruntime")
        if onnx_spec is None:
            return False, "rapidocr unavailable: No module named 'onnxruntime'"
        return True, ""

    @staticmethod
    def _probe_tesseract_status() -> tuple[bool, str]:
        binary = shutil.which("tesseract")
        if binary:
            return True, binary
        return False, "tesseract is not installed"

    def ocr_status(self) -> dict[str, Any]:
        rapidocr_available, rapidocr_detail = self._probe_rapidocr_status()
        tesseract_available, tesseract_detail = self._probe_tesseract_status()
        warning = ""
        if rapidocr_available:
            warning = ""
        elif not tesseract_available:
            warning = f"{rapidocr_detail}; {tesseract_detail}"
        elif not rapidocr_available:
            warning = rapidocr_detail
        return {
            "rapidocr_available": rapidocr_available,
            "rapidocr_detail": rapidocr_detail,
            "tesseract_available": tesseract_available,
            "tesseract_detail": tesseract_detail,
            "default_engine": "rapidocr" if rapidocr_available else ("tesseract" if tesseract_available else ""),
            "warning": warning,
        }

    @staticmethod
    def _image_has_alpha(image: Image.Image) -> bool:
        if image.mode in {"RGBA", "LA"}:
            return True
        return bool(image.info.get("transparency"))

    def _prepare_image_for_ocr(self, path: str) -> tuple[str, Callable[[], None], list[str]]:
        notes: list[str] = []
        try:
            with Image.open(path) as raw_image:
                image = ImageOps.exif_transpose(raw_image)
                if self._image_has_alpha(image):
                    base = Image.new("RGBA", image.size, (255, 255, 255, 255))
                    base.alpha_composite(image.convert("RGBA"))
                    image = base.convert("RGB")
                    notes.append("flattened_alpha")
                elif image.mode not in {"RGB", "L"}:
                    image = image.convert("RGB")
                    notes.append(f"converted_mode:{raw_image.mode}->{image.mode}")

                long_edge = max(image.size)
                if long_edge:
                    target_long_edge = long_edge
                    if long_edge < 1600:
                        target_long_edge = min(2400, max(1600, long_edge * 3))
                    elif long_edge > 2400:
                        target_long_edge = 2400
                    if target_long_edge != long_edge:
                        scale = float(target_long_edge) / float(long_edge)
                        target_size = (
                            max(1, int(round(image.width * scale))),
                            max(1, int(round(image.height * scale))),
                        )
                        image = image.resize(target_size, Image.Resampling.LANCZOS)
                        notes.append(f"resized_for_ocr:{target_size[0]}x{target_size[1]}")

                if image.mode != "L":
                    image = ImageOps.grayscale(image)
                    notes.append("grayscale")
                image = ImageOps.autocontrast(image)
                image = ImageEnhance.Contrast(image).enhance(1.35)
                notes.append("contrast_enhanced")

                with tempfile.NamedTemporaryFile(prefix="vp_ocr_", suffix=".png", delete=False) as handle:
                    temp_path = Path(handle.name)
                image.save(temp_path, format="PNG", optimize=True)

            def _cleanup() -> None:
                try:
                    temp_path.unlink(missing_ok=True)
                except Exception:
                    pass

            return str(temp_path), _cleanup, notes
        except Exception as exc:
            return path, (lambda: None), [f"ocr_preprocess_failed:{exc}"]

    def _run_rapidocr_ocr(self, path: str, max_output_chars: int) -> dict[str, Any]:
        rapidocr_available, rapidocr_detail = self._probe_rapidocr_status()
        if not rapidocr_available:
            return {
                "ok": False,
                "engine": "rapidocr",
                "available": False,
                "error": rapidocr_detail or "rapidocr unavailable",
            }
        rapidocr_module = importlib.import_module("rapidocr_onnxruntime")

        try:
            engine = rapidocr_module.RapidOCR()
            raw_result = engine(str(path))
        except Exception as exc:
            return {
                "ok": False,
                "engine": "rapidocr",
                "available": True,
                "error": f"rapidocr failed: {exc}",
            }

        text = self._normalize_ocr_text(
            self._extract_rapidocr_text(raw_result),
            max_output_chars=max_output_chars,
        )
        if not text:
            return {
                "ok": False,
                "engine": "rapidocr",
                "available": True,
                "error": "rapidocr returned no readable text",
            }
        return {
            "ok": True,
            "engine": "rapidocr",
            "available": True,
            "visible_text": text,
        }

    def _run_tesseract_ocr(self, path: str, max_output_chars: int) -> dict[str, Any]:
        available, binary_or_error = self._probe_tesseract_status()
        if not available:
            return {
                "ok": False,
                "engine": "tesseract",
                "available": False,
                "error": binary_or_error or "tesseract is not installed",
            }
        binary = binary_or_error
        try:
            proc = subprocess.run(
                [binary, str(path), "stdout", "--psm", "6"],
                capture_output=True,
                text=True,
                timeout=20,
                check=False,
            )
        except Exception as exc:
            return {
                "ok": False,
                "engine": "tesseract",
                "available": True,
                "error": f"tesseract failed: {exc}",
            }

        text = self._normalize_ocr_text(proc.stdout or "", max_output_chars=max_output_chars)
        if not text:
            stderr_text = str(proc.stderr or "").strip()
            error = stderr_text or "tesseract returned no readable text"
            return {
                "ok": False,
                "engine": "tesseract",
                "available": True,
                "error": error,
            }
        warning = str(proc.stderr or "").strip()
        return {
            "ok": True,
            "engine": "tesseract",
            "available": True,
            "visible_text": text,
            "warning": warning or "",
        }

    def _perform_local_image_ocr(self, path: str, max_output_chars: int) -> dict[str, Any]:
        engines_tried: list[str] = []
        warnings: list[str] = []
        preprocess_notes: list[str] = []
        available = False
        last_error = ""
        prepared_path, cleanup_prepared_path, prep_notes = self._prepare_image_for_ocr(path)
        preprocess_notes.extend(note for note in prep_notes if note)
        try:
            for runner in (self._run_rapidocr_ocr, self._run_tesseract_ocr):
                result = runner(prepared_path, max_output_chars)
                engine = str(result.get("engine") or "").strip()
                if engine:
                    engines_tried.append(engine)
                available = available or bool(result.get("available"))
                warning = str(result.get("warning") or "").strip()
                if warning:
                    warnings.append(warning)
                if bool(result.get("ok")):
                    return {
                        "ok": True,
                        "visible_text": str(result.get("visible_text") or ""),
                        "ocr_available": available,
                        "engines_tried": engines_tried,
                        "warning": "; ".join(item for item in warnings if item) or "",
                        "ocr_engine": engine,
                        "preprocess_notes": preprocess_notes,
                    }
                error = str(result.get("error") or "").strip()
                if error:
                    warnings.append(error)
                    last_error = error
        finally:
            cleanup_prepared_path()
        return {
            "ok": False,
            "visible_text": "",
            "ocr_available": available,
            "engines_tried": engines_tried,
            "warning": "; ".join(item for item in warnings if item) or "",
            "error": last_error or ("ocr_unavailable" if not available else "ocr returned no readable text"),
            "preprocess_notes": preprocess_notes,
        }

    def _current_access_roots(self) -> list[Path]:
        roots: list[Path] = []
        seen: set[str] = set()

        def add(path: Path | None) -> None:
            if path is None:
                return
            resolved = path.resolve()
            key = str(resolved)
            if key in seen:
                return
            seen.add(key)
            roots.append(resolved)

        boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
        if isinstance(boundary, dict):
            for raw in list(boundary.get("allowed_roots") or []):
                try:
                    add(Path(str(raw)).expanduser())
                except Exception:
                    continue
        if roots:
            return roots
        add(self._current_project_root())
        try:
            add(Path(self.config.uploads_dir))
        except Exception:
            pass
        return roots

    def _current_command_roots(self) -> list[Path]:
        roots: list[Path] = []
        seen: set[str] = set()

        def add(path: Path | None) -> None:
            if path is None:
                return
            try:
                resolved = path.resolve()
            except Exception:
                return
            key = str(resolved)
            if key in seen:
                return
            seen.add(key)
            roots.append(resolved)

        boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
        if isinstance(boundary, dict):
            for raw in list(boundary.get("command_allowed_roots") or []):
                try:
                    add(Path(str(raw)).expanduser())
                except Exception:
                    continue
        if roots:
            return roots
        add(self._current_project_root())
        return roots

    def _current_shell_allowed(self) -> bool:
        boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
        if isinstance(boundary, dict) and "shell_allowed" in boundary:
            return bool(boundary.get("shell_allowed"))
        return self._current_permission_profile() != "default"

    def _current_permission_profile(self) -> str:
        boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
        if isinstance(boundary, dict) and str(boundary.get("permission_profile") or "").strip():
            return normalize_permission_profile(str(boundary.get("permission_profile") or ""))
        return normalize_permission_profile(
            str(getattr(self._runtime_ctx, "permission_profile", "") or getattr(self.config, "permission_profile", "auto"))
        )

    def _unrestricted_path_access(self) -> bool:
        return self._current_permission_profile() == "full_access"

    def _current_network_allowed(self) -> bool:
        boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
        if isinstance(boundary, dict) and "network_allowed" in boundary:
            return bool(boundary.get("network_allowed"))
        return self._current_permission_profile() == "full_access"

    def _supply_chain_approval_allowed(self) -> bool:
        return self._current_permission_profile() == "full_access" and self._current_network_allowed()

    def _resolve_path(self, raw_path: str) -> Path:
        return _resolve_workspace_path(
            self.config,
            raw_path,
            workspace_root=self._current_project_root(),
            access_roots=self._current_access_roots(),
            allow_any_path=self._unrestricted_path_access(),
        )

    def _resolve_source_path(self, raw_path: str) -> Path:
        return _resolve_source_path(
            self.config,
            raw_path,
            workspace_root=self._current_project_root(),
            access_roots=self._current_access_roots(),
            allow_any_path=self._unrestricted_path_access(),
        )

    def _docker_sandbox_for_context(self) -> DockerSandboxManager:
        project_root = self._current_project_root()
        access_roots = [path for path in self._current_access_roots() if path != project_root]
        cache_key = tuple([str(project_root), *[str(path) for path in access_roots]])
        with self._docker_cache_lock:
            sandbox = self._docker_sandbox_cache.get(cache_key)
            if sandbox is None:
                sandbox = DockerSandboxManager(
                    workspace_root=project_root,
                    allowed_roots=access_roots,
                    docker_bin=self.config.docker_bin,
                    image=self.config.docker_image,
                    network=self.config.docker_network,
                    memory=self.config.docker_memory,
                    cpus=self.config.docker_cpus,
                    pids_limit=self.config.docker_pids_limit,
                    container_prefix=self.config.docker_container_prefix,
                )
                self._docker_sandbox_cache[cache_key] = sandbox
            return sandbox

    def _decorate_result(self, result: dict[str, Any]) -> dict[str, Any]:
        if not isinstance(result, dict):
            return {"ok": False, "error": "Tool returned non-dict result"}
        payload = dict(result)
        project_root = str(self._current_project_root())
        payload.setdefault("project_root", project_root)
        payload.setdefault("cwd", str(payload.get("path") or self._current_cwd_hint() or project_root))
        payload.setdefault("project_id", self._current_project_id())
        return payload

    @staticmethod
    def _utc_timestamp() -> str:
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    @staticmethod
    def _sha256_file(path: Path) -> str:
        digest = hashlib.sha256()
        with open(path, "rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                if not chunk:
                    break
                digest.update(chunk)
        return digest.hexdigest()

    def _load_taint_registry_unlocked(self) -> dict[str, Any]:
        if not self._taint_registry_path.exists():
            return {"version": 1, "files": {}, "approvals": {}}
        try:
            payload = json.loads(self._taint_registry_path.read_text(encoding="utf-8"))
        except Exception:
            return {"version": 1, "files": {}, "approvals": {}}
        if not isinstance(payload, dict):
            return {"version": 1, "files": {}, "approvals": {}}
        if not isinstance(payload.get("files"), dict):
            payload["files"] = {}
        if not isinstance(payload.get("approvals"), dict):
            payload["approvals"] = {}
        payload["version"] = 1
        return payload

    def _save_taint_registry_unlocked(self, payload: dict[str, Any]) -> None:
        self._taint_registry_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = self._taint_registry_path.with_suffix(".tmp")
        tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
        tmp_path.replace(self._taint_registry_path)

    def _register_tainted_file(
        self,
        path: Path,
        *,
        source_url: str,
        source_tool: str,
        content_type: str = "",
        parent_path: str = "",
        parent_sha256: str = "",
        entry_name: str = "",
    ) -> dict[str, Any]:
        resolved = path.expanduser().resolve()
        if not resolved.exists() or not resolved.is_file():
            return {}
        parsed = urllib.parse.urlparse(str(source_url or ""))
        record = {
            "path": str(resolved),
            "sha256": self._sha256_file(resolved),
            "size": int(resolved.stat().st_size),
            "source_url": str(source_url or "").strip(),
            "source_domain": str(parsed.hostname or "").lower(),
            "source_tool": str(source_tool or "").strip(),
            "content_type": str(content_type or "").strip(),
            "parent_path": str(parent_path or "").strip(),
            "parent_sha256": str(parent_sha256 or "").strip(),
            "entry_name": str(entry_name or "").strip(),
            "session_id": self._current_session_id(),
            "project_id": self._current_project_id(),
            "run_id": self._current_run_id(),
            "marked_at": self._utc_timestamp(),
        }
        with self._taint_registry_lock:
            registry = self._load_taint_registry_unlocked()
            registry.setdefault("files", {})[str(resolved)] = record
            self._save_taint_registry_unlocked(registry)
        return dict(record)

    def _taint_record_for_path(self, path: Path) -> dict[str, Any] | None:
        try:
            resolved = path.expanduser().resolve()
        except Exception:
            return None
        with self._taint_registry_lock:
            registry = self._load_taint_registry_unlocked()
            raw = dict((registry.get("files") or {}).get(str(resolved)) or {})
        if not raw:
            return None
        raw["path"] = str(resolved)
        if resolved.exists() and resolved.is_file():
            try:
                raw["current_sha256"] = self._sha256_file(resolved)
                raw["current_size"] = int(resolved.stat().st_size)
            except Exception:
                raw["current_sha256"] = ""
        else:
            raw["missing"] = True
            raw["current_sha256"] = ""
        return raw

    @staticmethod
    def _approval_file_records(tainted_files: list[dict[str, Any]]) -> list[dict[str, str]]:
        return [
            {
                "path": str(item.get("path") or ""),
                "sha256": str(item.get("current_sha256") or item.get("sha256") or ""),
            }
            for item in tainted_files
        ]

    @staticmethod
    def _approval_risk_records(risks: list[dict[str, Any]]) -> list[dict[str, str]]:
        records: list[dict[str, str]] = []
        for item in list(risks or []):
            if not isinstance(item, dict):
                continue
            records.append(
                {
                    "kind": str(item.get("kind") or ""),
                    "category": str(item.get("category") or ""),
                    "message": str(item.get("message") or ""),
                    "base_command": str(item.get("base_command") or ""),
                    "blocked_argument": str(item.get("blocked_argument") or ""),
                    "subcommand": str(item.get("subcommand") or ""),
                    "operation": str(item.get("operation") or ""),
                    "repository_root": str(item.get("repository_root") or ""),
                    "remote": str(item.get("remote") or ""),
                    "remote_url": str(item.get("remote_url") or ""),
                    "remote_fingerprint": str(item.get("remote_fingerprint") or ""),
                    "branch": str(item.get("branch") or ""),
                    "head": str(item.get("head") or ""),
                    "refspecs": str(item.get("refspecs") or ""),
                    "force": str(bool(item.get("force"))).lower(),
                    "delete": str(bool(item.get("delete"))).lower(),
                }
            )
        return records

    def _create_command_execution_approval(
        self,
        *,
        command: str,
        purpose: str,
        cwd: str,
        risks: list[dict[str, Any]],
        tainted_files: list[dict[str, Any]],
    ) -> str:
        file_records = self._approval_file_records(tainted_files)
        risk_records = self._approval_risk_records(risks)
        token = secrets.token_urlsafe(24)
        approval = {
            "type": "command_execution",
            "token": token,
            "command": str(command or "").strip(),
            "purpose": str(purpose or "").strip()[:240],
            "cwd": str(cwd or "").strip(),
            "risks": risk_records,
            "files": file_records,
            "session_id": self._current_session_id(),
            "project_id": self._current_project_id(),
            "run_id": self._current_run_id(),
            "created_at": self._utc_timestamp(),
            "used": False,
        }
        with self._taint_registry_lock:
            registry = self._load_taint_registry_unlocked()
            registry.setdefault("approvals", {})[token] = approval
            self._save_taint_registry_unlocked(registry)
        return token

    def _create_tainted_execution_approval(self, *, command: str, tainted_files: list[dict[str, Any]]) -> str:
        return self._create_command_execution_approval(
            command=command,
            purpose="",
            cwd=self._current_cwd_hint(),
            risks=[],
            tainted_files=tainted_files,
        )

    @staticmethod
    def _sorted_approval_file_signature(files: list[dict[str, Any]]) -> list[tuple[str, str]]:
        return sorted(
            (
                str(item.get("path") or ""),
                str(item.get("current_sha256") or item.get("sha256") or ""),
            )
            for item in files
        )

    @staticmethod
    def _sorted_approval_risk_signature(risks: list[dict[str, Any]]) -> list[str]:
        records = LocalToolExecutor._approval_risk_records(risks)
        return sorted(json.dumps(item, ensure_ascii=False, sort_keys=True) for item in records)

    def _consume_command_execution_approval(
        self,
        *,
        token: str,
        command: str,
        cwd: str,
        risks: list[dict[str, Any]],
        tainted_files: list[dict[str, Any]],
    ) -> tuple[bool, str]:
        normalized_token = str(token or "").strip()
        if not normalized_token:
            return False, "A command execution approval token is required."
        current_files = self._sorted_approval_file_signature(tainted_files)
        current_risks = self._sorted_approval_risk_signature(risks)
        with self._taint_registry_lock:
            registry = self._load_taint_registry_unlocked()
            approvals = registry.setdefault("approvals", {})
            approval = dict(approvals.get(normalized_token) or {})
            if not approval:
                return False, "Command execution approval token was not found."
            if bool(approval.get("used")):
                return False, "Command execution approval token was already used."
            if str(approval.get("type") or "tainted_code_execution") not in {"command_execution", "tainted_code_execution"}:
                return False, "Command execution approval token type is not supported."
            approval_session_id = str(approval.get("session_id") or "").strip()
            current_session_id = self._current_session_id()
            if approval_session_id and current_session_id and approval_session_id != current_session_id:
                return False, "Command execution approval token does not match this session."
            approval_project_id = str(approval.get("project_id") or "").strip()
            current_project_id = self._current_project_id()
            if approval_project_id and current_project_id and approval_project_id != current_project_id:
                return False, "Command execution approval token does not match this project."
            if str(approval.get("command") or "").strip() != str(command or "").strip():
                return False, "Command execution approval token does not match this command."
            approval_cwd = str(approval.get("cwd") or "").strip()
            if approval_cwd and approval_cwd != str(cwd or "").strip():
                return False, "Command execution approval token does not match this cwd."
            approved_files = sorted(
                (
                    str(item.get("path") or ""),
                    str(item.get("sha256") or ""),
                )
                for item in list(approval.get("files") or [])
                if isinstance(item, dict)
            )
            if approved_files != current_files:
                return False, "Command execution approval token does not match the current file hashes."
            approved_risks = sorted(
                json.dumps(dict(item), ensure_ascii=False, sort_keys=True)
                for item in list(approval.get("risks") or [])
                if isinstance(item, dict)
            )
            if approved_risks != current_risks:
                return False, "Command execution approval token does not match these risk details."
            approval["used"] = True
            approval["used_at"] = self._utc_timestamp()
            approval["used_run_id"] = self._current_run_id()
            approvals[normalized_token] = approval
            self._save_taint_registry_unlocked(registry)
        return True, ""

    def _consume_tainted_execution_approval(
        self,
        *,
        token: str,
        command: str,
        tainted_files: list[dict[str, Any]],
    ) -> tuple[bool, str]:
        return self._consume_command_execution_approval(
            token=token,
            command=command,
            cwd=self._current_cwd_hint(),
            risks=[],
            tainted_files=tainted_files,
        )

    @staticmethod
    def _resolve_command_file_candidate(raw: str, *, cwd: Path) -> Path:
        candidate = Path(str(raw or "")).expanduser()
        if not candidate.is_absolute():
            candidate = cwd / candidate
        return candidate.resolve(strict=False)

    @staticmethod
    def _command_base_name(raw: str) -> str:
        return str(raw or "").replace("\\", "/").rsplit("/", 1)[-1].lower()

    @classmethod
    def _is_direct_path_command(cls, raw: str) -> bool:
        text = str(raw or "").strip()
        return bool(text) and (text.startswith(("/", "./", "../", "~")) or "/" in text or "\\" in text)

    @classmethod
    def _is_source_builtin_command(cls, raw: str) -> bool:
        return cls._command_base_name(raw) in {"source", "."}

    @staticmethod
    def _is_python_command(argv0: str) -> bool:
        return LocalToolExecutor._command_base_name(argv0) in {"python", "python3", "py", "python.exe"}

    @staticmethod
    def _is_node_command(argv0: str) -> bool:
        return LocalToolExecutor._command_base_name(argv0) in {"node", "node.exe"}

    @staticmethod
    def _is_shell_command(argv0: str) -> bool:
        return LocalToolExecutor._command_base_name(argv0) in {"sh", "bash", "zsh"}

    def _candidate_execution_paths_from_argv(self, argv: list[str], *, cwd: Path) -> list[Path]:
        if not argv:
            return []
        candidates: list[Path] = []
        base = str(argv[0] or "").strip()
        base_name = self._command_base_name(base)
        if self._is_direct_path_command(base):
            candidates.append(self._resolve_command_file_candidate(base, cwd=cwd))

        args = [str(item or "").strip() for item in list(argv[1:] or []) if str(item or "").strip()]
        if self._is_python_command(base) or self._is_node_command(base) or self._is_shell_command(base):
            skip_next = False
            for arg in args:
                if skip_next:
                    skip_next = False
                    continue
                if arg in {"-m", "-c", "-e", "--eval", "--command"}:
                    skip_next = arg in {"-c", "-e", "--eval", "--command"}
                    if arg == "-m":
                        break
                    continue
                if arg.startswith("-"):
                    continue
                candidates.append(self._resolve_command_file_candidate(arg, cwd=cwd))
                break
        elif base_name == "pytest":
            for arg in args:
                if arg.startswith("-"):
                    continue
                candidates.append(self._resolve_command_file_candidate(arg, cwd=cwd))
        elif base_name in {"source", "."} and args:
            candidates.append(self._resolve_command_file_candidate(args[0], cwd=cwd))

        unique: list[Path] = []
        seen: set[str] = set()
        for candidate in candidates:
            key = str(candidate)
            if key in seen:
                continue
            seen.add(key)
            unique.append(candidate)
        return unique

    def _tainted_execution_matches_for_argv(self, argv: list[str], *, cwd: Path) -> list[dict[str, Any]]:
        matches: list[dict[str, Any]] = []
        for candidate in self._candidate_execution_paths_from_argv(argv, cwd=cwd):
            record = self._taint_record_for_path(candidate)
            if record is not None:
                matches.append(record)
        return matches

    def _tainted_execution_matches(
        self,
        *,
        argv: list[str],
        cwd: Path,
        compound_validation: dict[str, Any] | None = None,
    ) -> list[dict[str, Any]]:
        matches: list[dict[str, Any]] = []
        if compound_validation:
            for item in list(compound_validation.get("subcommands") or []):
                if not isinstance(item, dict):
                    continue
                sub_argv = [str(token) for token in list(item.get("argv") or []) if str(token or "").strip()]
                sub_cwd = Path(str(item.get("effective_cwd") or cwd)).expanduser().resolve()
                matches.extend(self._tainted_execution_matches_for_argv(sub_argv, cwd=sub_cwd))
        else:
            matches.extend(self._tainted_execution_matches_for_argv(argv, cwd=cwd))

        unique: list[dict[str, Any]] = []
        seen: set[str] = set()
        for item in matches:
            key = str(item.get("path") or "")
            if not key or key in seen:
                continue
            seen.add(key)
            unique.append(item)
        return unique

    def _tainted_file_path_validation_error(self, tainted_files: list[dict[str, Any]]) -> dict[str, Any] | None:
        if self._unrestricted_path_access():
            return None
        command_roots = [root.expanduser().resolve() for root in self._current_command_roots()]
        for item in tainted_files:
            raw_path = str(item.get("path") or "").strip()
            if not raw_path:
                continue
            try:
                resolved = Path(raw_path).expanduser().resolve(strict=False)
            except Exception:
                continue
            boundary_path = resolved if resolved.exists() else resolved.parent
            if any(_is_within(boundary_path, root) for root in command_roots):
                continue
            return {
                "kind": "command_path_outside_allowed_roots",
                "message": "Tainted executable path is outside command allowed roots.",
                "argument": raw_path,
                "resolved_path": str(resolved),
                "command_allowed_roots": [str(root) for root in command_roots],
                }
        return None

    @staticmethod
    def _approval_files_public_payload(tainted_files: list[dict[str, Any]]) -> list[dict[str, str]]:
        return [
            {
                "path": str(item.get("path") or ""),
                "sha256": str(item.get("current_sha256") or item.get("sha256") or ""),
                "source_url": str(item.get("source_url") or ""),
                "source_domain": str(item.get("source_domain") or ""),
                "source_tool": str(item.get("source_tool") or ""),
                "content_type": str(item.get("content_type") or ""),
            }
            for item in tainted_files
        ]

    def _tainted_execution_risks(self, tainted_files: list[dict[str, Any]]) -> list[dict[str, str]]:
        if not tainted_files:
            return []
        return [
            {
                "kind": "tainted_code_execution",
                "category": "tainted_code",
                "message": "Command would execute code that originated from the network.",
                "base_command": "",
                "blocked_argument": "",
                "subcommand": "",
            }
        ]

    @staticmethod
    def _supply_chain_risk_from_block(
        block: dict[str, Any],
        *,
        subcommand: str = "",
        cwd: str = "",
    ) -> dict[str, str]:
        return {
            "kind": "supply_chain_command",
            "category": "supply_chain",
            "message": str(block.get("message") or "Command can fetch or execute network-origin package/code."),
            "base_command": str(block.get("base_command") or ""),
            "blocked_argument": str(block.get("blocked_argument") or ""),
            "subcommand": str(subcommand or ""),
            "cwd": str(cwd or ""),
        }

    def _supply_chain_risks(
        self,
        *,
        argv: list[str],
        cwd: Path,
        compound_validation: dict[str, Any] | None = None,
    ) -> list[dict[str, str]]:
        risks: list[dict[str, str]] = []
        if compound_validation:
            for item in list(compound_validation.get("subcommands") or []):
                if not isinstance(item, dict):
                    continue
                sub_argv = [str(token) for token in list(item.get("argv") or []) if str(token or "").strip()]
                block = blocked_supply_chain_command(sub_argv)
                if block is None:
                    continue
                risks.append(
                    self._supply_chain_risk_from_block(
                        block,
                        subcommand=str(item.get("text") or ""),
                        cwd=str(item.get("effective_cwd") or cwd),
                    )
                )
        else:
            block = blocked_supply_chain_command(argv)
            if block is not None:
                risks.append(self._supply_chain_risk_from_block(block, cwd=str(cwd)))

        unique: list[dict[str, str]] = []
        seen: set[str] = set()
        for item in risks:
            key = json.dumps(item, ensure_ascii=False, sort_keys=True)
            if key in seen:
                continue
            seen.add(key)
            unique.append(item)
        return unique

    @staticmethod
    def _git_push_index(argv: list[str]) -> int:
        """Return the concrete git subcommand index without interpreting aliases."""
        if not argv or LocalToolExecutor._command_base_name(argv[0]) not in {"git", "git.exe"}:
            return -1
        options_with_value = {
            "-C",
            "-c",
            "--config-env",
            "--exec-path",
            "--git-dir",
            "--work-tree",
            "--namespace",
            "--super-prefix",
        }
        index = 1
        while index < len(argv):
            token = str(argv[index] or "").strip()
            if not token:
                index += 1
                continue
            if token == "--":
                index += 1
                break
            if token in options_with_value:
                index += 2
                continue
            if any(token.startswith(prefix + "=") for prefix in options_with_value if prefix.startswith("--")):
                index += 1
                continue
            if token.startswith("-"):
                index += 1
                continue
            return index if token.lower() == "push" else -1
        if index < len(argv) and str(argv[index] or "").strip().lower() == "push":
            return index
        return -1

    @staticmethod
    def _git_command_cwd(argv: list[str], *, cwd: Path, subcommand_index: int) -> Path:
        effective = cwd.resolve()
        index = 1
        while index < max(1, subcommand_index):
            token = str(argv[index] or "").strip()
            if token == "-C" and index + 1 < subcommand_index:
                candidate = Path(str(argv[index + 1] or "")).expanduser()
                effective = (effective / candidate).resolve(strict=False) if not candidate.is_absolute() else candidate.resolve(strict=False)
                index += 2
                continue
            index += 1
        return effective

    @staticmethod
    def _sanitize_git_remote_url(raw: str) -> str:
        text = str(raw or "").strip()
        if not text:
            return ""
        parsed = urllib.parse.urlsplit(text)
        if parsed.scheme and parsed.hostname:
            host = str(parsed.hostname or "")
            if parsed.port:
                host = f"{host}:{parsed.port}"
            return urllib.parse.urlunsplit((parsed.scheme, host, parsed.path, "", ""))
        scp_match = re.match(r"^(?:[^@\s]+@)?([^:\s]+):(.+)$", text)
        if scp_match:
            return f"{scp_match.group(1)}:{scp_match.group(2)}"
        return text

    @staticmethod
    def _git_probe(cwd: Path, *args: str) -> str:
        git_bin = shutil.which("git")
        if not git_bin:
            return ""
        try:
            completed = subprocess.run(
                [git_bin, "-C", str(cwd), *[str(item) for item in args]],
                capture_output=True,
                text=True,
                timeout=3,
                check=False,
            )
        except Exception:
            return ""
        if int(completed.returncode or 0) != 0:
            return ""
        return str(completed.stdout or "").strip()

    @classmethod
    def _git_push_operands(cls, argv: list[str], *, push_index: int) -> dict[str, Any]:
        args = [str(item or "").strip() for item in argv[push_index + 1 :] if str(item or "").strip()]
        value_options = {
            "--exec",
            "--receive-pack",
            "--repo",
            "--push-option",
            "-o",
        }
        repository = ""
        refspecs: list[str] = []
        force = False
        delete = False
        index = 0
        while index < len(args):
            token = args[index]
            lowered = token.lower()
            if lowered in {"--force", "-f", "--force-if-includes"} or lowered.startswith("--force-with-lease"):
                force = True
            if lowered == "--delete":
                delete = True
            if token in value_options:
                if index + 1 < len(args):
                    if token == "--repo":
                        repository = args[index + 1]
                    index += 2
                    continue
                index += 1
                continue
            if token.startswith("--repo="):
                repository = token.split("=", 1)[1]
                index += 1
                continue
            if token.startswith("-"):
                index += 1
                continue
            if not repository:
                repository = token
            else:
                refspecs.append(token)
                if token.startswith("+"):
                    force = True
                if token.startswith(":") or token.endswith(":"):
                    delete = True
            index += 1
        return {
            "repository": repository,
            "refspecs": refspecs,
            "force": force,
            "delete": delete,
        }

    def _git_push_risk(self, argv: list[str], *, cwd: Path, subcommand: str = "") -> dict[str, Any] | None:
        push_index = self._git_push_index(argv)
        if push_index < 0:
            return None
        git_cwd = self._git_command_cwd(argv, cwd=cwd, subcommand_index=push_index)
        operands = self._git_push_operands(argv, push_index=push_index)
        repository_root = self._git_probe(git_cwd, "rev-parse", "--show-toplevel")
        branch = self._git_probe(git_cwd, "branch", "--show-current")
        head = self._git_probe(git_cwd, "rev-parse", "HEAD")
        remote_names_raw = self._git_probe(git_cwd, "remote")
        remote_names = {
            str(item or "").strip()
            for item in str(remote_names_raw or "").splitlines()
            if str(item or "").strip()
        }
        explicit_repository = str(operands.get("repository") or "").strip()
        remote = explicit_repository
        if not remote:
            for key in (
                f"branch.{branch}.pushRemote" if branch else "",
                "remote.pushDefault",
                f"branch.{branch}.remote" if branch else "",
            ):
                if not key:
                    continue
                remote = self._git_probe(git_cwd, "config", "--get", key)
                if remote:
                    break
        if not remote and "origin" in remote_names:
            remote = "origin"
        raw_remote_url = ""
        if remote and (remote in remote_names or not any(marker in remote for marker in ("://", "/", "\\"))):
            raw_remote_url = self._git_probe(git_cwd, "remote", "get-url", "--push", remote)
        if not raw_remote_url:
            raw_remote_url = remote
        remote_url = self._sanitize_git_remote_url(raw_remote_url)
        remote_fingerprint = hashlib.sha256(str(raw_remote_url or "").encode("utf-8")).hexdigest() if raw_remote_url else ""
        force = bool(operands.get("force"))
        delete = bool(operands.get("delete"))
        qualifiers = []
        if force:
            qualifiers.append("force update")
        if delete:
            qualifiers.append("remote deletion")
        qualifier_text = f" ({', '.join(qualifiers)})" if qualifiers else ""
        return {
            "kind": "external_side_effect",
            "category": "external_write",
            "operation": "git_push",
            "message": f"git push writes refs to a remote repository{qualifier_text} and requires explicit one-time approval.",
            "base_command": "git",
            "blocked_argument": "push",
            "subcommand": str(subcommand or ""),
            "repository_root": str(repository_root or git_cwd),
            "remote": remote or "(unresolved)",
            "remote_url": remote_url or "(unresolved)",
            "remote_fingerprint": remote_fingerprint,
            "branch": branch or "(detached or unresolved)",
            "head": head or "(unresolved)",
            "refspecs": " ".join(str(item) for item in list(operands.get("refspecs") or [])),
            "force": force,
            "delete": delete,
        }

    def _external_side_effect_risks(
        self,
        *,
        argv: list[str],
        cwd: Path,
        compound_validation: dict[str, Any] | None = None,
    ) -> list[dict[str, Any]]:
        risks: list[dict[str, Any]] = []
        if compound_validation:
            for item in list(compound_validation.get("subcommands") or []):
                if not isinstance(item, dict):
                    continue
                sub_argv = [str(token) for token in list(item.get("argv") or []) if str(token or "").strip()]
                sub_cwd = Path(str(item.get("effective_cwd") or cwd)).expanduser().resolve()
                risk = self._git_push_risk(sub_argv, cwd=sub_cwd, subcommand=str(item.get("text") or ""))
                if risk is not None:
                    risks.append(risk)
        else:
            risk = self._git_push_risk(argv, cwd=cwd)
            if risk is not None:
                risks.append(risk)
        return risks

    def _command_execution_approval_failure_result(
        self,
        *,
        command: str,
        purpose: str = "",
        cwd: str,
        risks: list[dict[str, Any]],
        tainted_files: list[dict[str, Any]],
        token_error: str = "",
    ) -> dict[str, Any]:
        all_risks = [*list(risks or []), *self._tainted_execution_risks(tainted_files)]
        approval_token = "" if token_error else self._create_command_execution_approval(
            command=command,
            purpose=purpose,
            cwd=cwd,
            risks=all_risks,
            tainted_files=tainted_files,
        )
        files = self._approval_files_public_payload(tainted_files)
        has_tainted = bool(files)
        has_external_write = any(str(item.get("category") or "") == "external_write" for item in all_risks)
        message = token_error or (
            "External write requires explicit one-time approval. Verify the repository, remote, branch, and command before continuing."
            if has_external_write
            else (
                "Execution requires approval because the command can fetch or execute network-origin code."
                if all_risks and not has_tainted
                else "Execution blocked because the command would run code that originated from the network."
            )
        )
        approval_request = {
            "type": "command_execution",
            "approval_token": approval_token,
            "command": str(command or "").strip(),
            "purpose": str(purpose or "").strip()[:240],
            "cwd": str(cwd or "").strip(),
            "risks": [dict(item) for item in all_risks],
            "files": files,
            "single_use": True,
            "default_action": "cancel",
        }
        payload = self._command_failure_result(
            command=command,
            cwd=cwd,
            error=message,
            stderr=message,
            error_kind=(
                "external_side_effect_approval_required"
                if has_external_write
                else ("tainted_code_approval_required" if has_tainted else "command_execution_approval_required")
            ),
            error_detail={
                "approval_token": approval_token,
                "command": str(command or "").strip(),
                "purpose": str(purpose or "").strip()[:240],
                "cwd": str(cwd or "").strip(),
                "risks": [dict(item) for item in all_risks],
                "files": files,
                "single_use": True,
            },
        )
        payload.update(
            {
                "approval_required": True,
                "approval_request": approval_request,
                "summary": (
                    "External write requires explicit approval."
                    if has_external_write
                    else "Command execution requires explicit approval."
                ),
            }
        )
        return payload

    def _tainted_execution_failure_result(
        self,
        *,
        command: str,
        cwd: str,
        tainted_files: list[dict[str, Any]],
        token_error: str = "",
    ) -> dict[str, Any]:
        return self._command_execution_approval_failure_result(
            command=command,
            cwd=cwd,
            risks=[],
            tainted_files=tainted_files,
            token_error=token_error,
        )

    def _mark_extracted_files_tainted_from_parent(
        self,
        *,
        source_path: str,
        extracted_files: list[Any],
    ) -> list[dict[str, Any]]:
        try:
            parent_path = self._resolve_source_path(source_path)
        except Exception:
            return []
        parent_taint = self._taint_record_for_path(parent_path)
        if not parent_taint:
            return []
        source_url = str(parent_taint.get("source_url") or "")
        parent_sha256 = str(parent_taint.get("current_sha256") or parent_taint.get("sha256") or "")
        marked: list[dict[str, Any]] = []
        for item in extracted_files:
            raw_path = ""
            entry_name = ""
            if isinstance(item, dict):
                raw_path = str(item.get("path") or item.get("resolved_path") or "")
                entry_name = str(item.get("entry_name") or item.get("name") or "")
            else:
                raw_path = str(item or "")
            if not raw_path:
                continue
            try:
                record = self._register_tainted_file(
                    Path(raw_path),
                    source_url=source_url,
                    source_tool="archive_extract",
                    content_type=str(parent_taint.get("content_type") or ""),
                    parent_path=str(parent_path),
                    parent_sha256=parent_sha256,
                    entry_name=entry_name,
                )
            except Exception:
                record = {}
            if record:
                marked.append(
                    {
                        "path": record.get("path"),
                        "sha256": record.get("sha256"),
                        "source_url": record.get("source_url"),
                        "source_domain": record.get("source_domain"),
                        "entry_name": record.get("entry_name"),
                    }
                )
        return marked

    def _project_python_candidates(self) -> list[Path]:
        project_root = self._current_project_root()
        candidates = [
            (project_root / ".venv" / "bin" / "python").resolve(),
            (project_root / ".venv" / "Scripts" / "python.exe").resolve(),
        ]
        return [candidate for candidate in candidates if candidate.exists()]

    def _preferred_python_command(self, *, execution_mode: str) -> str:
        if execution_mode == "docker":
            return "python3"
        project_python = next((str(candidate) for candidate in self._project_python_candidates()), "")
        if project_python:
            return project_python
        return str(self.config.python_command or "python").strip() or "python"

    def _is_project_python_command(self, raw_command: str) -> bool:
        raw = str(raw_command or "").strip()
        if not raw:
            return False
        normalized_raw = raw.replace("\\", "/")
        candidate_path = Path(normalized_raw).expanduser()
        project_root = self._current_project_root()
        try:
            resolved = candidate_path.resolve() if candidate_path.is_absolute() else (project_root / candidate_path).resolve()
        except Exception:
            return False
        allowed = {str(candidate.resolve()) for candidate in self._project_python_candidates()}
        return str(resolved) in allowed

    def _is_allowed_command(self, base_cmd: str) -> bool:
        raw = str(base_cmd or "").strip()
        if not raw:
            return False
        return raw in self.config.allowed_commands or self._is_project_python_command(raw)

    def _command_failure_result(
        self,
        *,
        command: str,
        cwd: str,
        error: str,
        returncode: int = 126,
        stderr: str | None = None,
        error_kind: str | None = None,
        error_detail: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        cwd_text = str(cwd or self._current_cwd_hint() or self._current_project_root())
        payload = {
            "ok": False,
            "command": str(command or "").strip(),
            "cwd": cwd_text,
            "host_cwd": cwd_text,
            "returncode": returncode,
            "stdout": "",
            "stderr": str(stderr if stderr is not None else error),
            "error": str(error or "").strip(),
            "execution_mode": self._current_execution_mode(),
        }
        if error_kind:
            payload["error_kind"] = str(error_kind)
            payload["error_detail"] = dict(error_detail or {})
            if str(error_kind) in {
                "command_not_allowed",
                "command_path_outside_allowed_roots",
                "dangerous_command",
                "reserved_skill_path",
            }:
                payload["failure_outcome"] = "rejected"
        return payload

    def _normalize_python_command_argv(self, argv: list[str], *, execution_mode: str) -> list[str]:
        if not argv:
            return argv
        base = str(argv[0] or "").strip().lower()
        if base not in {"python", "python3", "py"}:
            return argv
        normalized = list(argv)
        normalized[0] = self._preferred_python_command(execution_mode=execution_mode)
        return normalized

    def _safe_split_command(
        self,
        command: str,
        *,
        for_session: bool = False,
        allow_supply_chain_commands: bool = False,
    ) -> tuple[list[str], str | None]:
        raw = str(command or "").strip()
        if not raw:
            return [], "Empty command"
        if self._is_compound_shell_command(raw):
            return [], "Command requires compound shell validation."
        try:
            argv = shlex.split(raw)
        except Exception as exc:
            return [], f"Command parse failed: {exc}"
        if not argv:
            return [], "Empty command"
        execution_mode = self._current_execution_mode()
        argv = self._normalize_python_command_argv(argv, execution_mode=execution_mode)
        base_cmd = argv[0]
        supply_chain_block = blocked_supply_chain_command(argv)
        if not self._is_allowed_command(base_cmd):
            return [], f"Command not allowed: {base_cmd}. Allowed: {', '.join(self.config.allowed_commands)}"
        missing_supply_chain_commands = missing_supply_chain_allowed_commands(supply_chain_block, set(self.config.allowed_commands))
        if missing_supply_chain_commands:
            return [], f"Command not allowed: {', '.join(missing_supply_chain_commands)}. Allowed: {', '.join(self.config.allowed_commands)}"
        if supply_chain_block is not None and not allow_supply_chain_commands:
            return [], str(supply_chain_block.get("message") or "Command is blocked by supply-chain policy.")
        if for_session and execution_mode == "docker":
            return [], "Interactive exec_command sessions are only supported in host mode."
        return argv, None

    def _is_compound_shell_command(self, raw: str) -> bool:
        return shell_command_uses_compound_syntax(raw)

    def _parse_compound_shell_for_validation(self, raw: str) -> dict[str, Any]:
        return parse_compound_shell_command(raw)

    def _validate_compound_shell_command(self, raw: str, cwd: Path) -> tuple[bool, dict[str, Any]]:
        unrestricted_paths = self._unrestricted_path_access()
        ok, detail = validate_compound_shell_command_shared(
            raw,
            cwd=cwd.resolve(),
            command_allowed_roots=self._current_command_roots(),
            writable_roots=[Path(item) for item in self._current_writable_roots()],
            allowed_commands=self.config.allowed_commands,
            allow_supply_chain_commands=self._supply_chain_approval_allowed(),
            allow_any_path=unrestricted_paths,
        )
        if not ok:
            return False, dict(detail)
        parsed = self._parse_compound_shell_for_validation(raw)
        command_roots = self._current_command_roots()
        effective_cwd = cwd.resolve()
        parsed_subcommands = [str(item) for item in list(parsed.get("parsed_subcommands") or []) if str(item or "").strip()]
        validated_subcommands: list[dict[str, Any]] = []

        def reject(index: int, subcommand_text: str, reason: str, detail: dict[str, Any] | None = None) -> tuple[bool, dict[str, Any]]:
            payload = {
                "ok": False,
                "error_kind": "compound_shell_subcommand_rejected",
                "summary": f"Compound command subcommand #{index} did not pass validation.",
                "failed_subcommand": str(subcommand_text or "").strip(),
                "failed_index": int(index),
                "reason": str(reason or "").strip(),
                "parsed_subcommands": list(parsed_subcommands),
            }
            if detail:
                payload["detail"] = dict(detail)
            return False, payload

        for index, subcommand in enumerate(list(parsed.get("subcommands") or []), start=1):
            item = dict(subcommand or {})
            argv = [str(token) for token in list(item.get("argv") or []) if str(token or "").strip()]
            text = str(item.get("text") or "").strip()
            if not argv:
                return reject(index, text, "Subcommand is empty.")
            base_command = str(item.get("base_command") or "").strip() or Path(argv[0]).name.lower()
            if base_command == "cd":
                if len(argv) != 2:
                    return reject(index, text, "Only simple `cd <path>` subcommands are supported.")
                target = str(argv[1] or "").strip()
                resolved = (effective_cwd / Path(target).expanduser()).resolve(strict=False) if not Path(target).expanduser().is_absolute() else Path(target).expanduser().resolve(strict=False)
                boundary_path = resolved if resolved.exists() else resolved.parent
                if not unrestricted_paths and not any(_is_within(boundary_path, root) for root in command_roots):
                    return reject(
                        index,
                        text,
                        "cd target is outside command allowed roots.",
                        {
                            "kind": "command_path_outside_allowed_roots",
                            "message": "cd target is outside command allowed roots.",
                            "argument": target,
                            "resolved_path": str(resolved),
                            "command_allowed_roots": [str(root) for root in command_roots],
                        },
                    )
                if not resolved.exists() or not resolved.is_dir():
                    return reject(index, text, "cd target does not exist or is not a directory.")
                effective_cwd = resolved
                item["effective_cwd"] = str(effective_cwd)
                validated_subcommands.append(item)
                continue
            item["effective_cwd"] = str(effective_cwd)
            validated_subcommands.append(item)
        return True, {
            "ok": True,
            "compound_shell": bool(parsed.get("compound_shell", True)),
            "parsed_subcommands": parsed_subcommands,
            "subcommands": validated_subcommands,
        }

    def _shell_argv_for_compound_command(self, raw: str) -> list[str]:
        shell_bin = shutil.which("bash") or shutil.which("zsh") or shutil.which("sh") or "/bin/sh"
        return [shell_bin, "-lc", str(raw or "").strip()]

    def _command_path_validation_error(self, argv: list[str], *, cwd: Path) -> dict[str, Any] | None:
        if self._unrestricted_path_access():
            return None
        ok, detail = validate_command_path_args(
            argv,
            cwd=cwd.resolve(),
            command_allowed_roots=self._current_command_roots(),
            writable_roots=[Path(item) for item in self._current_writable_roots()],
        )
        if ok:
            return None
        return dict(detail or {})

    def _current_writable_roots(self) -> list[str]:
        boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
        if isinstance(boundary, dict):
            roots = [str(item) for item in list(boundary.get("writable_roots") or []) if str(item or "").strip()]
            if roots:
                return roots
            if boundary.get("workspace_write_allowed") is False:
                return []
        return [str(self._current_project_root())]

    def _current_workspace_write_allowed(self) -> bool:
        boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
        if isinstance(boundary, dict) and "workspace_write_allowed" in boundary:
            return bool(boundary.get("workspace_write_allowed"))
        profile = normalize_permission_profile(
            str(getattr(self._runtime_ctx, "permission_profile", "") or getattr(self.config, "permission_profile", "auto"))
        )
        return profile != "default"

    def _write_path_error(self, path: Path) -> str:
        if not self._current_workspace_write_allowed():
            return "Workspace write is not allowed for the active permission profile."
        if self._unrestricted_path_access():
            return ""
        try:
            resolved = path.resolve(strict=False)
        except Exception:
            resolved = path.absolute()
        roots = [Path(item).expanduser().resolve() for item in self._current_writable_roots()]
        target = resolved if resolved.exists() else resolved.parent
        if any(_is_within(target, root) for root in roots):
            return ""
        allowed = ", ".join(str(root) for root in roots[:6])
        return f"Write path outside writable roots: {resolved}. Writable roots: {allowed}"

    def _reserved_skill_write_error(self, path: Path) -> str:
        """Keep Built-in read-only and enforce path capabilities for Team Skills."""

        try:
            resolved = path.resolve(strict=False)
        except Exception:
            resolved = path.absolute()
        for raw_root in list(getattr(self._runtime_ctx, "builtin_skill_roots", []) or []):
            try:
                root = Path(str(raw_root)).expanduser().resolve()
            except Exception:
                continue
            if _is_within(resolved, root):
                return "Built-in Skill files are read-only. Create or update a Team Skill instead."
        for raw_root in list(getattr(self._runtime_ctx, "team_skill_roots", []) or []):
            try:
                root = Path(str(raw_root)).expanduser().resolve()
            except Exception:
                continue
            if not _is_within(resolved, root):
                continue
            boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
            team_write_allowed = bool(
                isinstance(boundary, dict)
                and boundary.get("team_skill_write_allowed")
            )
            writable_roots = [
                Path(item).expanduser().resolve()
                for item in self._current_writable_roots()
            ]
            if team_write_allowed and any(_is_within(resolved, writable_root) for writable_root in writable_roots):
                return ""
            return "Team Skill path is outside the active RuntimeBoundary writable scope."
        for raw_root in list(getattr(self._runtime_ctx, "reserved_skill_roots", []) or []):
            try:
                root = Path(str(raw_root)).expanduser().resolve()
            except Exception:
                continue
            if _is_within(resolved, root):
                return "Skill catalog files must be changed through save_skill, not ordinary file tools."

        normalized = "/" + str(resolved).replace("\\", "/").lower().strip("/") + "/"
        project_skill_markers = (
            "/.agents/skills/",
            "/.codex/skills/",
            "/workspace/skills/",
        )
        if any(marker in normalized for marker in project_skill_markers):
            return "Project/workspace Skill directories are disabled. Use save_skill to write the global Team catalog."
        return ""

    def _reserved_skill_write_recovery(self, path: Path) -> str:
        try:
            resolved = path.resolve(strict=False)
        except Exception:
            resolved = path.absolute()
        for raw_root in list(getattr(self._runtime_ctx, "builtin_skill_roots", []) or []):
            try:
                if _is_within(resolved, Path(str(raw_root)).expanduser().resolve()):
                    return "Built-in Skills cannot be modified. Make the change in a Team Skill instead."
            except Exception:
                continue
        for raw_root in list(getattr(self._runtime_ctx, "team_skill_roots", []) or []):
            try:
                if _is_within(resolved, Path(str(raw_root)).expanduser().resolve()):
                    return "Use apply_patch when this Team Skill path is writable in the active RuntimeBoundary. save_skill only replaces SKILL.md and cannot edit bundled scripts or references."
            except Exception:
                continue
        return "Use save_skill to create or replace a Team SKILL.md; project-level Skill directories are disabled."

    def _reserved_skill_command_error(self, command: str, *, cwd: Path) -> str:
        text = str(command or "").replace("\\", "/").lower()
        if self._is_direct_enabled_skill_script_command(command, cwd=cwd):
            return ""
        known_mutator = bool(
            re.search(
                r"(?:^|[;&|]\s*|\s)(?:cp|mv|mkdir|touch|tee|sed\s+-i|perl\s+-pi|rm|del|copy|move|"
                r"python(?:3)?\b|py\b|node\b|sh\b|bash\b|zsh\b|powershell\b|pwsh\b|cmd\b|"
                r"install\b|rsync\b|patch\b|unzip\b)",
                text,
            )
        )
        redirection_targets = re.findall(r"(?:^|\s)(?:\d*>>?|&>)\s*([^\s;&|]+)", text)
        if not known_mutator and not redirection_targets:
            return ""

        project_markers = (".agents/skills/", ".codex/skills/", "workspace/skills/")
        if known_mutator and any(marker in text for marker in project_markers):
            return "Project/workspace Skill directories are disabled. Use save_skill to write the global Team catalog."
        if (
            known_mutator
            and any(marker in text for marker in ("skills/builtin/", "skills/team/"))
            and not list(getattr(self._runtime_ctx, "builtin_skill_roots", []) or [])
            and not list(getattr(self._runtime_ctx, "team_skill_roots", []) or [])
        ):
            return "Skill catalog files must be changed through save_skill, not ordinary file tools."

        for raw_target in redirection_targets:
            value = str(raw_target or "").strip().strip("'\"")
            if not value:
                continue
            candidate = Path(value).expanduser()
            if not candidate.is_absolute():
                candidate = cwd / candidate
            error = self._reserved_skill_write_error(candidate)
            if error:
                return error

        if not known_mutator:
            return ""

        try:
            argv = shlex.split(str(command or ""))
        except Exception:
            argv = []
        for token in argv[1:]:
            value = str(token or "").strip().strip("<>")
            if not value or value.startswith("-") or "://" in value:
                continue
            candidate = Path(value).expanduser()
            if not candidate.is_absolute():
                candidate = cwd / candidate
            error = self._reserved_skill_write_error(candidate)
            if error:
                return error
        return ""

    def _direct_enabled_skill_script_context(self, command: str, *, cwd: Path) -> dict[str, str]:
        """Resolve a directly executed script and its enabled Skill root."""

        if self._is_compound_shell_command(command):
            return {}
        try:
            argv = shlex.split(str(command or "").strip())
        except Exception:
            return {}
        if len(argv) < 2:
            return {}
        base = self._command_base_name(argv[0])
        script_arg = ""
        allowed_suffixes: set[str] = set()
        if self._is_python_command(base):
            if any(arg in {"-c", "-m"} for arg in argv[1:]):
                return {}
            script_arg = next((arg for arg in argv[1:] if arg and not arg.startswith("-")), "")
            allowed_suffixes = {".py"}
        elif self._is_shell_command(base):
            if any(arg in {"-c", "--command"} for arg in argv[1:]):
                return {}
            script_arg = next((arg for arg in argv[1:] if arg and not arg.startswith("-")), "")
            allowed_suffixes = {".sh"}
        elif self._is_node_command(base):
            if any(arg in {"-e", "--eval"} for arg in argv[1:]):
                return {}
            script_arg = next((arg for arg in argv[1:] if arg and not arg.startswith("-")), "")
            allowed_suffixes = {".js", ".mjs", ".cjs"}
        elif base in {"powershell", "powershell.exe", "pwsh", "pwsh.exe"}:
            lowered = [str(arg or "").strip().lower() for arg in argv[1:]]
            if any(arg in {"-command", "-encodedcommand", "-c"} for arg in lowered):
                return {}
            for index, arg in enumerate(lowered):
                if arg in {"-file", "-f"} and index + 2 < len(argv):
                    script_arg = argv[index + 2]
                    break
            allowed_suffixes = {".ps1"}
        if not script_arg:
            return {}
        candidate = Path(script_arg).expanduser()
        if not candidate.is_absolute():
            candidate = cwd / candidate
        try:
            resolved = candidate.resolve()
        except Exception:
            return {}
        if not resolved.is_file() or resolved.suffix.lower() not in allowed_suffixes:
            return {}
        boundary = getattr(self._runtime_ctx, "runtime_boundary", None)
        enabled_roots: list[Path] = []
        if isinstance(boundary, dict):
            for item in list(boundary.get("enabled_skill_roots") or []):
                if not str(item or "").strip():
                    continue
                try:
                    enabled_roots.append(Path(str(item)).expanduser().resolve())
                except Exception:
                    continue
        skill_root = next((root for root in enabled_roots if _is_within(resolved, root)), None)
        if skill_root is None:
            return {}
        return {
            "skill_root": str(skill_root),
            "script_path": str(resolved),
        }

    def _is_direct_enabled_skill_script_command(self, command: str, *, cwd: Path) -> bool:
        """Allow normal interpreter execution for a script in an enabled Skill."""

        return bool(self._direct_enabled_skill_script_context(command, cwd=cwd))

    def _skill_script_environment(self, skill_context: dict[str, str], *, cwd: Path) -> dict[str, str]:
        env = dict(os.environ)
        project_root_raw = str(getattr(self._runtime_ctx, "project_root", "") or "").strip()
        project_root = Path(project_root_raw).expanduser() if project_root_raw else cwd
        try:
            project_root = project_root.resolve()
        except Exception:
            project_root = cwd
        env.update(
            {
                "VP_SKILL_ROOT": str(skill_context.get("skill_root") or ""),
                "VP_SKILL_SCRIPT": str(skill_context.get("script_path") or ""),
                "VP_PROJECT_ROOT": str(project_root),
                "VP_PROJECT_CWD": str(cwd.resolve()),
            }
        )
        return env

    def _spawn_command_reader(self, session_id: int, proc: subprocess.Popen[bytes]) -> None:
        def reader() -> None:
            stream = proc.stdout
            if stream is None:
                return
            try:
                while True:
                    chunk = stream.read(4096)
                    if not chunk:
                        break
                    text = chunk.decode("utf-8", errors="replace")
                    with self._command_sessions_lock:
                        session = self._command_sessions.get(session_id)
                        if session is None:
                            return
                        session["buffer"] = str(session.get("buffer") or "") + text
            finally:
                try:
                    stream.close()
                except Exception:
                    pass

        threading.Thread(target=reader, daemon=True).start()

    def _command_session_snapshot(self, session_id: int, *, max_output_chars: int) -> dict[str, Any]:
        with self._command_sessions_lock:
            session = self._command_sessions.get(session_id)
            if session is None:
                return {"ok": False, "error": f"Unknown session_id: {session_id}"}
            proc = session.get("proc")
            buffer_text = str(session.get("buffer") or "")
            cursor = int(session.get("cursor") or 0)
            if cursor > len(buffer_text):
                cursor = len(buffer_text)
            new_output = buffer_text[cursor:]
            session["cursor"] = len(buffer_text)
            cwd = str(session.get("cwd") or "")
            command = str(session.get("command") or "")
            execution_mode = str(session.get("execution_mode") or self._current_execution_mode())
            tty = bool(session.get("tty"))
            compound_shell = bool(session.get("compound_shell"))
            compound_validation = dict(session.get("compound_validation") or {}) if compound_shell else {}
            command_execution_approved = dict(session.get("command_execution_approved") or {})
            tainted_execution_approved = dict(session.get("tainted_execution_approved") or {})
        returncode = proc.poll() if isinstance(proc, subprocess.Popen) else 0
        status = "running" if returncode is None else "completed"
        payload: dict[str, Any] = {
            "ok": True,
            "session_id": int(session_id),
            "status": status,
            "running": returncode is None,
            "returncode": None if returncode is None else int(returncode),
            "output": _truncate_output(new_output, max_output_chars),
            "cwd": cwd,
            "command": command,
            "execution_mode": execution_mode,
            "tty": tty,
        }
        if compound_shell:
            payload["compound_shell"] = True
            payload["compound_validation"] = compound_validation
        if command_execution_approved:
            payload["command_execution_approved"] = command_execution_approved
        if tainted_execution_approved:
            payload["tainted_execution_approved"] = tainted_execution_approved
        if returncode is not None:
            payload["summary"] = f"command exited with {returncode}"
        return payload

    def _cancel_command_sessions(self, *, run_id: str = "") -> int:
        target_run_id = str(run_id or "").strip()
        with self._command_sessions_lock:
            sessions = [
                (session_id, session, session.get("proc"))
                for session_id, session in self._command_sessions.items()
                if (
                    (not target_run_id or str(session.get("run_id") or "").strip() == target_run_id)
                    and isinstance(session.get("proc"), subprocess.Popen)
                    and session["proc"].poll() is None
                )
            ]
            for _session_id, session, _proc in sessions:
                session["cancelled"] = True
        for _session_id, _session, proc in sessions:
            try:
                stdin = getattr(proc, "stdin", None)
                if stdin is not None:
                    stdin.close()
            except Exception:
                pass
            try:
                proc.terminate()
            except Exception:
                pass
        deadline = time.monotonic() + 0.5
        remaining = [proc for _session_id, _session, proc in sessions]
        while remaining and time.monotonic() < deadline:
            remaining = [proc for proc in remaining if proc.poll() is None]
            if remaining:
                time.sleep(0.01)
        for proc in remaining:
            try:
                proc.kill()
            except Exception:
                pass
        return len(sessions)

    def _apply_update_hunks(self, path: Path, current_text: str, hunks: list[list[str]]) -> str:
        lines = current_text.splitlines()
        cursor = 0
        for hunk in hunks:
            old_chunk = [entry[1:] for entry in hunk if entry[:1] in {" ", "-"}]
            new_chunk = [entry[1:] for entry in hunk if entry[:1] in {" ", "+"}]
            start = _find_subsequence(lines, old_chunk, cursor)
            if start < 0:
                start = _find_subsequence(lines, old_chunk, 0)
            if start < 0:
                raise ValueError(f"Patch context not found for {path}")
            end = start + len(old_chunk)
            lines = lines[:start] + new_chunk + lines[end:]
            cursor = start + len(new_chunk)
        updated = "\n".join(lines)
        if current_text.endswith("\n") or updated:
            updated += "\n"
        return updated

    def _web_cache_path(self, prefix: str, payload: dict[str, Any]) -> Path:
        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        digest = hashlib.sha256(raw).hexdigest()
        return self._web_cache_dir / f"{prefix}_{digest}.json"

    def _load_web_cache(self, prefix: str, payload: dict[str, Any], max_age_sec: int = 900) -> dict[str, Any] | None:
        path = self._web_cache_path(prefix, payload)
        if not path.is_file():
            return None
        try:
            with self._web_cache_lock:
                cached = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return None
        saved_at = float(cached.get("saved_at") or 0)
        if saved_at <= 0 or (time.time() - saved_at) > max_age_sec:
            return None
        payload_data = cached.get("payload")
        return payload_data if isinstance(payload_data, dict) else None

    def _save_web_cache(self, prefix: str, payload: dict[str, Any], result: dict[str, Any]) -> None:
        path = self._web_cache_path(prefix, payload)
        body = {
            "saved_at": time.time(),
            "payload": result,
        }
        try:
            with self._web_cache_lock:
                path.write_text(json.dumps(body, ensure_ascii=False), encoding="utf-8")
        except Exception:
            return

    def docker_available(self) -> bool:
        return self._docker_sandbox_for_context().docker_available()

    def docker_status(self) -> tuple[bool, str]:
        sandbox = self._docker_sandbox_for_context()
        ok = sandbox.docker_available()
        return ok, sandbox.docker_status_message()

    @property
    def tool_specs(self) -> list[dict[str, Any]]:
        return [
            {
                "type": "function",
                "name": "exec_command",
                "description": "Run a workspace command and keep a resumable command session for follow-up polling or stdin.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "cmd": {"type": "string", "description": "Command string, e.g. `rg TODO .` or `pytest tests/test_app.py`"},
                        "purpose": {
                            "type": "string",
                            "minLength": 1,
                            "maxLength": 240,
                            "description": "One concise, user-facing sentence explaining why this command is needed. This is display-only and never grants permission.",
                        },
                        "cwd": {"type": "string", "description": "Working directory relative to workspace", "default": "."},
                        "yield_time_ms": {"type": "integer", "minimum": 0, "maximum": 10000, "default": 1000},
                        "max_output_chars": {"type": "integer", "minimum": 256, "maximum": 60000, "default": 12000},
                        "tty": {
                            "type": "boolean",
                            "default": False,
                            "description": "Compatibility flag reported in the result; the current host runner uses pipes and does not allocate a PTY.",
                        },
                        "approval_token": {
                            "type": "string",
                            "default": "",
                            "description": "Single-use token required to run an approved high-risk command execution.",
                        },
                        "tainted_approval_token": {
                            "type": "string",
                            "default": "",
                            "description": "Deprecated alias for approval_token when running network-origin tainted files.",
                        },
                    },
                    "required": ["cmd", "purpose"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "write_stdin",
                "description": "Write characters to a running exec_command session, or poll for fresh output.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "session_id": {"type": "integer", "minimum": 1},
                        "chars": {"type": "string", "default": ""},
                        "yield_time_ms": {"type": "integer", "minimum": 0, "maximum": 10000, "default": 1000},
                        "max_output_chars": {"type": "integer", "minimum": 256, "maximum": 60000, "default": 12000},
                    },
                    "required": ["session_id"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "read_tool_result",
                "description": "Continue reading a tool result that was truncated for model context. This reads the original execution result and never reruns the tool.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "result_ref": {"type": "string", "description": "Opaque result_ref from a truncated tool response."},
                        "cursor": {"type": "integer", "minimum": 0, "default": 0, "description": "Character cursor returned by the previous chunk."},
                        "max_tokens": {"type": "integer", "minimum": 512, "maximum": 8000, "default": 4000},
                    },
                    "required": ["result_ref"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "apply_patch",
                "description": APPLY_PATCH_TOOL_DESCRIPTION,
                "parameters": {
                    "type": "object",
                    "properties": {
                        "patch": {"type": "string", "description": APPLY_PATCH_ARGUMENT_DESCRIPTION},
                        "cwd": {
                            "type": "string",
                            "default": ".",
                            "description": "Base directory used to resolve relative patch paths.",
                        },
                        "check": {
                            "type": "boolean",
                            "default": False,
                            "description": "Validate the complete patch without changing files when true.",
                        },
                    },
                    "required": ["patch"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "read_file",
                "description": "Read one local file. Supports chunked reads plus Office/PDF text extraction for large document formats.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "default": "."},
                        "start_char": {
                            "type": "integer",
                            "minimum": 0,
                            "default": 0,
                            "description": "Zero-based character offset in character mode.",
                        },
                        "max_chars": {
                            "type": "integer",
                            "minimum": 128,
                            "maximum": 1000000,
                            "default": 200000,
                            "description": "Maximum extracted characters returned in either mode.",
                        },
                        "start_line": {
                            "type": "integer",
                            "minimum": 0,
                            "default": 0,
                            "description": "One-based first line; 0 keeps character mode unless max_lines is set.",
                        },
                        "max_lines": {
                            "type": "integer",
                            "minimum": 0,
                            "maximum": 200000,
                            "default": 0,
                            "description": "Maximum lines; a value above 0 enables line mode.",
                        },
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "list_dir",
                "description": "List files and directories under one local directory path without reading file contents.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "default": "."},
                        "max_entries": {"type": "integer", "minimum": 1, "maximum": 500, "default": 200},
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "glob_file_search",
                "description": "Find files by glob pattern relative to the workspace or a given directory root.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "pattern": {"type": "string"},
                        "path": {"type": "string", "default": "."},
                        "max_results": {"type": "integer", "minimum": 1, "maximum": 500, "default": 200},
                    },
                    "required": ["pattern"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "search_contents_in_file",
                "description": "Search text inside one known local file or extracted document text and return evidence snippets with read hints.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "query": {"type": "string"},
                        "max_matches": {"type": "integer", "minimum": 1, "maximum": 20, "default": 8},
                        "context_chars": {"type": "integer", "minimum": 40, "maximum": 2000, "default": 280},
                    },
                    "required": ["path", "query"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "search_contents_in_file_multi",
                "description": "Run multiple text searches against one known local file or extracted document text and merge the evidence snippets.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "queries": {
                            "type": "array",
                            "items": {"type": "string"},
                        },
                        "per_query_max_matches": {"type": "integer", "minimum": 1, "maximum": 10, "default": 3},
                        "context_chars": {"type": "integer", "minimum": 40, "maximum": 2000, "default": 280},
                    },
                    "required": ["path", "queries"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "read_section",
                "description": "Read a document section by matching a heading or section number and returning that section's content.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "heading": {"type": "string"},
                        "max_chars": {"type": "integer", "minimum": 512, "maximum": 50000, "default": 12000},
                    },
                    "required": ["path", "heading"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "table_extract",
                "description": "Extract tables from a local PDF or OpenXML Excel workbook, optionally narrowed by query or page hint.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "query": {"type": "string", "default": ""},
                        "page_hint": {"type": "integer", "minimum": 0, "default": 0},
                        "max_tables": {"type": "integer", "minimum": 1, "maximum": 20, "default": 5},
                        "max_rows": {"type": "integer", "minimum": 1, "maximum": 200, "default": 25},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "fact_check_file",
                "description": "Retrieve file snippets related to a claim and return a heuristic evidence verdict that still requires model judgment.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "claim": {"type": "string"},
                        "queries": {
                            "type": "array",
                            "items": {"type": "string"},
                            "default": [],
                        },
                        "max_evidence": {"type": "integer", "minimum": 1, "maximum": 12, "default": 6},
                    },
                    "required": ["path", "claim"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "search_codebase",
                "description": "Search code or text files under a local root and return structured file, line, and text matches.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"},
                        "root": {"type": "string", "default": "."},
                        "max_matches": {"type": "integer", "minimum": 1, "maximum": 100, "default": 20},
                        "file_glob": {"type": "string", "default": ""},
                        "use_regex": {"type": "boolean", "default": False},
                        "case_sensitive": {"type": "boolean", "default": False},
                    },
                    "required": ["query"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "web_search",
                "description": "Search the web using the local hosted provider and return candidate URLs and snippets.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"},
                        "max_results": {"type": "integer", "minimum": 1, "maximum": 20, "default": 5},
                        "timeout_sec": {"type": "integer", "minimum": 3, "maximum": 30, "default": 12},
                    },
                    "required": ["query"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "web_fetch",
                "description": "Fetch one web page or document URL through the local hosted web fetcher.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string"},
                        "max_chars": {"type": "integer", "minimum": 512, "maximum": 500000, "default": 120000},
                        "timeout_sec": {"type": "integer", "minimum": 3, "maximum": 30, "default": 12},
                    },
                    "required": ["url"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "web_download",
                "description": "Download a web file (binary-safe) under allowed writable roots. Downloaded content is marked untrusted; executing it may require approval.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string"},
                        "dst_path": {"type": "string", "default": ""},
                        "overwrite": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                        "timeout_sec": {"type": "integer", "minimum": 3, "maximum": 120, "default": 20},
                        "max_bytes": {"type": "integer", "minimum": 1024, "maximum": 209715200, "default": 52428800},
                    },
                    "required": ["url"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "sessions_list",
                "description": "List recent local chat sessions for the current project so the agent can locate past context.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "limit": {"type": "integer", "minimum": 1, "maximum": 200, "default": 20},
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "sessions_history",
                "description": "Read one local chat session summary and recent turns by session_id.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "session_id": {"type": "string"},
                        "max_turns": {"type": "integer", "minimum": 1, "maximum": 800, "default": 80},
                    },
                    "required": ["session_id"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "image_inspect",
                "description": "Inspect a local image and return basic metadata such as size, mode, and format.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "image_read",
                "description": "Read a local image with zero-config OCR first, then optional multimodal analysis, and return visible text plus a concise analysis.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "prompt": {"type": "string", "default": ""},
                        "max_output_chars": {"type": "integer", "minimum": 256, "maximum": 24000, "default": 12000},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "archive_extract",
                "description": "Extract a local .zip archive under allowed writable roots; files inherit untrusted provenance from a downloaded archive.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "zip_path": {"type": "string"},
                        "dst_dir": {"type": "string", "default": ""},
                        "overwrite": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                        "max_entries": {"type": "integer", "minimum": 1, "maximum": 100000, "default": 20000},
                        "max_total_bytes": {"type": "integer", "minimum": 1024, "maximum": 2147483648, "default": 524288000},
                    },
                    "required": ["zip_path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "mail_extract_attachments",
                "description": "Extract attachments from a local Outlook .msg email into a target directory.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "msg_path": {"type": "string"},
                        "dst_dir": {"type": "string", "default": ""},
                        "overwrite": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                        "max_attachments": {"type": "integer", "minimum": 1, "maximum": 5000, "default": 500},
                        "max_total_bytes": {"type": "integer", "minimum": 1024, "maximum": 2147483648, "default": 524288000},
                    },
                    "required": ["msg_path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "spawn_subagent",
                "description": "Start one bounded Subagent task in an isolated context and immediately return its id. Independent Subagents can run in parallel; call wait_subagents to collect their results before using them.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "task": {
                            "type": "string",
                            "description": "A self-contained task with the scope, relevant paths, and expected result.",
                        },
                        "role": {
                            "type": "string",
                            "enum": ["explorer", "tester", "analyst", "summarizer"],
                            "default": "explorer",
                        },
                        "label": {
                            "type": "string",
                            "default": "",
                            "description": "Short user-facing label for the delegated work.",
                        },
                    },
                    "required": ["task"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "wait_subagents",
                "description": "Wait for selected running Subagents, or all current Subagents when ids are omitted, and return completed summaries plus any still-running ids.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "subagent_ids": {
                            "type": "array",
                            "items": {"type": "string"},
                            "default": [],
                        },
                        "timeout_seconds": {
                            "type": "number",
                            "minimum": 0,
                            "maximum": 300,
                            "default": 30,
                        },
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "update_plan",
                "description": "Synchronize a lightweight checklist. Keep exactly one step in_progress until every step is completed.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "explanation": {"type": "string", "default": ""},
                        "plan": {
                            "type": "array",
                            "items": {
                                "type": ["object", "string"],
                                "properties": {
                                    "step": {"type": "string"},
                                    "description": {"type": "string"},
                                    "status": {
                                        "type": "string",
                                        "enum": ["pending", "in_progress", "completed"],
                                    },
                                },
                                "additionalProperties": False,
                            },
                        },
                        "steps": {"type": "array", "items": {"type": ["object", "string"]}},
                        "items": {"type": "array", "items": {"type": ["object", "string"]}},
                        "tasks": {"type": "array", "items": {"type": ["object", "string"]}},
                        "plan_state": {"type": "array", "items": {"type": ["object", "string"]}},
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "request_user_input",
                "description": "Pause the turn and ask the user one to three structured follow-up questions.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "questions": {
                            "type": "array",
                            "minItems": 1,
                            "maxItems": 3,
                            "items": {
                                "type": "object",
                                "properties": {
                                    "header": {"type": "string", "maxLength": 12},
                                    "id": {"type": "string"},
                                    "question": {"type": "string"},
                                    "options": {
                                        "type": "array",
                                        "minItems": 2,
                                        "maxItems": 3,
                                        "items": {
                                            "type": "object",
                                            "properties": {
                                                "label": {"type": "string"},
                                                "description": {"type": "string"},
                                            },
                                            "required": ["label", "description"],
                                            "additionalProperties": False,
                                        },
                                    },
                                },
                                "required": ["header", "id", "question", "options"],
                                "additionalProperties": False,
                            },
                        },
                    },
                    "required": ["questions"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "save_skill",
                "description": "Create a repository-shared Team SKILL.md, or replace it only when overwrite is true. The Skill Registry resolves its location independently of the active project; Built-in Skills are never modified.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {
                            "type": "string",
                            "description": "Team Skill name, e.g. protocol-analysis. Use lowercase letters, digits, hyphens, or underscores.",
                        },
                        "description": {
                            "type": "string",
                            "description": "Trigger description that tells the model when this skill should be used.",
                        },
                        "body": {
                            "type": "string",
                            "description": "Markdown instruction body only. Do not include YAML frontmatter.",
                        },
                        "enabled": {"type": "boolean", "default": True},
                        "overwrite": {
                            "type": "boolean",
                            "default": False,
                            "description": "Set true to replace an existing Team Skill with the same name.",
                        },
                    },
                    "required": ["name", "description", "body"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "save_task",
                "description": "Create a durable Task snapshot for the current project, or replace the loaded Task snapshot when task_id is provided. Use it when the user asks to summarize/save the current work as a Task, and to checkpoint material progress on a loaded Task.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "task_id": {
                            "type": "string",
                            "description": "Existing Task id to update. Leave empty only when creating a new Task.",
                            "default": "",
                        },
                        "title": {
                            "type": "string",
                            "description": "Short, recognizable Task title for the Tasks list.",
                        },
                        "goal": {
                            "type": "string",
                            "description": "The concrete outcome that defines what this Task is trying to achieve.",
                        },
                        "summary": {
                            "type": "string",
                            "description": "Self-contained continuation summary with enough context to resume without opening the source Thread.",
                        },
                        "progress": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Important work already completed or verified.",
                            "default": [],
                        },
                        "next_steps": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Ordered concrete actions that should happen next.",
                            "default": [],
                        },
                        "decisions": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Key decisions and constraints that future work must preserve.",
                            "default": [],
                        },
                        "blockers": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Known blockers, missing inputs, or unresolved risks.",
                            "default": [],
                        },
                        "artifacts": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Relevant files, branches, commits, pull requests, or other durable artifacts.",
                            "default": [],
                        },
                        "status": {
                            "type": "string",
                            "enum": ["active", "blocked", "completed", "archived"],
                            "description": "Current lifecycle status of the Task.",
                            "default": "active",
                        },
                    },
                    "required": ["title", "goal", "summary"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "browser_open",
                "description": "Open a webpage in a headless browser session and capture the current page state.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string"},
                        "timeout_ms": {"type": "integer", "minimum": 1000, "maximum": 60000, "default": 20000},
                    },
                    "required": ["url"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "browser_click",
                "description": "Click one element in the current browser session by CSS selector.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "selector": {"type": "string"},
                        "timeout_ms": {"type": "integer", "minimum": 1000, "maximum": 60000, "default": 12000},
                    },
                    "required": ["selector"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "browser_type",
                "description": "Type or fill text into the current browser session by CSS selector.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "selector": {"type": "string"},
                        "text": {"type": "string"},
                        "submit": {"type": "boolean", "default": False},
                        "clear": {"type": "boolean", "default": True},
                        "timeout_ms": {"type": "integer", "minimum": 1000, "maximum": 60000, "default": 12000},
                    },
                    "required": ["selector", "text"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "browser_wait",
                "description": "Wait for a selector state, or wait only for timeout_ms when selector is empty.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "selector": {"type": "string", "default": ""},
                        "timeout_ms": {"type": "integer", "minimum": 250, "maximum": 60000, "default": 5000},
                        "state": {
                            "type": "string",
                            "enum": ["attached", "detached", "visible", "hidden"],
                            "default": "visible",
                        },
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "browser_scroll",
                "description": "Scroll the page by direction/amount, or ignore those fields and bring selector into view when selector is set.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "direction": {
                            "type": "string",
                            "enum": ["down", "up", "left", "right"],
                            "default": "down",
                        },
                        "amount": {"type": "integer", "minimum": 1, "maximum": 5000, "default": 900},
                        "selector": {"type": "string", "default": ""},
                        "timeout_ms": {"type": "integer", "minimum": 250, "maximum": 60000, "default": 5000},
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "browser_snapshot",
                "description": "Capture the current browser page title, URL, text excerpt, and top links.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "max_chars": {"type": "integer", "minimum": 400, "maximum": 50000, "default": 12000},
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "browser_screenshot",
                "description": "Save a screenshot from the current browser session to local storage.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "full_page": {"type": "boolean", "default": True},
                    },
                    "additionalProperties": False,
                },
            },
        ]

    def execute(self, name: str, arguments: dict[str, Any]) -> dict[str, Any]:
        arguments = self._normalize_public_tool_arguments(name, arguments)
        result: dict[str, Any]
        try:
            return self._execute_impl(name, arguments)
        except Exception as exc:
            return self._decorate_result(
                {
                    "ok": False,
                    "error": {
                        "kind": "tool_execution_error",
                        "tool": str(name or ""),
                        "message": safe_error_message(exc),
                    },
                    "summary": safe_error_message(exc),
                }
            )

    def _execute_impl(self, name: str, arguments: dict[str, Any]) -> dict[str, Any]:
        if name == "exec_command":
            result = self.exec_command(**arguments)
            return self._decorate_result(result)
        if name == "write_stdin":
            result = self.write_stdin(**arguments)
            return self._decorate_result(result)
        if name == "read_tool_result":
            result = self.read_tool_result(**arguments)
            return self._decorate_result(result)
        if name == "read_file":
            result = self.read_file(**arguments)
            return self._decorate_result(result)
        if name == "list_dir":
            result = self.list_dir(**arguments)
            return self._decorate_result(result)
        if name == "glob_file_search":
            result = self.glob_file_search(**arguments)
            return self._decorate_result(result)
        if name == "search_contents_in_file":
            result = self.search_contents_in_file(**arguments)
            return self._decorate_result(result)
        if name == "search_contents_in_file_multi":
            result = self.search_contents_in_file_multi(**arguments)
            return self._decorate_result(result)
        if name == "read_section":
            result = self.read_section(**arguments)
            return self._decorate_result(result)
        if name == "table_extract":
            result = self.table_extract(**arguments)
            return self._decorate_result(result)
        if name == "fact_check_file":
            result = self.fact_check_file(**arguments)
            return self._decorate_result(result)
        if name == "search_codebase":
            result = self.search_codebase(**arguments)
            return self._decorate_result(result)
        if name == "web_search":
            result = self.web_search(**arguments)
            return self._decorate_result(result)
        if name == "web_fetch":
            result = self.web_fetch(**arguments)
            return self._decorate_result(result)
        if name == "web_download":
            result = self.web_download(**arguments)
            return self._decorate_result(result)
        if name == "sessions_list":
            result = self.sessions_list(**arguments)
            return self._decorate_result(result)
        if name == "sessions_history":
            result = self.sessions_history(**arguments)
            return self._decorate_result(result)
        if name == "image_inspect":
            result = self.image_inspect(**arguments)
            return self._decorate_result(result)
        if name == "image_read":
            result = self.image_read(**arguments)
            return self._decorate_result(result)
        if name == "archive_extract":
            result = self.archive_extract(**arguments)
            return self._decorate_result(result)
        if name == "mail_extract_attachments":
            result = self.mail_extract_attachments(**arguments)
            return self._decorate_result(result)
        if name == "update_plan":
            result = self.update_plan(**arguments)
            return self._decorate_result(result)
        if name == "spawn_subagent":
            result = self.spawn_subagent(**arguments)
            return self._decorate_result(result)
        if name == "wait_subagents":
            result = self.wait_subagents(**arguments)
            return self._decorate_result(result)
        if name == "request_user_input":
            result = self.request_user_input(**arguments)
            return self._decorate_result(result)
        if name == "save_skill":
            result = self.save_skill(**arguments)
            return self._decorate_result(result)
        if name == "save_task":
            result = self.save_task(**arguments)
            return self._decorate_result(result)
        if name == "browser_open":
            result = self.browser_open(**arguments)
            return self._decorate_result(result)
        if name == "browser_click":
            result = self.browser_click(**arguments)
            return self._decorate_result(result)
        if name == "browser_type":
            result = self.browser_type(**arguments)
            return self._decorate_result(result)
        if name == "browser_wait":
            result = self.browser_wait(**arguments)
            return self._decorate_result(result)
        if name == "browser_scroll":
            result = self.browser_scroll(**arguments)
            return self._decorate_result(result)
        if name == "browser_snapshot":
            result = self.browser_snapshot(**arguments)
            return self._decorate_result(result)
        if name == "browser_screenshot":
            result = self.browser_screenshot(**arguments)
            return self._decorate_result(result)
        if name == "apply_patch":
            result = self.apply_patch(**arguments)
            return self._decorate_result(result)
        return self._decorate_result(
            {
                "ok": False,
                "error": {
                    "kind": "unknown_tool",
                    "tool": str(name or ""),
                    "message": f"Unknown tool: {name}",
                },
                "summary": f"unknown tool: {name}",
            }
        )

    @staticmethod
    def _normalize_public_tool_arguments(name: str, arguments: dict[str, Any] | None) -> dict[str, Any]:
        normalized = dict(arguments or {})
        tool_name = str(name or "").strip()
        if tool_name in {"image_read", "image_inspect"} and "path" not in normalized and "image_path" in normalized:
            normalized["path"] = normalized.pop("image_path")
        return normalized

    def exec_command(
        self,
        cmd: str,
        purpose: str = "",
        cwd: str = ".",
        yield_time_ms: int = 1000,
        max_output_chars: int = 12000,
        tty: bool = False,
        approval_token: str = "",
        tainted_approval_token: str = "",
    ) -> dict[str, Any]:
        if self._current_cancel_requested():
            return self._command_failure_result(
                command=cmd,
                cwd=cwd,
                error="Command execution was cancelled before it started.",
                returncode=130,
                error_kind="tool_cancelled",
                error_detail={"message": "The owning Agent run was cancelled."},
            )
        if not self._current_shell_allowed():
            return self._command_failure_result(command=cmd, cwd=cwd, error="Shell execution is not allowed for the active permission profile.")
        if str(cmd or "").strip() and is_dangerous_command(str(cmd)):
            message = "Command is blocked by the runtime boundary."
            return self._command_failure_result(
                command=cmd,
                cwd=cwd,
                error=message,
                stderr=message,
                error_kind="dangerous_command",
                error_detail={"message": message},
            )
        execution_mode = self._current_execution_mode()
        if execution_mode == "docker":
            return self._command_failure_result(command=cmd, cwd=cwd, error="Interactive exec_command sessions are only supported in host mode.")
        try:
            real_cwd = self._resolve_path(cwd)
        except Exception as exc:
            return self._command_failure_result(command=cmd, cwd=cwd, error=str(exc), returncode=1)
        if not real_cwd.exists() or not real_cwd.is_dir():
            return self._command_failure_result(command=cmd, cwd=cwd, error=f"Invalid cwd: {cwd}", returncode=1)
        skill_script_context = self._direct_enabled_skill_script_context(cmd, cwd=real_cwd)
        reserved_skill_error = self._reserved_skill_command_error(cmd, cwd=real_cwd)
        if reserved_skill_error:
            return self._command_failure_result(
                command=cmd,
                cwd=str(real_cwd),
                error=reserved_skill_error,
                stderr=reserved_skill_error,
                error_kind="reserved_skill_path",
                error_detail={
                    "message": reserved_skill_error,
                    "recovery": (
                        "Use apply_patch for a Team Skill path permitted by the active RuntimeBoundary. "
                        "Built-in Skills remain read-only; save_skill only creates or replaces SKILL.md."
                    ),
                },
            )
        compound_shell = self._is_compound_shell_command(cmd)
        compound_validation: dict[str, Any] = {}
        tainted_matches: list[dict[str, Any]] = []
        supply_chain_risks: list[dict[str, Any]] = []
        path_args_validated = False
        allow_supply_chain_commands = self._supply_chain_approval_allowed()
        if compound_shell:
            ok, detail = self._validate_compound_shell_command(cmd, real_cwd)
            if not ok:
                message = str(detail.get("reason") or detail.get("summary") or detail.get("message") or "Compound shell command could not be validated safely.")
                return self._command_failure_result(
                    command=cmd,
                    cwd=str(real_cwd),
                    error=message,
                    stderr=message,
                    error_kind=str(detail.get("error_kind") or "compound_shell_subcommand_rejected"),
                    error_detail=detail,
                )
            compound_validation = {
                "ok": True,
                "parsed_subcommands": [str(item) for item in list(detail.get("parsed_subcommands") or []) if str(item or "").strip()],
                "subcommands": list(detail.get("subcommands") or []),
            }
            argv = self._shell_argv_for_compound_command(cmd)
        else:
            raw_argv: list[str] = []
            try:
                raw_argv = shlex.split(str(cmd or "").strip())
            except Exception:
                raw_argv = []
            raw_tainted_matches = self._tainted_execution_matches_for_argv(raw_argv, cwd=real_cwd) if raw_argv else []
            argv, error = self._safe_split_command(
                cmd,
                for_session=True,
                allow_supply_chain_commands=allow_supply_chain_commands,
            )
            if error:
                if not raw_tainted_matches:
                    error_kind = (
                        "command_not_allowed"
                        if str(error).lower().startswith("command not allowed:")
                        else "invalid_arguments"
                    )
                    return self._command_failure_result(
                        command=cmd,
                        cwd=cwd,
                        error=error,
                        error_kind=error_kind,
                        error_detail={"message": error},
                    )
                supply_chain_block = blocked_supply_chain_command(raw_argv)
                if supply_chain_block is not None:
                    message = str(supply_chain_block.get("message") or "Command is blocked by supply-chain policy.")
                    return self._command_failure_result(
                        command=cmd,
                        cwd=str(real_cwd),
                        error=message,
                        stderr=message,
                        error_kind="blocked_supply_chain_command",
                        error_detail=supply_chain_block,
                    )
                tainted_path_error = self._tainted_file_path_validation_error(raw_tainted_matches)
                if tainted_path_error:
                    message = str(tainted_path_error.get("message") or "Tainted executable path is outside command allowed roots.")
                    return self._command_failure_result(
                        command=cmd,
                        cwd=str(real_cwd),
                        error=message,
                        stderr=message,
                        error_kind="command_path_outside_allowed_roots",
                        error_detail=tainted_path_error,
                    )
                path_error = self._command_path_validation_error(raw_argv, cwd=real_cwd)
                if path_error:
                    message = str(path_error.get("message") or "Command path argument is outside command allowed roots.")
                    return self._command_failure_result(
                        command=cmd,
                        cwd=str(real_cwd),
                        error=message,
                        stderr=message,
                        error_kind="command_path_outside_allowed_roots",
                        error_detail=path_error,
                    )
                argv = self._shell_argv_for_compound_command(cmd) if self._is_source_builtin_command(raw_argv[0]) else raw_argv
                tainted_matches = raw_tainted_matches
                path_args_validated = True
            if not path_args_validated:
                path_error = self._command_path_validation_error(argv, cwd=real_cwd)
                if path_error:
                    message = str(path_error.get("message") or "Command path argument is outside command allowed roots.")
                    return self._command_failure_result(
                        command=cmd,
                        cwd=str(real_cwd),
                        error=message,
                        stderr=message,
                        error_kind="command_path_outside_allowed_roots",
                        error_detail=path_error,
                    )

        if not tainted_matches:
            tainted_matches = self._tainted_execution_matches(
                argv=argv,
                cwd=real_cwd,
                compound_validation=compound_validation if compound_shell else None,
            )
        supply_chain_risks = self._supply_chain_risks(
            argv=argv,
            cwd=real_cwd,
            compound_validation=compound_validation if compound_shell else None,
        )
        external_side_effect_risks = self._external_side_effect_risks(
            argv=argv,
            cwd=real_cwd,
            compound_validation=compound_validation if compound_shell else None,
        )
        approval_risks = [
            *supply_chain_risks,
            *external_side_effect_risks,
            *self._tainted_execution_risks(tainted_matches),
        ]
        approval_token_value = str(approval_token or tainted_approval_token or "").strip()
        command_execution_approval_payload: dict[str, Any] = {}
        tainted_approval_payload: dict[str, Any] = {}
        if approval_risks:
            if self._current_subagent_read_only():
                message = (
                    "Read-only Subagents cannot request interactive command approval. "
                    "Use file/search tools or an existing read-oriented workspace script instead of inline code, "
                    "network-origin code, or external writes."
                )
                return self._command_failure_result(
                    command=cmd,
                    cwd=str(real_cwd),
                    error=message,
                    stderr=message,
                    error_kind="subagent_safe_alternative_required",
                    error_detail={
                        "retryability": "change_tool_or_arguments",
                        "recovery": (
                            "Choose a provenance-preserving alternative: read/search the file directly, "
                            "or run an existing read-oriented script/module under the allowed workspace roots."
                        ),
                    },
                )
            approved, approval_error = self._consume_command_execution_approval(
                token=approval_token_value,
                command=str(cmd or "").strip(),
                cwd=str(real_cwd),
                risks=approval_risks,
                tainted_files=tainted_matches,
            )
            if not approved:
                return self._command_execution_approval_failure_result(
                    command=cmd,
                    purpose=purpose,
                    cwd=str(real_cwd),
                    risks=[*supply_chain_risks, *external_side_effect_risks],
                    tainted_files=tainted_matches,
                    token_error="" if not approval_token_value else approval_error,
                )
            command_execution_approval_payload = {
                "approved": True,
                "approval_token": approval_token_value,
                "command": str(cmd or "").strip(),
                "purpose": str(purpose or "").strip()[:240],
                "cwd": str(real_cwd),
                "risks": [dict(item) for item in approval_risks],
                "files": self._approval_files_public_payload(tainted_matches),
            }
        if tainted_matches and command_execution_approval_payload:
            tainted_approval_payload = {
                "approved": True,
                "approval_token": approval_token_value,
                "files": self._approval_files_public_payload(tainted_matches),
            }

        try:
            proc = subprocess.Popen(
                argv,
                cwd=str(real_cwd),
                env=(
                    self._skill_script_environment(skill_script_context, cwd=real_cwd)
                    if skill_script_context
                    else None
                ),
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
            )
        except Exception as exc:
            return self._command_failure_result(command=cmd, cwd=str(real_cwd), error=f"exec_command failed: {exc}", returncode=1)

        session_id = next(self._command_session_ids)
        with self._command_sessions_lock:
            self._command_sessions[session_id] = {
                "proc": proc,
                "buffer": "",
                "cursor": 0,
                "cwd": str(real_cwd),
                "command": str(cmd or "").strip() if compound_shell else " ".join(shlex.quote(token) for token in argv),
                "run_id": self._current_run_id(),
                "execution_mode": execution_mode,
                "tty": bool(tty),
            }
            if command_execution_approval_payload:
                self._command_sessions[session_id]["command_execution_approved"] = dict(command_execution_approval_payload)
            if tainted_approval_payload:
                self._command_sessions[session_id]["tainted_execution_approved"] = dict(tainted_approval_payload)
            if compound_shell:
                self._command_sessions[session_id]["compound_shell"] = True
                self._command_sessions[session_id]["compound_validation"] = dict(compound_validation)
        self._spawn_command_reader(session_id, proc)
        if self._current_cancel_requested():
            self._cancel_command_sessions(run_id=self._current_run_id())
            return self._command_failure_result(
                command=cmd,
                cwd=str(real_cwd),
                error="Command execution was cancelled after it started.",
                returncode=130,
                error_kind="tool_cancelled",
                error_detail={"message": "The owning Agent run was cancelled."},
            )
        time.sleep(max(0.0, min(float(yield_time_ms) / 1000.0, 10.0)))
        payload = self._command_session_snapshot(session_id, max_output_chars=max_output_chars)
        if command_execution_approval_payload:
            payload["command_execution_approved"] = dict(command_execution_approval_payload)
        if tainted_approval_payload:
            payload["tainted_execution_approved"] = dict(tainted_approval_payload)
        payload.setdefault("summary", "command started")
        return payload

    def write_stdin(
        self,
        session_id: int,
        chars: str = "",
        yield_time_ms: int = 1000,
        max_output_chars: int = 12000,
    ) -> dict[str, Any]:
        try:
            normalized_session_id = int(session_id)
        except Exception:
            return {"ok": False, "error": "session_id must be an integer"}
        if self._current_cancel_requested():
            self._cancel_command_sessions(run_id=self._current_run_id())
            return self._command_failure_result(
                command="write_stdin",
                cwd=self._current_cwd_hint(),
                error="Command session was cancelled with its owning Agent run.",
                returncode=130,
                error_kind="tool_cancelled",
                error_detail={"session_id": normalized_session_id},
            )
        with self._command_sessions_lock:
            session = self._command_sessions.get(normalized_session_id)
            if session is None:
                return {"ok": False, "error": f"Unknown session_id: {normalized_session_id}"}
            proc = session.get("proc")
            stdin = getattr(proc, "stdin", None)
            if chars and proc.poll() is not None:
                return {"ok": False, "error": f"Session {normalized_session_id} is already completed"}
            if chars and stdin is not None:
                try:
                    stdin.write(str(chars).encode("utf-8"))
                    stdin.flush()
                except Exception as exc:
                    return {"ok": False, "error": f"write_stdin failed: {exc}"}
        time.sleep(max(0.0, min(float(yield_time_ms) / 1000.0, 10.0)))
        return self._command_session_snapshot(normalized_session_id, max_output_chars=max_output_chars)

    def update_plan(
        self,
        plan: list[dict[str, Any]] | None = None,
        explanation: str = "",
        **kwargs: Any,
    ) -> dict[str, Any]:
        if plan is None:
            plan = (
                kwargs.get("steps")
                or kwargs.get("items")
                or kwargs.get("tasks")
                or kwargs.get("plan_state")
            )
        if plan is None:
            return {
                "ok": False,
                "error": {
                    "kind": "bad_tool_arguments",
                    "tool": "update_plan",
                    "message": "update_plan requires a plan field.",
                    "received_keys": sorted(kwargs.keys()),
                },
                "summary": "missing plan argument",
            }
        if not isinstance(plan, list):
            return {
                "ok": False,
                "error": {
                    "kind": "bad_tool_arguments",
                    "tool": "update_plan",
                    "message": "plan must be a list.",
                    "received_type": type(plan).__name__,
                },
                "summary": "plan must be a list",
            }
        def _normalize_status(value: Any) -> str:
            raw = str(value or "").strip().lower()
            if raw in {"completed", "complete", "done", "success", "succeeded"}:
                return "completed"
            if raw in {"in_progress", "in-progress", "active", "doing", "working"}:
                return "in_progress"
            if raw in {"pending", "todo", "not_started", "not-started", ""}:
                return "pending"
            return raw

        def _looks_placeholder_step(value: str) -> bool:
            text = str(value or "").strip().lower()
            return bool(text) and bool(re.fullmatch(r"(?:step\s*)?\d+", text))

        normalized_plan: list[dict[str, Any]] = []
        for item in list(plan or []):
            if isinstance(item, str):
                item = {"step": item, "status": "pending"}
            if not isinstance(item, dict):
                continue
            raw_step = str(item.get("step") or item.get("title") or item.get("label") or "").strip()
            description = str(item.get("description") or item.get("content") or "").strip()
            step = description if description and (_looks_placeholder_step(raw_step) or not raw_step) else (raw_step or description)
            status = _normalize_status(item.get("status"))
            if not step:
                continue
            if status not in {"pending", "in_progress", "completed"}:
                return {
                    "ok": False,
                    "error": {
                        "kind": "bad_tool_arguments",
                        "tool": "update_plan",
                        "message": f"Invalid plan status: {status or '(empty)'}",
                    },
                    "summary": f"invalid plan status: {status or '(empty)'}",
                }
            normalized_item: dict[str, Any] = {
                "step": step,
                "status": status,
            }
            normalized_plan.append(normalized_item)
        if not normalized_plan:
            return {
                "ok": False,
                "error": {
                    "kind": "bad_tool_arguments",
                    "tool": "update_plan",
                    "message": "update_plan `plan` must be a non-empty list.",
                },
                "summary": "update_plan `plan` must be a non-empty list.",
            }
        in_progress_count = sum(1 for item in normalized_plan if item.get("status") == "in_progress")
        all_completed = all(item.get("status") == "completed" for item in normalized_plan)
        if not all_completed and in_progress_count == 0:
            first_open = next(
                (item for item in normalized_plan if item.get("status") != "completed"),
                None,
            )
            if first_open is not None:
                first_open["status"] = "in_progress"
                in_progress_count = 1
        if (all_completed and in_progress_count != 0) or (not all_completed and in_progress_count != 1):
            return {
                "ok": False,
                "error": {
                    "kind": "bad_tool_arguments",
                    "tool": "update_plan",
                    "message": (
                        "A non-complete plan must contain exactly one in_progress step; "
                        "a complete plan must contain only completed steps."
                    ),
                    "in_progress_count": in_progress_count,
                },
                "summary": "invalid plan state: expected exactly one in_progress step",
            }
        return {
            "ok": True,
            "plan": normalized_plan,
            "explanation": str(explanation or "").strip(),
            "summary": str(explanation or "").strip() or f"plan updated ({len(normalized_plan)} steps)",
        }

    def request_user_input(self, questions: list[dict[str, Any]]) -> dict[str, Any]:
        normalized_questions: list[dict[str, Any]] = []
        for item in list(questions or [])[:3]:
            if not isinstance(item, dict):
                continue
            header = str(item.get("header") or "").strip()
            question_id = str(item.get("id") or "").strip()
            question = str(item.get("question") or "").strip()
            raw_options = list(item.get("options") or [])
            options: list[dict[str, str]] = []
            for raw_option in raw_options[:3]:
                if not isinstance(raw_option, dict):
                    continue
                label = str(raw_option.get("label") or "").strip()
                description = str(raw_option.get("description") or "").strip()
                if not label or not description:
                    continue
                options.append({"label": label, "description": description})
            if not header or not question_id or not question or len(options) < 2:
                continue
            normalized_questions.append(
                {
                    "header": header[:12],
                    "id": question_id,
                    "question": question,
                    "options": options,
                }
            )
        if not normalized_questions:
            return {"ok": False, "error": "request_user_input requires at least one well-formed question"}
        return {
            "ok": True,
            "pending": True,
            "questions": normalized_questions,
            "summary": "user input required",
        }

    def spawn_subagent(
        self,
        task: str,
        role: str = "explorer",
        label: str = "",
    ) -> dict[str, Any]:
        runner = getattr(self._runtime_ctx, "subagent_runner", None)
        task_text = str(task or "").strip()
        if not task_text:
            return {
                "ok": False,
                "error_kind": "invalid_arguments",
                "error": "spawn_subagent requires a non-empty task",
            }
        if not callable(runner):
            return {
                "ok": False,
                "error_kind": "subagent_unavailable",
                "error": "Subagent execution is unavailable in this runtime.",
            }
        return dict(
            runner(
                task=task_text,
                role=str(role or "explorer").strip() or "explorer",
                label=str(label or "").strip(),
            )
            or {}
        )

    def wait_subagents(
        self,
        subagent_ids: list[str] | None = None,
        timeout_seconds: float = 30,
    ) -> dict[str, Any]:
        waiter = getattr(self._runtime_ctx, "subagent_waiter", None)
        if not callable(waiter):
            return {
                "ok": False,
                "error_kind": "subagent_unavailable",
                "error": "Subagent waiting is unavailable in this runtime.",
            }
        return dict(
            waiter(
                subagent_ids=[str(item) for item in list(subagent_ids or []) if str(item or "").strip()],
                timeout_seconds=max(0.0, min(300.0, float(timeout_seconds or 0.0))),
            )
            or {}
        )

    def save_skill(
        self,
        name: str,
        description: str,
        body: str,
        enabled: bool = True,
        overwrite: bool = False,
    ) -> dict[str, Any]:
        skill_name = str(name or "").strip()
        if not skill_name:
            return {
                "ok": False,
                "error": {
                    "kind": "invalid_skill_name",
                    "tool": "save_skill",
                    "message": "skill name is required",
                },
                "summary": "skill name is required",
            }
        writer = getattr(self._runtime_ctx, "skill_writer", None)
        if not callable(writer):
            return {
                "ok": False,
                "error": {
                    "kind": "skill_writer_unavailable",
                    "tool": "save_skill",
                    "message": "No skill writer is available for the current run.",
                },
                "summary": "skill writer unavailable",
            }
        try:
            payload = writer(
                name=skill_name,
                description=str(description or ""),
                body=str(body or ""),
                enabled=bool(enabled),
                overwrite=bool(overwrite),
            )
        except FileExistsError as exc:
            message = safe_error_message(exc)
            return {
                "ok": False,
                "error": {
                    "kind": "skill_already_exists",
                    "tool": "save_skill",
                    "message": message,
                },
                "summary": message,
            }
        except Exception as exc:
            message = safe_error_message(exc)
            return {
                "ok": False,
                "error": {
                    "kind": "skill_save_failed",
                    "tool": "save_skill",
                    "message": message,
                },
                "summary": message,
            }
        if not isinstance(payload, dict):
            return {
                "ok": False,
                "error": {
                    "kind": "invalid_skill_payload",
                    "tool": "save_skill",
                    "message": "Skill writer returned an invalid payload.",
                },
                "summary": "invalid skill payload",
            }
        return payload if "ok" in payload else {"ok": True, **payload}

    def save_task(
        self,
        title: str,
        goal: str,
        summary: str,
        task_id: str = "",
        progress: list[str] | None = None,
        next_steps: list[str] | None = None,
        decisions: list[str] | None = None,
        blockers: list[str] | None = None,
        artifacts: list[str] | None = None,
        status: str = "active",
    ) -> dict[str, Any]:
        writer = getattr(self._runtime_ctx, "task_writer", None)
        if not callable(writer):
            return {
                "ok": False,
                "error": {
                    "kind": "task_writer_unavailable",
                    "tool": "save_task",
                    "message": "No Task writer is available for the current run.",
                },
                "summary": "Task writer unavailable",
            }
        try:
            payload = writer(
                task_id=str(task_id or ""),
                title=str(title or ""),
                goal=str(goal or ""),
                summary=str(summary or ""),
                progress=[str(item) for item in list(progress or [])],
                next_steps=[str(item) for item in list(next_steps or [])],
                decisions=[str(item) for item in list(decisions or [])],
                blockers=[str(item) for item in list(blockers or [])],
                artifacts=[str(item) for item in list(artifacts or [])],
                status=str(status or "active"),
            )
        except FileNotFoundError as exc:
            message = safe_error_message(exc)
            return {
                "ok": False,
                "error": {"kind": "task_not_found", "tool": "save_task", "message": message},
                "summary": message,
            }
        except (PermissionError, ValueError) as exc:
            message = safe_error_message(exc)
            return {
                "ok": False,
                "error": {"kind": "invalid_task", "tool": "save_task", "message": message},
                "summary": message,
            }
        except Exception as exc:
            message = safe_error_message(exc)
            return {
                "ok": False,
                "error": {"kind": "task_save_failed", "tool": "save_task", "message": message},
                "summary": message,
            }
        if not isinstance(payload, dict):
            return {
                "ok": False,
                "error": {
                    "kind": "invalid_task_payload",
                    "tool": "save_task",
                    "message": "Task writer returned an invalid payload.",
                },
                "summary": "invalid Task payload",
            }
        return payload if "ok" in payload else {"ok": True, **payload}

    def web_search(self, query: str, max_results: int = 5, timeout_sec: int = 12) -> dict[str, Any]:
        result = self._web_search_impl(query=query, max_results=max_results, timeout_sec=timeout_sec)
        if not isinstance(result, dict):
            return {"ok": False, "error": "web_search failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "web_search")
        return payload

    def read_file(
        self,
        path: str = ".",
        start_char: int = 0,
        max_chars: int = 200000,
        start_line: int = 0,
        max_lines: int = 0,
    ) -> dict[str, Any]:
        try:
            real_path = self._resolve_source_path(path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if real_path.is_dir():
                return {
                    "ok": False,
                    "error": f"Path is a directory: {path}. Use list_dir instead.",
                }

            result = self._read_file_impl(
                path=path,
                start_char=start_char,
                max_chars=max_chars,
                start_line=start_line,
                max_lines=max_lines,
            )
            if isinstance(result, dict) and bool(result.get("ok")):
                payload = dict(result)
                payload.update(_path_payload(real_path, project_root=self._current_project_root(), cwd=Path(self._current_cwd_hint())))
                payload.setdefault("kind", "file")
                payload.setdefault("tool_name", "read_file")
                return payload
            return result
        except Exception as exc:
            return {"ok": False, "error": f"read_file failed: {exc}"}

    def list_dir(
        self,
        path: str = ".",
        max_entries: int = 200,
    ) -> dict[str, Any]:
        result = self._list_dir_impl(path=path, max_entries=max_entries)
        if not isinstance(result, dict):
            return {"ok": False, "error": "list_dir failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "list_dir")
        return payload

    def glob_file_search(
        self,
        pattern: str,
        path: str = ".",
        max_results: int = 200,
    ) -> dict[str, Any]:
        try:
            normalized_pattern = str(pattern or "").strip()
            if not normalized_pattern:
                return {"ok": False, "error": "pattern is empty"}
            real_root = self._resolve_path(path)
            if not real_root.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_root.is_dir():
                return {"ok": False, "error": f"Not a directory: {path}"}
            limit = max(1, min(500, int(max_results)))
            root_payload = _path_payload(real_root, project_root=self._current_project_root(), cwd=Path(self._current_cwd_hint()))
            all_matches = [candidate for candidate in sorted(real_root.glob(normalized_pattern)) if candidate.is_file()]
            if _is_broad_glob_pattern(normalized_pattern) and len(all_matches) > _BROAD_GLOB_GUIDANCE_THRESHOLD:
                return {
                    "ok": False,
                    "tool_name": "glob_file_search",
                    "error": {
                        "kind": "broad_glob_on_large_directory",
                        "message": "The glob pattern is too broad for this directory. Use list_dir to inspect subdirectories or use a narrower filename pattern.",
                    },
                    "path": root_payload["path"],
                    "root": root_payload["path"],
                    "root_ref": root_payload["root_ref"],
                    "resolved_root": str(real_root.resolve()),
                    "pattern": normalized_pattern,
                    "total_matches": len(all_matches),
                    "max_results": limit,
                    "truncated": True,
                    "suggested_next_steps": [
                        "Use list_dir on the current root to identify likely subdirectories.",
                        "Use a narrower glob such as '*runner*', '*.py', or '*target_name*'.",
                        "Use search_codebase with a concrete function, class, or filename keyword.",
                    ],
                    "summary": "glob pattern too broad for large directory",
                }
            matches: list[str] = []
            for candidate in all_matches[:limit]:
                matches.append(_display_model_path(candidate, project_root=self._current_project_root(), cwd=Path(self._current_cwd_hint())))
            truncated = len(all_matches) > limit
            return {
                "ok": True,
                "tool_name": "glob_file_search",
                "path": root_payload["path"],
                "root": root_payload["path"],
                "root_ref": root_payload["root_ref"],
                "resolved_root": str(real_root.resolve()),
                "pattern": normalized_pattern,
                "count": len(matches),
                "matches": matches,
                "total_matches": len(all_matches),
                "max_results": limit,
                "truncated": truncated,
                "summary": f"matched {len(matches)} files",
            }
        except Exception as exc:
            return {"ok": False, "error": f"glob_file_search failed: {exc}"}

    def search_contents_in_file(
        self,
        path: str,
        query: str,
        max_matches: int = 8,
        context_chars: int = 280,
    ) -> dict[str, Any]:
        result = self._search_contents_in_file_impl(
            path=path,
            query=query,
            max_matches=max_matches,
            context_chars=context_chars,
        )
        if not isinstance(result, dict):
            return {"ok": False, "error": "search_contents_in_file failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "search_contents_in_file")
        return payload

    def search_contents_in_file_multi(
        self,
        path: str,
        queries: list[str],
        per_query_max_matches: int = 3,
        context_chars: int = 280,
    ) -> dict[str, Any]:
        result = self._search_contents_in_file_multi_impl(
            path=path,
            queries=queries,
            per_query_max_matches=per_query_max_matches,
            context_chars=context_chars,
        )
        if not isinstance(result, dict):
            return {"ok": False, "error": "search_contents_in_file_multi failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "search_contents_in_file_multi")
        return payload

    def read_section(self, path: str, heading: str, max_chars: int = 12000) -> dict[str, Any]:
        result = self._read_section_impl(path=path, heading=heading, max_chars=max_chars)
        if not isinstance(result, dict):
            return {"ok": False, "error": "read_section failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "read_section")
        return payload

    def web_fetch(self, url: str, max_chars: int = 120000, timeout_sec: int = 12) -> dict[str, Any]:
        result = self._web_fetch_impl(url=url, max_chars=max_chars, timeout_sec=timeout_sec)
        if not isinstance(result, dict):
            return {"ok": False, "error": "web_fetch failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "web_fetch")
        return payload

    def web_download(
        self,
        url: str,
        dst_path: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        timeout_sec: int = 20,
        max_bytes: int = 52428800,
    ) -> dict[str, Any]:
        result = self._web_download_impl(
            url=url,
            dst_path=dst_path,
            overwrite=overwrite,
            create_dirs=create_dirs,
            timeout_sec=timeout_sec,
            max_bytes=max_bytes,
        )
        if not isinstance(result, dict):
            return {"ok": False, "error": "web_download failed: invalid result"}
        payload = dict(result)
        if bool(payload.get("ok")) and str(payload.get("path") or "").strip():
            try:
                taint = self._register_tainted_file(
                    Path(str(payload.get("path") or "")),
                    source_url=str(payload.get("url") or url or ""),
                    source_tool="web_download",
                    content_type=str(payload.get("content_type") or ""),
                )
                if taint:
                    payload["taint"] = {
                        "tainted": True,
                        "path": taint.get("path"),
                        "sha256": taint.get("sha256"),
                        "source_url": taint.get("source_url"),
                        "source_domain": taint.get("source_domain"),
                    }
            except Exception as exc:
                payload["taint_warning"] = f"failed to mark downloaded file as tainted: {exc}"
        payload.setdefault("tool_name", "web_download")
        return payload

    def sessions_list(self, limit: int = 20) -> dict[str, Any]:
        result = self._sessions_list_impl(max_sessions=limit)
        if not isinstance(result, dict):
            return {"ok": False, "error": "sessions_list failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "sessions_list")
        return payload

    def sessions_history(self, session_id: str, max_turns: int = 80) -> dict[str, Any]:
        result = self._sessions_history_impl(session_id=session_id, max_turns=max_turns)
        if not isinstance(result, dict):
            return {"ok": False, "error": "sessions_history failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "sessions_history")
        return payload

    def image_inspect(self, path: str = "", image_path: str = "") -> dict[str, Any]:
        resolved_path = str(path or image_path or "").strip()
        result = self._image_inspect_impl(path=resolved_path)
        if not isinstance(result, dict):
            return {"ok": False, "error": "image_inspect failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "image_inspect")
        return payload

    def image_read(
        self,
        path: str = "",
        prompt: str = "",
        max_output_chars: int = 12000,
        image_path: str = "",
    ) -> dict[str, Any]:
        resolved_path = str(path or image_path or "").strip()
        inspect_payload = self.image_inspect(path=resolved_path)
        if not bool(inspect_payload.get("ok")):
            error = str(inspect_payload.get("error") or "image inspect failed")
            return {
                "ok": False,
                "tool_name": "image_read",
                "path": resolved_path,
                "error": error,
                "model_capability_status": "read_error",
            }

        inspected_path = str(inspect_payload.get("path") or resolved_path)
        ocr_payload = self._perform_local_image_ocr(inspected_path, max_output_chars=max_output_chars)

        multimodal_payload: dict[str, Any] = {}
        if callable(self._image_read_handler):
            try:
                handler_result = self._image_read_handler(
                    path=inspected_path,
                    prompt=str(prompt or ""),
                    max_output_chars=max_output_chars,
                    model=self._current_model_hint(),
                )
            except Exception as exc:
                multimodal_payload = {
                    "ok": False,
                    "error": f"image_read failed: {exc}",
                    "model_capability_status": "read_error",
                    "visible_text": "",
                    "analysis": "",
                }
            else:
                if not isinstance(handler_result, dict):
                    multimodal_payload = {
                        "ok": False,
                        "error": "image_read failed: invalid result",
                        "model_capability_status": "read_error",
                        "visible_text": "",
                        "analysis": "",
                    }
                else:
                    multimodal_payload = dict(handler_result)

        ocr_text = str(ocr_payload.get("visible_text") or "").strip()
        multimodal_ok = bool(multimodal_payload.get("ok"))
        multimodal_text = str(multimodal_payload.get("visible_text") or "").strip()
        multimodal_analysis = str(multimodal_payload.get("analysis") or "").strip()

        warning_parts = [
            str(inspect_payload.get("warning") or "").strip(),
            str(ocr_payload.get("warning") or "").strip(),
            str(multimodal_payload.get("warning") or "").strip(),
        ]
        warning_text = "; ".join(item for item in warning_parts if item)

        payload = dict(inspect_payload)
        payload.setdefault("tool_name", "image_read")
        payload["engines_tried"] = list(ocr_payload.get("engines_tried") or [])
        payload["ocr_available"] = bool(ocr_payload.get("ocr_available"))
        payload["warning"] = warning_text or None
        payload["ocr_engine"] = str(ocr_payload.get("ocr_engine") or "").strip()
        payload["preprocess_notes"] = list(ocr_payload.get("preprocess_notes") or [])
        payload["effective_model"] = str(multimodal_payload.get("effective_model") or "").strip() or None

        if ocr_text and multimodal_ok:
            payload.update(
                {
                    "ok": True,
                    "visible_text": ocr_text,
                    "analysis": multimodal_analysis or "Extracted visible text via local OCR and supplemented the image analysis with the model.",
                    "model_capability_status": str(multimodal_payload.get("model_capability_status") or "ok"),
                    "read_strategy": "hybrid",
                    "fallback_reason": "",
                }
            )
            payload["summary"] = f"image_read · hybrid · {payload.get('ocr_engine') or 'ocr'}"
            payload["diagnostics"] = {
                "engines_tried": list(payload.get("engines_tried") or []),
                "ocr_available": bool(payload.get("ocr_available")),
                "ocr_engine": str(payload.get("ocr_engine") or ""),
                "preprocess_notes": list(payload.get("preprocess_notes") or []),
                "fallback_reason": "",
                "read_strategy": "hybrid",
                "model_capability_status": str(payload.get("model_capability_status") or ""),
                "visible_text_preview": self._short_preview(payload.get("visible_text"), limit=240),
                "analysis_preview": self._short_preview(payload.get("analysis"), limit=240),
                "warning": payload.get("warning"),
            }
            return payload

        if ocr_text:
            fallback_reason = ""
            if not callable(self._image_read_handler):
                fallback_reason = "no_runtime_image_reader"
            elif multimodal_payload:
                fallback_reason = str(multimodal_payload.get("model_capability_status") or "").strip() or "multimodal_unavailable"
            payload.update(
                {
                    "ok": True,
                    "visible_text": ocr_text,
                    "analysis": multimodal_analysis or "Extracted visible text from the image using local OCR.",
                    "model_capability_status": str(multimodal_payload.get("model_capability_status") or ("not_invoked" if not callable(self._image_read_handler) else "read_error")),
                    "read_strategy": "ocr_only",
                    "fallback_reason": fallback_reason,
                }
            )
            engine_label = str(payload.get("ocr_engine") or "ocr").strip()
            payload["summary"] = f"image_read · ocr_only · {engine_label}"
            payload["diagnostics"] = {
                "engines_tried": list(payload.get("engines_tried") or []),
                "ocr_available": bool(payload.get("ocr_available")),
                "ocr_engine": engine_label,
                "preprocess_notes": list(payload.get("preprocess_notes") or []),
                "fallback_reason": fallback_reason,
                "read_strategy": "ocr_only",
                "model_capability_status": str(payload.get("model_capability_status") or ""),
                "visible_text_preview": self._short_preview(payload.get("visible_text"), limit=240),
                "analysis_preview": self._short_preview(payload.get("analysis"), limit=240),
                "warning": payload.get("warning"),
            }
            return payload

        if multimodal_ok:
            payload.update(
                {
                    "ok": True,
                    "visible_text": multimodal_text,
                    "analysis": multimodal_analysis,
                    "model_capability_status": str(multimodal_payload.get("model_capability_status") or "ok"),
                    "read_strategy": "multimodal_only",
                    "fallback_reason": str(ocr_payload.get("error") or "").strip(),
                }
            )
            payload["summary"] = "image_read · multimodal_only"
            payload["diagnostics"] = {
                "engines_tried": list(payload.get("engines_tried") or []),
                "ocr_available": bool(payload.get("ocr_available")),
                "ocr_engine": str(payload.get("ocr_engine") or ""),
                "preprocess_notes": list(payload.get("preprocess_notes") or []),
                "fallback_reason": str(payload.get("fallback_reason") or ""),
                "read_strategy": "multimodal_only",
                "model_capability_status": str(payload.get("model_capability_status") or ""),
                "visible_text_preview": self._short_preview(payload.get("visible_text"), limit=240),
                "analysis_preview": self._short_preview(payload.get("analysis"), limit=240),
                "warning": payload.get("warning"),
            }
            return payload

        if not bool(ocr_payload.get("ocr_available")):
            ocr_reason = str(ocr_payload.get("warning") or ocr_payload.get("error") or "").strip()
            error_text = ocr_reason or "local OCR is unavailable"
            if not callable(self._image_read_handler) and not error_text:
                error_text = "local OCR is unavailable and no runtime image reader is configured"
            elif str(multimodal_payload.get("error") or "").strip():
                error_text = str(multimodal_payload.get("error") or "").strip()
            payload.update(
                {
                    "ok": False,
                    "visible_text": "",
                    "analysis": "",
                    "error": error_text,
                    "model_capability_status": str(multimodal_payload.get("model_capability_status") or ("not_invoked" if not callable(self._image_read_handler) else "read_error")),
                    "read_strategy": "",
                    "fallback_reason": "ocr_unavailable",
                }
            )
            payload["summary"] = "image_read · ocr_unavailable"
            payload["diagnostics"] = {
                "engines_tried": list(payload.get("engines_tried") or []),
                "ocr_available": bool(payload.get("ocr_available")),
                "ocr_engine": str(payload.get("ocr_engine") or ""),
                "preprocess_notes": list(payload.get("preprocess_notes") or []),
                "fallback_reason": "ocr_unavailable",
                "read_strategy": "",
                "model_capability_status": str(payload.get("model_capability_status") or ""),
                "visible_text_preview": "",
                "analysis_preview": "",
                "warning": payload.get("warning"),
                "error": payload.get("error"),
            }
            return payload

        payload.update(
            {
                "ok": False,
                "visible_text": "",
                "analysis": "",
                "error": str(multimodal_payload.get("error") or ocr_payload.get("error") or "image_read failed").strip(),
                "model_capability_status": str(multimodal_payload.get("model_capability_status") or ("not_invoked" if not callable(self._image_read_handler) else "read_error")),
                "read_strategy": "",
                "fallback_reason": str(ocr_payload.get("error") or "").strip() or "no_readable_text_detected",
            }
        )
        payload["summary"] = "image_read · no_readable_text_detected"
        payload["diagnostics"] = {
            "engines_tried": list(payload.get("engines_tried") or []),
            "ocr_available": bool(payload.get("ocr_available")),
            "ocr_engine": str(payload.get("ocr_engine") or ""),
            "preprocess_notes": list(payload.get("preprocess_notes") or []),
            "fallback_reason": str(payload.get("fallback_reason") or ""),
            "read_strategy": "",
            "model_capability_status": str(payload.get("model_capability_status") or ""),
            "visible_text_preview": "",
            "analysis_preview": "",
            "warning": payload.get("warning"),
            "error": payload.get("error"),
        }
        return payload

    def archive_extract(
        self,
        zip_path: str,
        dst_dir: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        max_entries: int = 20000,
        max_total_bytes: int = 524288000,
    ) -> dict[str, Any]:
        result = self._archive_extract_impl(
            zip_path=zip_path,
            dst_dir=dst_dir,
            overwrite=overwrite,
            create_dirs=create_dirs,
            max_entries=max_entries,
            max_total_bytes=max_total_bytes,
        )
        if not isinstance(result, dict):
            return {"ok": False, "error": "archive_extract failed: invalid result"}
        payload = dict(result)
        if bool(payload.get("ok")):
            payload["tainted_files"] = self._mark_extracted_files_tainted_from_parent(
                source_path=zip_path,
                extracted_files=list(payload.get("entries") or []),
            )
        payload.setdefault("tool_name", "archive_extract")
        return payload

    def mail_extract_attachments(
        self,
        msg_path: str,
        dst_dir: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        max_attachments: int = 500,
        max_total_bytes: int = 524288000,
    ) -> dict[str, Any]:
        result = self._mail_extract_attachments_impl(
            msg_path=msg_path,
            dst_dir=dst_dir,
            overwrite=overwrite,
            create_dirs=create_dirs,
            max_attachments=max_attachments,
            max_total_bytes=max_total_bytes,
        )
        if not isinstance(result, dict):
            return {"ok": False, "error": "mail_extract_attachments failed: invalid result"}
        payload = dict(result)
        payload.setdefault("tool_name", "mail_extract_attachments")
        return payload

    def _list_dir_impl(self, path: str = ".", max_entries: int = 200) -> dict[str, Any]:
        try:
            real_path = self._resolve_source_path(path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_dir():
                return {"ok": False, "error": f"Not a directory: {path}"}

            limit = max(1, min(500, int(max_entries)))
            ordered = sorted(real_path.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
            entries = []
            for child in ordered[:limit]:
                child_path = _display_model_path(child, project_root=self._current_project_root(), cwd=real_path)
                entries.append(
                    {
                        "name": child.name,
                        "path": child_path,
                        "type": "symlink" if child.is_symlink() else ("directory" if child.is_dir() else "file"),
                        "kind": "symlink" if child.is_symlink() else ("directory" if child.is_dir() else "file"),
                        "is_dir": child.is_dir(),
                        "size": child.stat().st_size if child.is_file() else None,
                    }
                )
            truncated = len(ordered) > limit
            path_payload = _path_payload(real_path, project_root=self._current_project_root(), cwd=Path(self._current_cwd_hint()))
            return {
                "ok": True,
                "path": path_payload["path"],
                "root_ref": path_payload["root_ref"],
                "resolved_path": path_payload["resolved_path"],
                "entries": entries,
                "entry_count": len(entries),
                "total_entries": len(ordered),
                "max_entries": limit,
                "truncated": truncated,
                "has_more": truncated,
                "source_format": "directory_listing",
                "summary": f"listed {len(entries)} entries",
            }
        except Exception as exc:
            return {"ok": False, "error": f"list_dir failed: {exc}"}

    def _sessions_list_impl(self, max_sessions: int = 20) -> dict[str, Any]:
        try:
            limit = max(1, min(200, int(max_sessions)))
            current_project_id = self._current_project_id()
            rows: list[dict[str, Any]] = []
            files = sorted(
                self.config.sessions_dir.glob("*.json"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
            for path in files:
                try:
                    payload = json.loads(path.read_text(encoding="utf-8"))
                except Exception:
                    continue
                if current_project_id and str(payload.get("project_id") or "").strip() != current_project_id:
                    continue
                sid = str(payload.get("id") or path.stem)
                turns = payload.get("turns", [])
                if not isinstance(turns, list):
                    turns = []
                title = "新会话"
                preview = ""
                for turn in turns:
                    if not isinstance(turn, dict):
                        continue
                    role = str(turn.get("role") or "")
                    text = str(turn.get("text") or "").strip()
                    if role == "user" and text:
                        title = text.replace("\n", " ")[:60]
                        break
                if turns:
                    last = turns[-1]
                    if isinstance(last, dict):
                        preview = str(last.get("text") or "").replace("\n", " ").strip()[:120]
                rows.append(
                    {
                        "session_id": sid,
                        "title": title,
                        "preview": preview,
                        "turn_count": len(turns),
                        "project_id": str(payload.get("project_id") or ""),
                        "project_title": str(payload.get("project_title") or ""),
                        "project_root": str(payload.get("project_root") or ""),
                        "git_branch": str(payload.get("git_branch") or ""),
                        "cwd": str(payload.get("cwd") or ""),
                        "updated_at": str(payload.get("updated_at") or ""),
                        "created_at": str(payload.get("created_at") or ""),
                    }
                )
                if len(rows) >= limit:
                    break
            return {"ok": True, "count": len(rows), "sessions": rows}
        except Exception as exc:
            return {"ok": False, "error": f"sessions_list failed: {exc}"}

    def _sessions_history_impl(self, session_id: str, max_turns: int = 80) -> dict[str, Any]:
        sid = str(session_id or "").strip()
        if not sid:
            return {"ok": False, "error": "session_id cannot be empty"}
        if "/" in sid or "\\" in sid or ".." in sid:
            return {"ok": False, "error": "Invalid session_id"}
        try:
            session_path = (self.config.sessions_dir / f"{sid}.json").resolve()
            if not _is_within(session_path, self.config.sessions_dir):
                return {"ok": False, "error": "Invalid session path"}
            if not session_path.exists():
                return {"ok": False, "error": f"Session not found: {sid}"}
            payload = json.loads(session_path.read_text(encoding="utf-8"))
            turns = payload.get("turns", [])
            if not isinstance(turns, list):
                turns = []
            keep = max(1, min(800, int(max_turns)))
            sliced = turns[-keep:]
            trimmed_turns: list[dict[str, str]] = []
            for turn in sliced:
                if not isinstance(turn, dict):
                    continue
                trimmed_turns.append(
                    {
                        "role": str(turn.get("role") or "user"),
                        "text": str(turn.get("text") or ""),
                        "created_at": str(turn.get("created_at") or ""),
                    }
                )
            return {
                "ok": True,
                "session_id": sid,
                "project_id": str(payload.get("project_id") or ""),
                "project_title": str(payload.get("project_title") or ""),
                "project_root": str(payload.get("project_root") or ""),
                "cwd": str(payload.get("cwd") or ""),
                "summary": str(payload.get("summary") or ""),
                "turn_count": len(turns),
                "turns": trimmed_turns,
            }
        except Exception as exc:
            return {"ok": False, "error": f"sessions_history failed: {exc}"}

    def _browser_session_id(self) -> str:
        return self._current_session_id()

    def browser_open(self, url: str, timeout_ms: int = 20000) -> dict[str, Any]:
        return self._browser_manager.open(
            session_id=self._browser_session_id(),
            url=str(url or "").strip(),
            timeout_ms=timeout_ms,
        )

    def browser_click(self, selector: str, timeout_ms: int = 12000) -> dict[str, Any]:
        return self._browser_manager.click(
            session_id=self._browser_session_id(),
            selector=str(selector or "").strip(),
            timeout_ms=timeout_ms,
        )

    def browser_type(
        self,
        selector: str,
        text: str,
        submit: bool = False,
        clear: bool = True,
        timeout_ms: int = 12000,
    ) -> dict[str, Any]:
        return self._browser_manager.type(
            session_id=self._browser_session_id(),
            selector=str(selector or "").strip(),
            text=str(text or ""),
            submit=bool(submit),
            clear=bool(clear),
            timeout_ms=timeout_ms,
        )

    def browser_wait(self, selector: str = "", timeout_ms: int = 5000, state: str = "visible") -> dict[str, Any]:
        return self._browser_manager.wait(
            session_id=self._browser_session_id(),
            selector=str(selector or "").strip(),
            timeout_ms=timeout_ms,
            state=str(state or "visible"),
        )

    def browser_scroll(
        self,
        direction: str = "down",
        amount: int = 900,
        selector: str = "",
        timeout_ms: int = 5000,
    ) -> dict[str, Any]:
        return self._browser_manager.scroll(
            session_id=self._browser_session_id(),
            direction=str(direction or "down").strip(),
            amount=amount,
            selector=str(selector or "").strip(),
            timeout_ms=timeout_ms,
        )

    def browser_snapshot(self, max_chars: int = 12000) -> dict[str, Any]:
        return self._browser_manager.snapshot(
            session_id=self._browser_session_id(),
            max_chars=max_chars,
        )

    def browser_screenshot(self, path: str = "", full_page: bool = True) -> dict[str, Any]:
        try:
            target = (
                self._resolve_path(path)
                if str(path or "").strip()
                else self._browser_manager.default_screenshot_path(session_id=self._browser_session_id())
            )
            return self._browser_manager.screenshot(
                session_id=self._browser_session_id(),
                target_path=target,
                full_page=bool(full_page),
            )
        except Exception as exc:
            return {"ok": False, "error": f"browser_screenshot failed: {exc}"}

    def _image_inspect_impl(self, path: str) -> dict[str, Any]:
        try:
            real_path = self._resolve_source_path(path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}
            with Image.open(real_path) as image:
                width, height = image.size
                image_format = str(image.format or "")
                return {
                    "ok": True,
                    "path": str(real_path),
                    "format": image_format,
                    "mime": str(Image.MIME.get(image_format, "") or ""),
                    "mode": str(image.mode or ""),
                    "width": int(width),
                    "height": int(height),
                    "summary": f"{real_path.name} · {width}x{height} · {image_format or 'unknown'}",
                }
        except Exception as exc:
            return {"ok": False, "error": f"image_inspect failed: {exc}"}

    def apply_patch(self, patch: str, cwd: str = ".", check: bool = False) -> dict[str, Any]:
        patch_text = str(patch or "")
        if not patch_text.strip():
            return {"ok": False, "error": "patch cannot be empty"}
        try:
            operations = _parse_workspace_patch(patch_text)
            real_cwd = self._resolve_path(cwd)
            if not real_cwd.exists() or not real_cwd.is_dir():
                return {"ok": False, "error": f"Invalid cwd: {cwd}"}
            files: list[str] = []
            pending_writes: list[tuple[Path, str]] = []
            pending_deletes: list[Path] = []
            for op in operations:
                op_type = str(op.get("op") or "")
                raw_path = str(op.get("path") or "").strip()
                if op_type == "add":
                    target = _resolve_workspace_path(
                        self.config,
                        raw_path,
                        workspace_root=real_cwd,
                        access_roots=self._current_access_roots(),
                        allow_any_path=self._unrestricted_path_access(),
                    )
                    if target.exists():
                        return {
                            "ok": False,
                            "error": {
                                "kind": "file_already_exists",
                                "operation": "add",
                                "message": f"Cannot add file because it already exists: {raw_path}",
                                "recovery": (
                                    "Read the existing file if needed, then retry with "
                                    f"*** Update File: {raw_path}; do not repeat *** Add File."
                                ),
                            },
                            "files": files,
                            "summary": "apply_patch Add File rejected because the target already exists",
                        }
                    pending_writes.append((target, str(op.get("content") or "")))
                    files.append(str(target))
                    continue
                if op_type == "delete":
                    target = _resolve_workspace_path(
                        self.config,
                        raw_path,
                        workspace_root=real_cwd,
                        access_roots=self._current_access_roots(),
                        allow_any_path=self._unrestricted_path_access(),
                    )
                    if not target.exists():
                        return {"ok": False, "error": f"File not found: {raw_path}", "files": files}
                    pending_deletes.append(target)
                    files.append(str(target))
                    continue
                if op_type == "update":
                    source = _resolve_workspace_path(
                        self.config,
                        raw_path,
                        workspace_root=real_cwd,
                        access_roots=self._current_access_roots(),
                        allow_any_path=self._unrestricted_path_access(),
                    )
                    if not source.exists():
                        return {"ok": False, "error": f"File not found: {raw_path}", "files": files}
                    original_text = source.read_text(encoding="utf-8")
                    updated_text = self._apply_update_hunks(source, original_text, list(op.get("hunks") or []))
                    target_raw = str(op.get("move_to") or raw_path).strip() or raw_path
                    target = _resolve_workspace_path(
                        self.config,
                        target_raw,
                        workspace_root=real_cwd,
                        access_roots=self._current_access_roots(),
                        allow_any_path=self._unrestricted_path_access(),
                    )
                    pending_writes.append((target, updated_text))
                    files.append(str(target))
                    if target != source:
                        pending_deletes.append(source)
                    continue
                return {"ok": False, "error": f"Unsupported patch op: {op_type}", "files": files}

            for target, _content in pending_writes:
                reserved_error = self._reserved_skill_write_error(target)
                if reserved_error:
                    return {
                        "ok": False,
                        "error": {
                            "kind": "reserved_skill_path",
                            "tool": "apply_patch",
                            "message": reserved_error,
                            "recovery": self._reserved_skill_write_recovery(target),
                        },
                        "files": files,
                        "summary": reserved_error,
                    }
            for target in pending_deletes:
                reserved_error = self._reserved_skill_write_error(target)
                if reserved_error:
                    return {
                        "ok": False,
                        "error": {
                            "kind": "reserved_skill_path",
                            "tool": "apply_patch",
                            "message": reserved_error,
                            "recovery": self._reserved_skill_write_recovery(target),
                        },
                        "files": files,
                        "summary": reserved_error,
                    }

            if check:
                return {
                    "ok": True,
                    "cwd": str(real_cwd),
                    "files": files,
                    "summary": "patch validated",
                }

            for target, content in pending_writes:
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_text(content, encoding="utf-8")
            for target in pending_deletes:
                if target.exists():
                    target.unlink()
            return {
                "ok": True,
                "cwd": str(real_cwd),
                "files": files,
                "summary": "patch applied",
            }
        except Exception as exc:
            return {"ok": False, "error": f"apply_patch failed: {exc}"}

    def _read_file_impl(
        self,
        path: str,
        start_char: int = 0,
        max_chars: int = 200000,
        start_line: int = 0,
        max_lines: int = 0,
    ) -> dict[str, Any]:
        try:
            real_path = self._resolve_source_path(path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}
            suffix = real_path.suffix.lower()
            locale_hint = self._current_locale_hint()
            source_format = "text_utf8"
            full_text = ""
            msg_payload: dict[str, Any] | None = None

            # For office/binary documents, try structured extraction first
            # so users can "download then read" in one flow.
            if suffix == ".pdf":
                source_format = "pdf_text_extracted"
                try:
                    full_text = extract_pdf_text_from_path(real_path, max_chars=1_000_000)
                except Exception as exc:
                    full_text = f"[文档解析失败: {exc}]"
            elif suffix in {
                ".docx",
                ".msg",
                ".xlsx",
                ".xlsm",
                ".xltx",
                ".xltm",
                ".xls",
                ".pptx",
                ".pptm",
                ".ppt",
                ".atom",
                ".rss",
                ".xml",
            }:
                if suffix == ".msg":
                    from app.attachments import extract_outlook_msg_payload  # lazy import

                    msg_payload = self._extract_outlook_msg_payload_compat(
                        extract_outlook_msg_payload,
                        str(real_path),
                        max_chars=1_000_000,
                        locale=locale_hint,
                    )
                    full_text = str(msg_payload.get("content") or "")
                else:
                    from app.attachments import extract_document_text  # lazy import

                    extracted = extract_document_text(
                        str(real_path),
                        max_chars=1_000_000,
                        locale=locale_hint,
                    ) or ""
                    full_text = extracted
                if suffix == ".docx":
                    source_format = "docx_text_extracted"
                elif suffix == ".msg":
                    source_format = "msg_text_extracted"
                elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
                    source_format = "xlsx_text_extracted"
                elif suffix in {".pptx", ".pptm", ".ppt"}:
                    source_format = "pptx_text_extracted"
                elif suffix in {".atom", ".rss", ".xml"}:
                    source_format = "xml_text_extracted"
            else:
                # Content sniffing: handle docs saved without normal suffix.
                try:
                    with real_path.open("rb") as fp:
                        sniff = fp.read(512 * 1024)
                    head = sniff[:8]
                except Exception:
                    sniff = b""
                    head = b""
                if head.startswith(b"%PDF-"):
                    source_format = "pdf_text_extracted"
                    try:
                        full_text = extract_pdf_text_from_path(real_path, max_chars=1_000_000)
                    except Exception as exc:
                        full_text = f"[文档解析失败: {exc}]"
                elif head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
                    from app.attachments import extract_document_text, extract_outlook_msg_payload, looks_like_outlook_msg_bytes  # lazy import

                    if looks_like_outlook_msg_bytes(sniff):
                        source_format = "msg_text_extracted"
                        msg_payload = self._extract_outlook_msg_payload_compat(
                            extract_outlook_msg_payload,
                            str(real_path),
                            max_chars=1_000_000,
                            locale=locale_hint,
                        )
                        full_text = str(msg_payload.get("content") or "")
                    else:
                        full_text = real_path.read_text(encoding="utf-8", errors="ignore")
                elif head.startswith(b"PK\x03\x04"):
                    from app.attachments import extract_document_text, looks_like_pptx_file, looks_like_xlsx_file  # lazy import

                    if looks_like_xlsx_file(real_path):
                        source_format = "xlsx_text_extracted"
                        full_text = extract_document_text(
                            str(real_path),
                            max_chars=1_000_000,
                            locale=locale_hint,
                        ) or ""
                    elif looks_like_pptx_file(real_path):
                        source_format = "pptx_text_extracted"
                        full_text = extract_document_text(
                            str(real_path),
                            max_chars=1_000_000,
                            locale=locale_hint,
                        ) or ""
                    else:
                        full_text = real_path.read_text(encoding="utf-8", errors="ignore")
                else:
                    full_text = real_path.read_text(encoding="utf-8", errors="ignore")

            total_length = len(full_text)
            limit = max(128, min(1_000_000, int(max_chars)))
            line_start = max(0, int(start_line))
            line_limit = max(0, int(max_lines))

            if line_start > 0 or line_limit > 0:
                lines = full_text.splitlines()
                total_lines = len(lines)
                first_line = max(1, line_start if line_start > 0 else 1)
                if first_line > total_lines:
                    first_line = total_lines + 1
                start_idx = max(0, first_line - 1)
                take_lines = max(1, min(200_000, line_limit)) if line_limit > 0 else 400
                end_idx = min(total_lines, start_idx + take_lines)
                chunk_lines = lines[start_idx:end_idx]
                text = "\n".join(chunk_lines)

                if len(text) > limit:
                    text = text[:limit]
                    truncated = True
                else:
                    truncated = end_idx < total_lines

                start_char_calc = sum(len(line) + 1 for line in lines[:start_idx])
                end_char_calc = start_char_calc + len(text)
                payload = {
                    "ok": True,
                    "path": str(real_path),
                    "content": text,
                    "length": len(text),
                    "start_char": start_char_calc,
                    "end_char": end_char_calc,
                    "total_length": total_length,
                    "truncated": truncated,
                    "has_more": truncated,
                    "line_mode": True,
                    "start_line": first_line if total_lines else 0,
                    "end_line": min(total_lines, end_idx),
                    "total_lines": total_lines,
                    "source_format": source_format,
                }
                if source_format == "msg_text_extracted" and isinstance(msg_payload, dict):
                    payload["email_meta"] = dict(msg_payload.get("email_meta") or {})
                    payload["attachment_list"] = list(msg_payload.get("attachment_list") or [])
                return payload

            start = max(0, int(start_char))
            if start > total_length:
                start = total_length
            end = min(total_length, start + limit)
            text = full_text[start:end]
            truncated = end < total_length
            payload = {
                "ok": True,
                "path": str(real_path),
                "content": text,
                "length": len(text),
                "start_char": start,
                "end_char": end,
                "total_length": total_length,
                "truncated": truncated,
                "has_more": truncated,
                "source_format": source_format,
            }
            if source_format == "msg_text_extracted" and isinstance(msg_payload, dict):
                payload["email_meta"] = dict(msg_payload.get("email_meta") or {})
                payload["attachment_list"] = list(msg_payload.get("attachment_list") or [])
            return payload
        except Exception as exc:
            return {"ok": False, "error": f"read_file failed: {exc}"}

    def _search_contents_in_file_impl(
        self,
        path: str,
        query: str,
        max_matches: int = 8,
        context_chars: int = 280,
    ) -> dict[str, Any]:
        try:
            normalized_query = _normalize_search_query(query)
            if not normalized_query:
                return {"ok": False, "error": "query is empty"}

            variants = _expand_search_variants(normalized_query)
            limit = max(1, min(20, int(max_matches)))
            window = max(40, min(2000, int(context_chars)))
            matches: list[dict[str, Any]] = []

            real_path = self._resolve_source_path(path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}

            if _looks_like_pdf_path(real_path):
                pages = extract_pdf_page_texts_from_path(real_path)
                for variant in variants:
                    pattern = _build_search_pattern(variant)
                    if pattern is None:
                        continue
                    for page_num, body in pages:
                        for found in pattern.finditer(body):
                            span = found.span()
                            start = max(0, span[0] - window)
                            end = min(len(body), span[1] + window)
                            matches.append(
                                {
                                    "query_variant": variant,
                                    "matched_text": found.group(0),
                                    "start_char": span[0],
                                    "end_char": span[1],
                                    "page_hint": page_num,
                                    "context": body[start:end].strip(),
                                    "read_hint": {
                                        "page_hint": page_num,
                                        "start_char": max(0, span[0] - 2000),
                                        "max_chars": 6000,
                                    },
                                }
                            )
                            if len(matches) >= limit:
                                break
                        if len(matches) >= limit:
                            break
                    if len(matches) >= limit:
                        break

                return {
                    "ok": True,
                    "path": str(real_path),
                    "source_format": "pdf_text_extracted",
                    "query": normalized_query,
                    "searched_variants": variants,
                    "match_count": len(matches),
                    "matches": matches,
                    "note": (
                        "Search was run page-by-page over extracted PDF text. "
                        "If match_count=0, only conclude that the current extracted PDF text did not show a hit."
                    ),
                }

            base = self._read_file_impl(path=path, start_char=0, max_chars=1_000_000)
            if not bool(base.get("ok")):
                return base

            text = str(base.get("content") or "")
            seen_spans: list[tuple[int, int]] = []
            for variant in variants:
                pattern = _build_search_pattern(variant)
                if pattern is None:
                    continue
                for found in pattern.finditer(text):
                    span = found.span()
                    if any(_spans_overlap(span, prior) for prior in seen_spans):
                        continue
                    seen_spans.append(span)

                    start = max(0, span[0] - window)
                    end = min(len(text), span[1] + window)
                    page_hint = _page_hint_for_offset(text, span[0])
                    matches.append(
                        {
                            "query_variant": variant,
                            "matched_text": found.group(0),
                            "start_char": span[0],
                            "end_char": span[1],
                            "page_hint": page_hint,
                            "context": text[start:end].strip(),
                            "read_hint": {
                                "start_char": max(0, span[0] - 2000),
                                "max_chars": 6000,
                            },
                        }
                    )
                    if len(matches) >= limit:
                        break
                if len(matches) >= limit:
                    break

            return {
                "ok": True,
                "path": str(base.get("path") or real_path),
                "source_format": base.get("source_format") or "text_utf8",
                "query": normalized_query,
                "searched_variants": variants,
                "match_count": len(matches),
                "matches": matches,
                "note": (
                    "Search was run over extracted document text. "
                    "If match_count=0, only conclude that the current extracted text did not show a hit."
                ),
            }
        except Exception as exc:
            return {"ok": False, "error": f"search_contents_in_file failed: {exc}"}

    def _search_contents_in_file_multi_impl(
        self,
        path: str,
        queries: list[str],
        per_query_max_matches: int = 3,
        context_chars: int = 280,
    ) -> dict[str, Any]:
        try:
            cleaned_queries = [_normalize_search_query(item) for item in (queries or []) if str(item or "").strip()]
            if not cleaned_queries:
                return {"ok": False, "error": "queries is empty"}

            merged: list[dict[str, Any]] = []
            seen: set[tuple[Any, ...]] = set()
            for query in cleaned_queries[:20]:
                result = self._search_contents_in_file_impl(
                    path=path,
                    query=query,
                    max_matches=max(1, min(10, int(per_query_max_matches))),
                    context_chars=context_chars,
                )
                if not bool(result.get("ok")):
                    return result
                for match in result.get("matches") or []:
                    if not isinstance(match, dict):
                        continue
                    key = (
                        match.get("page_hint"),
                        match.get("start_char"),
                        match.get("end_char"),
                        str(match.get("matched_text") or ""),
                    )
                    if key in seen:
                        continue
                    seen.add(key)
                    merged.append(match)

            return {
                "ok": True,
                "path": str(self._resolve_source_path(path)),
                "queries": cleaned_queries[:20],
                "match_count": len(merged),
                "matches": merged,
            }
        except Exception as exc:
            return {"ok": False, "error": f"search_contents_in_file_multi failed: {exc}"}

    def _read_section_impl(self, path: str, heading: str, max_chars: int = 12000) -> dict[str, Any]:
        try:
            real_path = self._resolve_source_path(path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}

            limit = max(512, min(50000, int(max_chars)))
            if _looks_like_pdf_path(real_path):
                pages = extract_pdf_page_texts_from_path(real_path)
                headings = extract_heading_entries_from_pages(pages, max_headings=1000)
                section = _extract_section_from_pdf_pages(pages, headings, heading, limit)
                if not bool(section.get("ok")):
                    return section
                return {
                    "ok": True,
                    "path": str(real_path),
                    "matched_heading": section.get("matched_heading"),
                    "matched_section": section.get("matched_section"),
                    "page_start": section.get("page_start"),
                    "page_end": section.get("page_end"),
                    "content": section.get("content"),
                }

            base = self._read_file_impl(path=path, start_char=0, max_chars=1_000_000)
            if not bool(base.get("ok")):
                return base
            text = str(base.get("content") or "")
            lines = text.splitlines()
            pages = [(1, text)]
            headings = extract_heading_entries_from_pages(pages, max_headings=1000)
            section = _extract_section_from_pdf_pages([(1, text)], headings, heading, limit)
            if bool(section.get("ok")):
                return {
                    "ok": True,
                    "path": str(real_path),
                    "matched_heading": section.get("matched_heading"),
                    "matched_section": section.get("matched_section"),
                    "page_start": 1,
                    "page_end": 1,
                    "content": section.get("content"),
                }
            return {"ok": False, "error": f"Heading not found: {heading}", "path": str(real_path), "line_count": len(lines)}
        except Exception as exc:
            return {"ok": False, "error": f"read_section failed: {exc}"}

    def table_extract(
        self,
        path: str,
        query: str = "",
        page_hint: int = 0,
        max_tables: int = 5,
        max_rows: int = 25,
    ) -> dict[str, Any]:
        try:
            real_path = self._resolve_source_path(path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}

            limit_tables = max(1, min(20, int(max_tables)))
            limit_rows = max(1, min(200, int(max_rows)))
            query_norm = _normalize_search_query(query)

            if _looks_like_pdf_path(real_path):
                candidate_pages: list[int] = []
                if page_hint > 0:
                    candidate_pages.append(int(page_hint))
                if query_norm:
                    search = self._search_contents_in_file_impl(path=path, query=query_norm, max_matches=8, context_chars=120)
                    if bool(search.get("ok")):
                        candidate_pages.extend(
                            int(item.get("page_hint") or 0)
                            for item in (search.get("matches") or [])
                            if int(item.get("page_hint") or 0) > 0
                        )
                page_numbers = sorted(set(page for page in candidate_pages if page > 0)) or None
                tables = extract_pdf_tables_from_path(
                    real_path,
                    page_numbers=page_numbers,
                    max_tables=limit_tables,
                    max_rows=limit_rows,
                )
                if query_norm:
                    query_tokens = [normalize_lookup_text(query_norm)]
                    filtered: list[dict[str, object]] = []
                    for table in tables:
                        rows = [str(row) for row in table.get("rows") or []]
                        joined = normalize_lookup_text("\n".join(rows))
                        if any(token in joined for token in query_tokens):
                            filtered.append(table)
                    tables = filtered
                return {
                    "ok": True,
                    "path": str(real_path),
                    "table_count": len(tables),
                    "tables": tables[:limit_tables],
                }

            if real_path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                try:
                    from openpyxl import load_workbook  # lazy import
                except Exception as exc:
                    return {"ok": False, "error": f"table_extract requires openpyxl: {exc}"}
                wb = load_workbook(filename=str(real_path), read_only=True, data_only=True)
                try:
                    tables: list[dict[str, Any]] = []
                    for sheet in wb.worksheets:
                        rows: list[str] = []
                        for row in sheet.iter_rows(values_only=True):
                            cells = [_xlsx_cell_to_text(cell) for cell in row]
                            if not any(cells):
                                continue
                            row_line = " | ".join(cells)
                            if query_norm and normalize_lookup_text(query_norm) not in normalize_lookup_text(row_line):
                                continue
                            rows.append(row_line)
                            if len(rows) >= limit_rows:
                                break
                        if rows:
                            tables.append({"sheet": sheet.title or "Sheet", "rows": rows})
                        if len(tables) >= limit_tables:
                            break
                    return {"ok": True, "path": str(real_path), "table_count": len(tables), "tables": tables}
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass

            return {"ok": False, "error": "table_extract currently supports PDF/XLSX files only"}
        except Exception as exc:
            return {"ok": False, "error": f"table_extract failed: {exc}"}

    def fact_check_file(
        self,
        path: str,
        claim: str,
        queries: list[str] | None = None,
        max_evidence: int = 6,
    ) -> dict[str, Any]:
        try:
            cleaned_claim = (claim or "").strip()
            if not cleaned_claim:
                return {"ok": False, "error": "claim is empty"}
            query_list = [_normalize_search_query(item) for item in (queries or []) if str(item or "").strip()]
            if not query_list:
                query_list = _derive_fact_check_queries(cleaned_claim)
            search = self._search_contents_in_file_multi_impl(
                path=path,
                queries=query_list,
                per_query_max_matches=max(1, min(6, int(max_evidence))),
                context_chars=220,
            )
            if not bool(search.get("ok")):
                return search

            evidence = list(search.get("matches") or [])[: max(1, min(12, int(max_evidence)))]
            verdict = "insufficient_evidence"
            if evidence:
                verdict = "conflicted" if _is_negative_claim(cleaned_claim) else "supported"
            return {
                "ok": True,
                "path": str(self._resolve_source_path(path)),
                "claim": cleaned_claim,
                "queries_used": query_list,
                "verdict": verdict,
                "evidence_count": len(evidence),
                "evidence": evidence,
                "note": (
                    "This tool checks whether the current extracted file text contains evidence related to the claim. "
                    "A 'supported' result still requires agent judgment about relevance and exact wording."
                ),
            }
        except Exception as exc:
            return {"ok": False, "error": f"fact_check_file failed: {exc}"}

    def search_codebase(
        self,
        query: str,
        root: str = ".",
        max_matches: int = 20,
        file_glob: str = "",
        use_regex: bool = False,
        case_sensitive: bool = False,
    ) -> dict[str, Any]:
        try:
            cleaned_query = str(query or "").strip()
            if not cleaned_query:
                return {"ok": False, "error_kind": "invalid_arguments", "error": "query is empty"}
            real_root = self._resolve_path(root)
            if not real_root.exists():
                return {"ok": False, "error_kind": "path_not_found", "error": f"Path not found: {root}"}
            if not real_root.is_dir():
                return {"ok": False, "error_kind": "not_a_directory", "error": f"Not a directory: {root}"}

            limit = max(1, min(100, int(max_matches)))
            matches: list[dict[str, Any]] = []
            parser_mode = "json"
            if shutil.which("rg"):
                argv_core = ["-n", "--color", "never", "--max-count", str(limit)]
                if not use_regex:
                    argv_core.append("-F")
                if case_sensitive:
                    argv_core.append("-s")
                else:
                    argv_core.append("-i")
                if file_glob.strip():
                    argv_core.extend(["-g", file_glob.strip()])
                argv_tail = [cleaned_query, str(real_root)]

                proc = subprocess.run(["rg", "--json", *argv_core, *argv_tail], capture_output=True, text=True, timeout=20)
                stderr_text = (proc.stderr or "").strip()
                if proc.returncode not in {0, 1} and "--json" in stderr_text.lower():
                    parser_mode = "text_fallback"
                    proc = subprocess.run(
                        ["rg", *argv_core, "--no-heading", *argv_tail],
                        capture_output=True,
                        text=True,
                        timeout=20,
                    )

                if proc.returncode not in {0, 1}:
                    return {"ok": False, "error": (proc.stderr or proc.stdout or "rg failed").strip()}

                if parser_mode == "json":
                    for raw_line in (proc.stdout or "").splitlines():
                        try:
                            event = json.loads(raw_line)
                        except Exception:
                            continue
                        if str(event.get("type") or "") != "match":
                            continue
                        data = event.get("data") if isinstance(event.get("data"), dict) else {}
                        path_block = data.get("path") if isinstance(data.get("path"), dict) else {}
                        lines_block = data.get("lines") if isinstance(data.get("lines"), dict) else {}
                        file_path = str(path_block.get("text") or "").strip()
                        if not file_path:
                            continue
                        resolved_file = Path(file_path)
                        if not resolved_file.is_absolute():
                            resolved_file = (real_root / resolved_file).resolve()
                        try:
                            line_no = int(data.get("line_number") or 0)
                        except Exception:
                            line_no = 0
                        text_line = str(lines_block.get("text") or "").rstrip("\r\n")
                        matches.append(
                            {
                                "path": _display_model_path(resolved_file, project_root=self._current_project_root(), cwd=real_root),
                                "resolved_path": str(resolved_file.resolve()),
                                "line": line_no,
                                "text": text_line.strip(),
                            }
                        )
                        if len(matches) >= limit:
                            break
                else:
                    for line in (proc.stdout or "").splitlines():
                        parts = line.rsplit(":", 2)
                        if len(parts) != 3:
                            continue
                        file_path, line_no_raw, text_line = parts
                        try:
                            line_no = int(line_no_raw)
                        except Exception:
                            line_no = 0
                        resolved_file = Path(file_path)
                        if not resolved_file.is_absolute():
                            resolved_file = (real_root / resolved_file).resolve()
                        matches.append(
                            {
                                "path": _display_model_path(resolved_file, project_root=self._current_project_root(), cwd=real_root),
                                "resolved_path": str(resolved_file.resolve()),
                                "line": line_no,
                                "text": text_line.strip(),
                            }
                        )
                        if len(matches) >= limit:
                            break
            else:
                parser_mode = "python_fallback"
                if use_regex:
                    flags = 0 if case_sensitive else re.IGNORECASE
                    pattern = re.compile(cleaned_query, flags)
                else:
                    needle = cleaned_query if case_sensitive else cleaned_query.lower()
                    pattern = None

                for file_path in real_root.rglob("*"):
                    if not file_path.is_file():
                        continue
                    if file_glob.strip():
                        rel = file_path.relative_to(real_root).as_posix()
                        if not fnmatch.fnmatch(rel, file_glob.strip()):
                            continue
                    try:
                        text = file_path.read_text(encoding="utf-8", errors="ignore")
                    except Exception:
                        continue
                    if "\x00" in text:
                        continue
                    for idx, line in enumerate(text.splitlines(), start=1):
                        hay = line if case_sensitive else line.lower()
                        matched = bool(pattern.search(line)) if pattern is not None else needle in hay
                        if not matched:
                            continue
                        matches.append(
                            {
                                "path": _display_model_path(file_path, project_root=self._current_project_root(), cwd=real_root),
                                "resolved_path": str(file_path.resolve()),
                                "line": idx,
                                "text": line.strip(),
                            }
                        )
                        if len(matches) >= limit:
                            break
                    if len(matches) >= limit:
                        break

            existing_paths = {
                str(item.get("resolved_path") or item.get("path") or "").strip()
                for item in matches
                if str(item.get("resolved_path") or item.get("path") or "").strip()
            }
            path_match_count = 0
            if len(matches) < limit:
                query_for_path = cleaned_query if case_sensitive else cleaned_query.lower()
                query_for_stem = query_for_path.rsplit(".", 1)[0] if "." in query_for_path else query_for_path
                path_pattern: re.Pattern[str] | None = None
                if use_regex:
                    flags = 0 if case_sensitive else re.IGNORECASE
                    try:
                        path_pattern = re.compile(cleaned_query, flags)
                    except Exception:
                        path_pattern = None

                file_candidates: list[Path] = []
                if shutil.which("rg"):
                    proc_files = subprocess.run(
                        ["rg", "--files", str(real_root)],
                        capture_output=True,
                        text=True,
                        timeout=20,
                    )
                    if proc_files.returncode == 0:
                        for raw_line in (proc_files.stdout or "").splitlines():
                            raw_item = str(raw_line or "").strip()
                            if not raw_item:
                                continue
                            candidate = Path(raw_item)
                            if not candidate.is_absolute():
                                candidate = (real_root / raw_item).resolve()
                            if candidate.is_file():
                                file_candidates.append(candidate)
                if not file_candidates:
                    file_candidates = [item for item in real_root.rglob("*") if item.is_file()]

                for candidate in file_candidates:
                    if len(matches) >= limit:
                        break
                    try:
                        rel = candidate.relative_to(real_root).as_posix()
                    except Exception:
                        rel = candidate.as_posix()
                    if file_glob.strip() and not fnmatch.fnmatch(rel, file_glob.strip()):
                        continue

                    rel_text = rel if case_sensitive else rel.lower()
                    stem_text = candidate.stem if case_sensitive else candidate.stem.lower()
                    matched = False
                    if path_pattern is not None:
                        matched = bool(path_pattern.search(rel))
                    else:
                        matched = (
                            (query_for_path in rel_text)
                            or (query_for_path in stem_text)
                            or (query_for_stem and query_for_stem in stem_text)
                        )
                    if not matched:
                        continue

                    candidate_path = str(candidate)
                    if candidate_path in existing_paths:
                        continue
                    existing_paths.add(candidate_path)
                    path_match_count += 1
                    matches.append(
                        {
                            "path": _display_model_path(candidate, project_root=self._current_project_root(), cwd=real_root),
                            "resolved_path": str(candidate.resolve()),
                            "line": 0,
                            "text": "[filename match]",
                            "match_type": "path",
                        }
                    )
            root_payload = _path_payload(real_root, project_root=self._current_project_root(), cwd=Path(self._current_cwd_hint()))
            return {
                "ok": True,
                "root": root_payload["path"],
                "root_ref": root_payload["root_ref"],
                "resolved_root": str(real_root.resolve()),
                "query": cleaned_query,
                "match_count": len(matches),
                "matches": matches,
                "path_match_count": path_match_count,
                "parser_mode": parser_mode,
            }
        except FileNotFoundError:
            return {"ok": False, "error": "rg not found"}
        except Exception as exc:
            return {"ok": False, "error": f"search_codebase failed: {exc}"}

    def _archive_extract_impl(
        self,
        zip_path: str,
        dst_dir: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        max_entries: int = 20000,
        max_total_bytes: int = 524288000,
    ) -> dict[str, Any]:
        try:
            zip_real = self._resolve_source_path(zip_path)
            if not zip_real.exists():
                return {"ok": False, "error": f"Zip path not found: {zip_path}"}
            if not zip_real.is_file():
                return {"ok": False, "error": f"Zip path is not a file: {zip_path}"}

            dst_raw = (dst_dir or "").strip()
            if not dst_raw:
                dst_raw = str(zip_real.with_suffix(""))

            dst_real = self._resolve_path(dst_raw)
            if dst_real.exists() and dst_real.is_file():
                return {"ok": False, "error": f"Destination is a file, not directory: {dst_raw}"}
            if not dst_real.exists():
                if not create_dirs:
                    return {"ok": False, "error": f"Destination directory not found: {dst_real}"}
                dst_real.mkdir(parents=True, exist_ok=True)

            entry_limit = max(1, min(100000, int(max_entries)))
            total_limit = max(1024, min(2147483648, int(max_total_bytes)))

            extracted_files = 0
            skipped_files = 0
            extracted_bytes = 0
            entries: list[dict[str, Any]] = []

            with zipfile.ZipFile(zip_real, "r") as zf:
                infos = zf.infolist()
                if len(infos) > entry_limit:
                    return {
                        "ok": False,
                        "error": f"Zip entries exceed max_entries limit ({len(infos)} > {entry_limit}).",
                    }

                total_uncompressed = sum(int(getattr(i, "file_size", 0) or 0) for i in infos)
                if total_uncompressed > total_limit:
                    return {
                        "ok": False,
                        "error": (
                            f"Zip uncompressed size exceeds max_total_bytes "
                            f"({total_uncompressed} > {total_limit})."
                        ),
                    }

                for info in infos:
                    name = (info.filename or "").replace("\\", "/")
                    if not name:
                        continue

                    rel = Path(name)
                    if rel.is_absolute() or ".." in rel.parts:
                        return {"ok": False, "error": f"Unsafe zip entry path detected: {name}"}

                    target = (dst_real / rel).resolve()
                    if not _is_within(target, dst_real):
                        return {"ok": False, "error": f"Unsafe zip entry path detected: {name}"}

                    if info.is_dir():
                        target.mkdir(parents=True, exist_ok=True)
                        continue

                    if target.exists() and not overwrite:
                        skipped_files += 1
                        continue

                    target.parent.mkdir(parents=True, exist_ok=True)
                    with zf.open(info, "r") as src, open(target, "wb") as out:
                        shutil.copyfileobj(src, out)
                    extracted_files += 1
                    size = int(target.stat().st_size)
                    extracted_bytes += size
                    if len(entries) < 1000:
                        entries.append(
                            {
                                "entry_name": name,
                                "path": str(target),
                                "size": size,
                            }
                        )

            return {
                "ok": True,
                "zip_path": str(zip_real),
                "dst_dir": str(dst_real),
                "files_extracted": extracted_files,
                "files_skipped": skipped_files,
                "bytes_extracted": extracted_bytes,
                "entries": entries,
                "overwrite": bool(overwrite),
            }
        except zipfile.BadZipFile:
            return {"ok": False, "error": f"Invalid zip archive: {zip_path}"}
        except Exception as exc:
            return {"ok": False, "error": f"archive_extract failed: {exc}"}

    def _mail_extract_attachments_impl(
        self,
        msg_path: str,
        dst_dir: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        max_attachments: int = 500,
        max_total_bytes: int = 524288000,
    ) -> dict[str, Any]:
        try:
            msg_real = self._resolve_source_path(msg_path)
            if not msg_real.exists():
                return {"ok": False, "error": f"MSG path not found: {msg_path}"}
            if not msg_real.is_file():
                return {"ok": False, "error": f"MSG path is not a file: {msg_path}"}

            from app.attachments import looks_like_outlook_msg_file  # lazy import

            suffix = msg_real.suffix.lower()
            if suffix != ".msg" and not looks_like_outlook_msg_file(msg_real):
                return {"ok": False, "error": f"Not an Outlook .msg file: {msg_path}"}

            dst_raw = (dst_dir or "").strip()
            if not dst_raw:
                dst_raw = str(msg_real.parent / f"{msg_real.stem}_attachments")

            dst_real = self._resolve_path(dst_raw)
            if dst_real.exists() and dst_real.is_file():
                return {"ok": False, "error": f"Destination is a file, not directory: {dst_raw}"}
            if not dst_real.exists():
                if not create_dirs:
                    return {"ok": False, "error": f"Destination directory not found: {dst_real}"}
                dst_real.mkdir(parents=True, exist_ok=True)

            attachment_limit = max(1, min(5000, int(max_attachments)))
            total_limit = max(1024, min(2147483648, int(max_total_bytes)))

            try:
                import extract_msg  # lazy import
            except Exception as exc:
                return {
                    "ok": False,
                    "error": (
                        "解析 .msg 附件需要依赖 extract-msg。请执行 "
                        "`pip install -r requirements.txt` 后重试。"
                    ),
                    "detail": str(exc),
                }

            msg = extract_msg.openMsg(str(msg_real), strict=False, delayAttachments=False)
            try:
                attachments = list(getattr(msg, "attachments", []) or [])
                if len(attachments) > attachment_limit:
                    return {
                        "ok": False,
                        "error": (
                            f"MSG attachments exceed max_attachments limit "
                            f"({len(attachments)} > {attachment_limit})."
                        ),
                        "msg_path": str(msg_real),
                        "dst_dir": str(dst_real),
                    }

                entries: list[dict[str, Any]] = []
                files_saved = 0
                files_skipped = 0
                bytes_extracted = 0

                for idx, att in enumerate(attachments, start=1):
                    raw_name = (
                        (getattr(att, "longFilename", None) or "")
                        or (getattr(att, "filename", None) or "")
                        or (getattr(att, "name", None) or "")
                        or f"attachment_{idx}"
                    )
                    safe_name = _safe_filename(str(raw_name or ""))
                    if safe_name == "download.bin":
                        safe_name = f"attachment_{idx}.bin"

                    att_type = str(getattr(att, "type", "") or "").upper()
                    if "MSG" in att_type and not safe_name.lower().endswith(".msg"):
                        safe_name = f"{safe_name}.msg"

                    target = (dst_real / safe_name).resolve()
                    if not _is_within(target, dst_real):
                        return {
                            "ok": False,
                            "error": f"Unsafe attachment path detected: {safe_name}",
                            "msg_path": str(msg_real),
                            "dst_dir": str(dst_real),
                        }

                    if target.exists() and not overwrite:
                        files_skipped += 1
                        entries.append(
                            {
                                "index": idx,
                                "name": safe_name,
                                "status": "skipped_exists",
                                "path": str(target),
                            }
                        )
                        continue

                    try:
                        save_result = att.save(
                            customPath=str(dst_real),
                            customFilename=safe_name,
                            overwriteExisting=bool(overwrite),
                            extractEmbedded=True,
                            skipEmbedded=False,
                        )
                    except Exception as exc:
                        entries.append(
                            {
                                "index": idx,
                                "name": safe_name,
                                "status": "error",
                                "error": str(exc),
                            }
                        )
                        continue

                    saved_paths: list[Path] = []
                    if (
                        isinstance(save_result, tuple)
                        and len(save_result) >= 2
                        and save_result[1] is not None
                    ):
                        payload = save_result[1]
                        if isinstance(payload, str):
                            saved_paths.append(Path(payload).resolve())
                        elif isinstance(payload, list):
                            for item in payload:
                                if isinstance(item, str):
                                    saved_paths.append(Path(item).resolve())

                    if not saved_paths and target.exists():
                        saved_paths.append(target)

                    saved_payload: list[dict[str, Any]] = []
                    for path_obj in saved_paths:
                        if not path_obj.exists():
                            continue
                        if not _is_within(path_obj, dst_real):
                            continue
                        size = path_obj.stat().st_size if path_obj.is_file() else None
                        if isinstance(size, int):
                            bytes_extracted += size
                        saved_payload.append(
                            {
                                "path": str(path_obj),
                                "is_dir": path_obj.is_dir(),
                                "bytes": size,
                            }
                        )

                    if bytes_extracted > total_limit:
                        return {
                            "ok": False,
                            "error": (
                                f"Extracted bytes exceed max_total_bytes limit "
                                f"({bytes_extracted} > {total_limit})."
                            ),
                            "msg_path": str(msg_real),
                            "dst_dir": str(dst_real),
                            "attachments_total": len(attachments),
                            "files_saved": files_saved,
                            "files_skipped": files_skipped,
                            "bytes_extracted": bytes_extracted,
                            "entries": entries,
                        }

                    if saved_payload:
                        files_saved += 1
                        entries.append(
                            {
                                "index": idx,
                                "name": safe_name,
                                "status": "saved",
                                "saved": saved_payload,
                            }
                        )
                    else:
                        entries.append(
                            {
                                "index": idx,
                                "name": safe_name,
                                "status": "no_output",
                            }
                        )

                return {
                    "ok": True,
                    "msg_path": str(msg_real),
                    "dst_dir": str(dst_real),
                    "attachments_total": len(attachments),
                    "files_saved": files_saved,
                    "files_skipped": files_skipped,
                    "bytes_extracted": bytes_extracted,
                    "entries": entries,
                    "overwrite": bool(overwrite),
                }
            finally:
                close = getattr(msg, "close", None)
                if callable(close):
                    try:
                        close()
                    except Exception:
                        pass
        except Exception as exc:
            return {"ok": False, "error": f"mail_extract_attachments failed: {exc}"}

    def _domain_allowed(self, host: str) -> bool:
        if self.config.web_allow_all_domains:
            return True

        host = host.lower().strip(".")
        for allowed in self.config.web_allowed_domains:
            d = allowed.lower().strip(".")
            if host == d or host.endswith("." + d):
                return True
        return False

    def _web_search_impl(self, query: str, max_results: int = 5, timeout_sec: int = 12) -> dict[str, Any]:
        q = (query or "").strip()
        if not q:
            return {"ok": False, "error": "query cannot be empty"}

        timeout_val = max(3, min(30, timeout_sec))
        limit = max(1, min(20, int(max_results)))
        cache_key = {"query": q, "max_results": limit, "algo_version": 4}
        cached = self._load_web_cache("web_search", cache_key, max_age_sec=900)
        if cached:
            return {**cached, "cached": True}
        read_limit = min(500000, max(20000, self.config.web_fetch_max_chars))
        ddg_allowed = self._domain_allowed("duckduckgo.com")
        if not ddg_allowed:
            return {
                "ok": False,
                "error": (
                    "Domain not allowed for search engine. "
                    f"Allowed: {', '.join(self.config.web_allowed_domains)}"
                ),
            }

        search_url = "https://duckduckgo.com/html/?q=" + urllib.parse.quote_plus(q)
        lite_url = "https://lite.duckduckgo.com/lite/?q=" + urllib.parse.quote_plus(q)

        if self.config.web_skip_tls_verify:
            ssl_context = ssl._create_unverified_context()
        elif self.config.web_ca_cert_path:
            try:
                ssl_context = ssl.create_default_context(cafile=self.config.web_ca_cert_path)
            except Exception as exc:
                return {
                    "ok": False,
                    "error": f"Invalid web CA cert path: {self.config.web_ca_cert_path} ({exc})",
                }
        else:
            ssl_context = ssl.create_default_context()

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.5",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }

        tls_warning: str | None = None
        active_context = ssl_context

        def _open(current_context: ssl.SSLContext | None, target_url: str):
            req = urllib.request.Request(
                url=target_url,
                headers=headers,
                method="GET",
            )
            opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=current_context))
            return opener.open(req, timeout=timeout_val)

        def _fetch_page(target_url: str, current_context: ssl.SSLContext | None) -> tuple[int, str, str, bool]:
            with _open(current_context, target_url) as resp:
                status = getattr(resp, "status", None) or 200
                content_type = (resp.headers.get("Content-Type") or "").lower()
                raw = resp.read(read_limit + 1)
                truncated = len(raw) > read_limit
                raw = raw[:read_limit]
                text = raw.decode("utf-8", errors="ignore")
                return status, content_type, text, truncated

        def _fetch_page_with_retry(target_url: str) -> tuple[int, str, str, bool]:
            nonlocal active_context, tls_warning
            try:
                return _fetch_page(target_url, active_context)
            except Exception as first_exc:
                if not self.config.web_skip_tls_verify and _is_cert_verify_error(first_exc):
                    tls_warning = "TLS verify failed; web_search auto-retried with verify disabled."
                    active_context = ssl._create_unverified_context()
                    return _fetch_page(target_url, active_context)
                raise

        try:
            results: list[dict[str, str]] = []
            source = "unknown"
            status = 200
            content_type = "text/html"
            truncated = False
            warning_parts: list[str] = []
            seen_result_keys: set[str] = set()

            def _append_results(items: list[dict[str, str]], source_name: str) -> int:
                added = 0
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    title = str(item.get("title", "")).strip()
                    url = str(item.get("url", "")).strip()
                    if not title and not url:
                        continue
                    key = f"{title}|{url}".lower()
                    if key in seen_result_keys:
                        continue
                    seen_result_keys.add(key)

                    row = dict(item)
                    row.setdefault("source", source_name)
                    results.append(row)
                    added += 1
                    if len(results) >= limit:
                        break
                return added

            if not results:
                try:
                    status, content_type, text, truncated = _fetch_page_with_retry(search_url)
                    ddg_results = _extract_ddg_results(text, max_results=limit)
                    if _append_results(ddg_results, "duckduckgo_html") > 0:
                        source = "duckduckgo_html"
                except Exception as exc:
                    warning_parts.append(f"DuckDuckGo HTML 搜索失败: {exc}")

            if not results:
                try:
                    status, content_type, text, truncated = _fetch_page_with_retry(lite_url)
                    ddg_results = _extract_ddg_results(text, max_results=limit)
                    if _append_results(ddg_results, "duckduckgo_lite") > 0:
                        source = "duckduckgo_lite"
                except Exception as exc:
                    warning_parts.append(f"DuckDuckGo Lite 搜索失败: {exc}")

            if not results:
                warning_parts.append("搜索结果页解析为空，可能被网关改写或反爬。")

            if tls_warning:
                warning_parts.insert(0, tls_warning)

            warning = " ".join(part.strip() for part in warning_parts if part and part.strip()) or None
            if source == "unknown":
                source = "none"

            normalized_results: list[dict[str, Any]] = []
            for item in results:
                if not isinstance(item, dict):
                    continue
                row = dict(item)
                url = str(row.get("url") or "").strip()
                try:
                    row["domain"] = (urllib.parse.urlsplit(url).hostname or "").strip().lower()
                except Exception:
                    row["domain"] = ""
                row["score"] = round(_score_web_result(q, row), 3)
                normalized_results.append(row)
            normalized_results.sort(
                key=lambda item: (
                    float(item.get("score") or 0.0),
                    bool(item.get("published_at")),
                    len(str(item.get("title") or "")),
                ),
                reverse=True,
            )

            payload = {
                "ok": True,
                "query": q,
                "engine": source,
                "status": status,
                "content_type": content_type,
                "count": len(normalized_results),
                "results": normalized_results,
                "truncated": truncated,
                "warning": warning,
                "cached": False,
            }
            self._save_web_cache("web_search", cache_key, payload)
            return payload
        except urllib.error.HTTPError as exc:
            body = exc.read(4000).decode("utf-8", errors="ignore") if hasattr(exc, "read") else ""
            return {"ok": False, "error": f"HTTP {exc.code}: {exc.reason}", "body_preview": body}
        except Exception as exc:
            return {"ok": False, "error": f"web_search failed: {exc}"}

    def _web_download_impl(
        self,
        url: str,
        dst_path: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        timeout_sec: int = 20,
        max_bytes: int = 52_428_800,
    ) -> dict[str, Any]:
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme not in {"http", "https"}:
            return {"ok": False, "error": "Only http/https URLs are supported"}
        if not parsed.netloc:
            return {"ok": False, "error": "Invalid URL"}

        try:
            request_url = _normalize_url_for_request(url)
        except Exception as exc:
            return {"ok": False, "error": f"Invalid URL: {exc}"}

        host = parsed.hostname or ""
        if not self._domain_allowed(host):
            return {
                "ok": False,
                "error": f"Domain not allowed: {host}. Allowed: {', '.join(self.config.web_allowed_domains)}",
            }

        timeout_val = max(3, min(120, int(timeout_sec)))
        byte_limit = max(1024, min(209_715_200, int(max_bytes)))

        ssl_context: ssl.SSLContext | None = None
        if parsed.scheme == "https":
            if self.config.web_skip_tls_verify:
                ssl_context = ssl._create_unverified_context()
            elif self.config.web_ca_cert_path:
                try:
                    ssl_context = ssl.create_default_context(cafile=self.config.web_ca_cert_path)
                except Exception as exc:
                    return {
                        "ok": False,
                        "error": f"Invalid web CA cert path: {self.config.web_ca_cert_path} ({exc})",
                    }
            else:
                ssl_context = ssl.create_default_context()

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "*/*",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }

        tls_warning: str | None = None

        def _open(current_context: ssl.SSLContext | None):
            req = urllib.request.Request(
                url=request_url,
                headers=headers,
                method="GET",
            )
            opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=current_context))
            return opener.open(req, timeout=timeout_val)

        try:
            try:
                resp_cm = _open(ssl_context)
            except Exception as first_exc:
                if not self.config.web_skip_tls_verify and _is_cert_verify_error(first_exc):
                    tls_warning = "TLS verify failed; web_download auto-retried with verify disabled."
                    resp_cm = _open(ssl._create_unverified_context())
                else:
                    raise

            with resp_cm as resp:
                status = getattr(resp, "status", None) or 200
                content_type = (resp.headers.get("Content-Type") or "").lower()
                content_disposition = resp.headers.get("Content-Disposition") or ""
                filename = _guess_filename_from_response(url=url, content_type=content_type, content_disposition=content_disposition)

                raw = resp.read(byte_limit + 1)
                truncated = len(raw) > byte_limit
                if truncated:
                    return {
                        "ok": False,
                        "error": f"Remote file exceeds max_bytes limit ({byte_limit} bytes).",
                        "status": status,
                        "url": url,
                        "content_type": content_type,
                        "filename": filename,
                    }

                target_raw = (dst_path or "").strip()
                if not target_raw:
                    target_raw = str(Path("downloads") / filename)

                target_path = self._resolve_path(target_raw)
                if target_path.exists() and target_path.is_dir():
                    return {"ok": False, "error": f"Destination is a directory: {target_raw}"}
                if target_path.exists() and not overwrite:
                    return {"ok": False, "error": f"Destination exists and overwrite=false: {target_raw}"}
                if not target_path.parent.exists():
                    if not create_dirs:
                        return {"ok": False, "error": f"Destination parent not found: {target_path.parent}"}
                    target_path.parent.mkdir(parents=True, exist_ok=True)

                action = "overwrite" if target_path.exists() else "create"
                target_path.write_bytes(raw)
                return {
                    "ok": True,
                    "url": url,
                    "status": status,
                    "content_type": content_type,
                    "path": str(target_path),
                    "bytes": len(raw),
                    "filename": filename,
                    "action": action,
                    "warning": tls_warning,
                }
        except urllib.error.HTTPError as exc:
            body = exc.read(4000).decode("utf-8", errors="ignore") if hasattr(exc, "read") else ""
            return {"ok": False, "error": f"HTTP {exc.code}: {exc.reason}", "body_preview": body}
        except Exception as exc:
            return {"ok": False, "error": f"web_download failed: {exc}"}

    def _web_fetch_impl(self, url: str, max_chars: int = 120000, timeout_sec: int = 12) -> dict[str, Any]:
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme not in {"http", "https"}:
            return {"ok": False, "error": "Only http/https URLs are supported"}
        if not parsed.netloc:
            return {"ok": False, "error": "Invalid URL"}

        try:
            request_url = _normalize_url_for_request(url)
        except Exception as exc:
            return {"ok": False, "error": f"Invalid URL: {exc}"}

        host = parsed.hostname or ""
        if not self._domain_allowed(host):
            return {
                "ok": False,
                "error": f"Domain not allowed: {host}. Allowed: {', '.join(self.config.web_allowed_domains)}",
            }

        timeout_val = max(3, min(30, timeout_sec))
        limit = max(512, min(500000, max_chars, self.config.web_fetch_max_chars))
        cache_key = {"url": request_url, "max_chars": limit}
        cached = self._load_web_cache("web_fetch", cache_key, max_age_sec=900)
        if cached:
            return {**cached, "cached": True}
        ssl_context: ssl.SSLContext | None = None
        if parsed.scheme == "https":
            if self.config.web_skip_tls_verify:
                ssl_context = ssl._create_unverified_context()
            elif self.config.web_ca_cert_path:
                try:
                    ssl_context = ssl.create_default_context(cafile=self.config.web_ca_cert_path)
                except Exception as exc:
                    return {
                        "ok": False,
                        "error": f"Invalid web CA cert path: {self.config.web_ca_cert_path} ({exc})",
                    }
            else:
                ssl_context = ssl.create_default_context()

        default_headers = {
            # Use a browser-like UA to reduce bot-block false positives.
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/json,text/plain,application/xml;q=0.9,*/*;q=0.5",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }

        tls_warning: str | None = None

        def _open(current_context: ssl.SSLContext | None, target_url: str):
            req = urllib.request.Request(
                url=target_url,
                headers=default_headers,
                method="GET",
            )
            opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=current_context))
            return opener.open(req, timeout=timeout_val)

        try:
            try:
                resp_cm = _open(ssl_context, request_url)
            except Exception as first_exc:
                # Pragmatic fallback for enterprise TLS chains:
                # if verification fails and user did not explicitly disable it,
                # retry once with verification off for web_fetch only.
                if not self.config.web_skip_tls_verify and _is_cert_verify_error(first_exc):
                    tls_warning = "TLS verify failed; web_fetch auto-retried with verify disabled."
                    resp_cm = _open(ssl._create_unverified_context(), request_url)
                else:
                    raise

            with resp_cm as resp:
                status = getattr(resp, "status", None) or 200
                content_type = (resp.headers.get("Content-Type") or "").lower()
                content_disposition = (resp.headers.get("Content-Disposition") or "").lower()
                pdf_like = (
                    "application/pdf" in content_type
                    or parsed.path.lower().endswith(".pdf")
                    or ".pdf" in content_disposition
                )
                pdf_byte_limit = min(20_000_000, max(1_000_000, self.config.web_fetch_max_chars * 40))
                raw_limit = pdf_byte_limit if pdf_like else limit

                raw = resp.read(raw_limit + 1)
                truncated = len(raw) > raw_limit
                raw = raw[:raw_limit]

                if pdf_like:
                    try:
                        pdf_text = _extract_pdf_text_from_bytes(raw, max_chars=limit)
                        warning = tls_warning
                        if truncated:
                            warning = (
                                f"{warning} PDF 文件较大，已按 {raw_limit} bytes 截断读取。"
                                if warning
                                else f"PDF 文件较大，已按 {raw_limit} bytes 截断读取。"
                            )
                        if not pdf_text.strip():
                            warning = (
                                f"{warning} PDF 可读文本为空（可能是扫描件图片）。"
                                if warning
                                else "PDF 可读文本为空（可能是扫描件图片）。"
                            )
                        payload = {
                            "ok": True,
                            "url": url,
                            "status": status,
                            "content_type": content_type,
                            "domain": host,
                            "binary": False,
                            "truncated": truncated,
                            "content": pdf_text,
                            "length": len(pdf_text),
                            "source_format": "pdf_text_extracted",
                            "warning": warning,
                            "cached": False,
                        }
                        self._save_web_cache("web_fetch", cache_key, payload)
                        return payload
                    except Exception as pdf_exc:
                        warning = (
                            f"{tls_warning} PDF 文本提取失败: {pdf_exc}"
                            if tls_warning
                            else f"PDF 文本提取失败: {pdf_exc}"
                        )
                        return {
                            "ok": True,
                            "url": url,
                            "status": status,
                            "content_type": content_type,
                            "binary": True,
                            "size_preview_bytes": len(raw),
                            "truncated": truncated,
                            "warning": warning,
                        }

                if any(x in content_type for x in ["application/octet-stream", "image/", "audio/", "video/"]):
                    return {
                        "ok": True,
                        "url": url,
                        "status": status,
                        "content_type": content_type,
                        "binary": True,
                        "size_preview_bytes": len(raw),
                        "truncated": truncated,
                        "warning": tls_warning,
                    }

                text = raw.decode("utf-8", errors="ignore")
                if _looks_like_html(content_type, text):
                    metadata = _extract_html_metadata(text, base_url=url)
                    extracted = _extract_html_text(text, max_chars=limit)
                    warning = None
                    if len(extracted.strip()) < 80:
                        warning = (
                            "页面正文较少，可能是 JS 动态渲染或反爬页面。"
                            "建议改用该站点公开 API，或换一个可直读正文的页面。"
                        )
                    if _looks_like_script_payload(extracted):
                        script_warning = (
                            "抓取内容疑似脚本/反爬响应，而非正文页面。"
                            "请不要据此下结论，建议改用官方 API 或可直读页面。"
                        )
                        warning = f"{script_warning} {warning}" if warning else script_warning

                        # Search-engine anti-bot fallback: try a text-friendly results page.
                        search_query = _extract_search_query(url)
                        if search_query and self._domain_allowed("duckduckgo.com"):
                            fallback_url = (
                                "https://duckduckgo.com/html/?q="
                                + urllib.parse.quote_plus(search_query)
                            )
                            try:
                                with _open(ssl_context, fallback_url) as fb_resp:
                                    fb_status = getattr(fb_resp, "status", None) or 200
                                    fb_ct = (fb_resp.headers.get("Content-Type") or "").lower()
                                    fb_raw = fb_resp.read(limit + 1)
                                    fb_truncated = len(fb_raw) > limit
                                    fb_raw = fb_raw[:limit]
                                    fb_text = fb_raw.decode("utf-8", errors="ignore")
                                    fb_extracted = _extract_html_text(fb_text, max_chars=limit)

                                if fb_extracted.strip() and not _looks_like_script_payload(fb_extracted):
                                    if tls_warning:
                                        warning = f"{tls_warning} {warning}" if warning else tls_warning
                                    fallback_warning = (
                                        f"{warning} 已自动回退到 DuckDuckGo HTML 结果页（query={search_query}）。"
                                        if warning
                                        else f"已自动回退到 DuckDuckGo HTML 结果页（query={search_query}）。"
                                    )
                                    fallback_payload = {
                                        "ok": True,
                                        "url": url,
                                        "status": fb_status,
                                        "content_type": fb_ct,
                                        "domain": host,
                                        "binary": False,
                                        "truncated": fb_truncated,
                                        "content": fb_extracted,
                                        "length": len(fb_extracted),
                                        "source_format": "search_fallback_duckduckgo_html",
                                        "warning": fallback_warning,
                                        "title": metadata.get("title") or "",
                                        "published_at": metadata.get("published_at") or "",
                                        "canonical_url": metadata.get("canonical_url") or "",
                                        "cached": False,
                                    }
                                    self._save_web_cache("web_fetch", cache_key, fallback_payload)
                                    return fallback_payload
                            except Exception as fb_exc:
                                warning = (
                                    f"{warning} DuckDuckGo 回退失败: {fb_exc}"
                                    if warning
                                    else f"DuckDuckGo 回退失败: {fb_exc}"
                                )

                        # Avoid passing noisy script payload to the model.
                        extracted = (
                            "[抓取到脚本/反爬页面，正文不可用。"
                            "请改用目标站点公开 API、可直读正文 URL，或非搜索结果页链接。]"
                        )
                    if tls_warning:
                        warning = f"{tls_warning} {warning}" if warning else tls_warning
                    payload = {
                        "ok": True,
                        "url": url,
                        "status": status,
                        "content_type": content_type,
                        "domain": host,
                        "binary": False,
                        "truncated": truncated,
                        "content": extracted,
                        "length": len(extracted),
                        "source_format": "html_text_extracted",
                        "warning": warning,
                        "title": metadata.get("title") or "",
                        "published_at": metadata.get("published_at") or "",
                        "canonical_url": metadata.get("canonical_url") or "",
                        "cached": False,
                    }
                    self._save_web_cache("web_fetch", cache_key, payload)
                    return payload

                payload = {
                    "ok": True,
                    "url": url,
                    "status": status,
                    "content_type": content_type,
                    "domain": host,
                    "binary": False,
                    "truncated": truncated,
                    "content": text,
                    "length": len(text),
                    "warning": tls_warning,
                    "cached": False,
                }
                self._save_web_cache("web_fetch", cache_key, payload)
                return payload
        except urllib.error.HTTPError as exc:
            body = exc.read(4000).decode("utf-8", errors="ignore") if hasattr(exc, "read") else ""
            return {"ok": False, "error": f"HTTP {exc.code}: {exc.reason}", "body_preview": body}
        except Exception as exc:
            return {"ok": False, "error": f"web_fetch failed: {exc}"}


def parse_json_arguments(raw: str | dict[str, Any] | None) -> dict[str, Any]:
    if raw is None:
        return {}
    if isinstance(raw, dict):
        return raw
    if not raw.strip():
        return {}
    try:
        return json.loads(raw)
    except Exception:
        return {}
